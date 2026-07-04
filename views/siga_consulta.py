"""views/siga_consulta.py — Consulta fiscal tipo SIGA/DTE com exportação Excel e PDF"""
import io
import calendar
import datetime

import streamlit as st

from auth.security import current_user
from db.database import listar_certificados, listar_resultados_por_periodo, reparar_dados_nfe
from views import nav


_MESES = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def _fmt_cnpj(c: str) -> str:
    c = (c or "").strip().zfill(14)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


def _fmt_brl(v) -> str:
    try:
        return f"R$ {float(v):_.2f}".replace("_", ".").replace(".", ",", 1)
    except Exception:
        return "R$ 0,00"


# ── Excel ────────────────────────────────────────────────────────────────────

def _gerar_excel(dados_rec, dados_emit, dados_nfce, cnpj, razao, periodo):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="EEF4FB")
    tot_fill = PatternFill("solid", fgColor="DBEAFE")
    bdr = Border(bottom=Side(style="thin", color="CBD5E1"))

    sheets = [
        ("NF-e Recebidas",  dados_rec,  "recebida"),
        ("NF-e Emitidas",   dados_emit, "emitida"),
        ("NFC-e Emitidas",  dados_nfce, "nfce"),
    ]

    for sheet_name, docs, tipo in sheets:
        ws = wb.create_sheet(sheet_name)

        ws.merge_cells("A1:H1")
        ws["A1"] = f"FISCAL HUB — {sheet_name} | {razao} ({_fmt_cnpj(cnpj)}) | {periodo}"
        ws["A1"].font = Font(bold=True, size=12, color="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        if tipo == "recebida":
            cols = ["Nº", "Série", "Data Emissão", "CNPJ Emitente", "Emitente",
                    "Nat. Operação", "Chave de Acesso", "Valor Total"]
        else:
            cols = ["Nº", "Série", "Data Emissão", "CNPJ Destinatário", "Destinatário",
                    "Nat. Operação", "Chave de Acesso", "Valor Total"]

        for ci, h in enumerate(cols, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 20

        total = 0.0
        for ri, doc in enumerate(docs, 3):
            fill = alt_fill if ri % 2 == 0 else None
            cnpj_p = _fmt_cnpj(doc.get("cnpj_emit", "") if tipo == "recebida" else doc.get("cnpj_dest", ""))
            nome_p = doc.get("nome_emit", "") if tipo == "recebida" else doc.get("nome_dest", "")
            val = float(doc.get("valor_total", 0))
            total += val
            row_vals = [
                doc.get("numero", ""), doc.get("serie", ""),
                (doc.get("data_emissao", "") or "")[:10],
                cnpj_p, nome_p,
                doc.get("nat_operacao", ""),
                doc.get("chave", ""),
                val,
            ]
            for ci, v in enumerate(row_vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                if fill:
                    cell.fill = fill
                cell.border = bdr
                if ci == 8:
                    cell.number_format = '"R$" #,##0.00'

        last = len(docs) + 3
        lbl = ws.cell(row=last, column=7, value="TOTAL")
        lbl.font = Font(bold=True)
        lbl.fill = tot_fill
        tcell = ws.cell(row=last, column=8, value=total)
        tcell.font = Font(bold=True, color="1E3A5F")
        tcell.fill = tot_fill
        tcell.number_format = '"R$" #,##0.00'

        for ci, w in enumerate([8, 7, 14, 19, 36, 26, 50, 16], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF DTE ──────────────────────────────────────────────────────────────────

def _gerar_pdf(dados_rec, dados_emit, dados_nfce, cnpj, razao, periodo):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    BLUE  = colors.HexColor("#1E3A5F")
    BLUE2 = colors.HexColor("#2563EB")
    GRAY  = colors.HexColor("#F0F4F8")
    LGRAY = colors.HexColor("#F8FAFC")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=12 * mm, bottomMargin=12 * mm,
        leftMargin=12 * mm, rightMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Normal"],
                              fontSize=14, textColor=BLUE,
                              alignment=TA_CENTER, fontName="Helvetica-Bold",
                              spaceAfter=4)
    sub_s   = ParagraphStyle("S", parent=styles["Normal"],
                              fontSize=9, textColor=colors.HexColor("#64748B"),
                              alignment=TA_CENTER, spaceAfter=4)
    sec_s   = ParagraphStyle("Sec", parent=styles["Normal"],
                              fontSize=10, textColor=BLUE,
                              fontName="Helvetica-Bold",
                              spaceBefore=10, spaceAfter=4)

    content = []

    content.append(Paragraph("DEMONSTRATIVO DE TRIBUTAÇÃO ELETRÔNICA — DTE", title_s))
    content.append(Paragraph(
        f"Empresa: <b>{razao}</b> &nbsp;|&nbsp; CNPJ: <b>{_fmt_cnpj(cnpj)}</b>"
        f" &nbsp;|&nbsp; Competência: <b>{periodo}</b>", sub_s))
    content.append(Paragraph(
        f"Emitido em {datetime.date.today().strftime('%d/%m/%Y')} via Fiscal Hub", sub_s))
    content.append(HRFlowable(width="100%", thickness=1, color=BLUE, spaceAfter=8))

    # Resumo
    tot_r = sum(float(d.get("valor_total", 0)) for d in dados_rec)
    tot_e = sum(float(d.get("valor_total", 0)) for d in dados_emit)
    tot_n = sum(float(d.get("valor_total", 0)) for d in dados_nfce)
    resumo = [
        ["Tipo de Documento",  "Qtd.", "Valor Total"],
        ["NF-e Recebidas",     str(len(dados_rec)),  _fmt_brl(tot_r)],
        ["NF-e Emitidas",      str(len(dados_emit)), _fmt_brl(tot_e)],
        ["NFC-e Emitidas",     str(len(dados_nfce)), _fmt_brl(tot_n)],
        ["TOTAL GERAL",
         str(len(dados_rec) + len(dados_emit) + len(dados_nfce)),
         _fmt_brl(tot_r + tot_e + tot_n)],
    ]
    resumo_style = TableStyle([
        ("BACKGROUND",   (0, 0), (-1,  0), BLUE),
        ("TEXTCOLOR",    (0, 0), (-1,  0), colors.white),
        ("FONTNAME",     (0, 0), (-1,  0), "Helvetica-Bold"),
        ("BACKGROUND",   (0, -1), (-1, -1), BLUE2),
        ("TEXTCOLOR",    (0, -1), (-1, -1), colors.white),
        ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, GRAY]),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ])
    rt = Table(resumo, colWidths=[140, 60, 120])
    rt.setStyle(resumo_style)
    content.append(rt)

    # Seções de detalhe
    for titulo, docs, tipo in [
        ("NF-e RECEBIDAS",  dados_rec,  "recebida"),
        ("NF-e EMITIDAS",   dados_emit, "emitida"),
        ("NFC-e EMITIDAS",  dados_nfce, "nfce"),
    ]:
        if not docs:
            continue

        content.append(Spacer(1, 6))
        content.append(Paragraph(titulo, sec_s))

        if tipo == "recebida":
            headers = ["Nº", "Sér.", "Data", "CNPJ Emitente", "Emitente", "Nat. Operação", "Valor"]
        else:
            headers = ["Nº", "Sér.", "Data", "CNPJ Dest.", "Destinatário", "Nat. Operação", "Valor"]

        col_ws = [40, 28, 52, 82, 150, 90, 62]

        rows = [headers]
        for item in docs:
            cnpj_p = _fmt_cnpj(item.get("cnpj_emit", "") if tipo == "recebida" else item.get("cnpj_dest", ""))
            nome_p = (item.get("nome_emit", "") if tipo == "recebida" else item.get("nome_dest", ""))[:35]
            nat    = (item.get("nat_operacao", "") or "")[:22]
            rows.append([
                item.get("numero", ""), item.get("serie", ""),
                (item.get("data_emissao", "") or "")[:10],
                cnpj_p, nome_p, nat,
                _fmt_brl(item.get("valor_total", 0)),
            ])

        subtotal = sum(float(d.get("valor_total", 0)) for d in docs)
        rows.append(["", "", "", "", "", "SUBTOTAL", _fmt_brl(subtotal)])

        det_style = TableStyle([
            ("BACKGROUND",    (0, 0), (-1,  0), BLUE),
            ("TEXTCOLOR",     (0, 0), (-1,  0), colors.white),
            ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
            ("BACKGROUND",    (0, -1), (-1, -1), GRAY),
            ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, LGRAY]),
            ("FONTSIZE",      (0, 0), (-1, -1), 7),
            ("ALIGN",         (6, 0), (6, -1),  "RIGHT"),
            ("ALIGN",         (5, 0), (5, -1),  "CENTER"),
            ("ALIGN",         (0, 0), (4, -1),  "CENTER"),
            ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 3),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
        ])
        dt = Table(rows, colWidths=col_ws, repeatRows=1)
        dt.setStyle(det_style)
        content.append(dt)

    doc.build(content)
    buf.seek(0)
    return buf.read()


# ── Tabela de dados ───────────────────────────────────────────────────────────

def _render_tabela(docs, tipo):
    if not docs:
        st.info("Nenhum documento encontrado para o período selecionado.")
        return

    import pandas as pd

    rows = []
    for d in docs:
        if tipo == "recebida":
            cnpj_p = _fmt_cnpj(d.get("cnpj_emit", ""))
            nome_p = d.get("nome_emit", "")
        else:
            cnpj_p = _fmt_cnpj(d.get("cnpj_dest", ""))
            nome_p = d.get("nome_dest", "")

        rows.append({
            "Nº":           d.get("numero", ""),
            "Série":        d.get("serie", ""),
            "Data Emissão": (d.get("data_emissao", "") or "")[:10],
            "CNPJ":         cnpj_p,
            "Empresa":      nome_p,
            "Nat. Operação":d.get("nat_operacao", ""),
            "Valor Total":  float(d.get("valor_total", 0)),
            "Chave":        d.get("chave", ""),
        })

    df = pd.DataFrame(rows)
    total = df["Valor Total"].sum()

    st.dataframe(
        df, use_container_width=True, hide_index=True,
        column_config={
            "Valor Total": st.column_config.NumberColumn("Valor Total", format="R$ %.2f"),
            "Chave": st.column_config.TextColumn("Chave de Acesso", width="medium"),
        },
    )
    st.caption(
        f"**{len(docs)} documentos** &nbsp;|&nbsp; Total: **{_fmt_brl(total)}**",
        unsafe_allow_html=True,
    )


# ── Render ────────────────────────────────────────────────────────────────────

def render():
    user = current_user()
    nav.render("siga_consulta")

    st.markdown("## Consulta Fiscal — DTE / SIGA")
    st.caption(
        "Visualize NF-e e NFC-e sincronizadas no acervo local. "
        "Exporte em Excel ou PDF no formato DTE da SEFAZ CE."
    )

    certs = listar_certificados(user["username"])
    if not certs:
        st.warning(
            "Nenhum certificado digital cadastrado. "
            "Acesse a página **Certificados** para adicionar."
        )
        return

    opcoes = {
        f"{c['razao_social']} ({_fmt_cnpj(c['cnpj'])})": c["cnpj"]
        for c in certs
    }

    hoje = datetime.date.today()

    c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
    with c1:
        empresa_label = st.selectbox("Empresa", list(opcoes.keys()), key="siga_empresa")
    with c2:
        mes_sel = st.selectbox(
            "Mês", range(1, 13), index=hoje.month - 1,
            format_func=lambda m: _MESES[m - 1], key="siga_mes",
        )
    with c3:
        anos = list(range(hoje.year, hoje.year - 6, -1))
        ano_sel = st.selectbox("Ano", anos, key="siga_ano")
    with c4:
        st.markdown("<br>", unsafe_allow_html=True)
        consultar = st.button("Consultar", type="primary",
                              use_container_width=True, key="siga_btn")

    cnpj_sel  = opcoes[empresa_label]
    razao_sel = next(c["razao_social"] for c in certs if c["cnpj"] == cnpj_sel)

    with st.expander("Ferramentas de manutenção"):
        st.caption(
            "Use **Reparar dados** se notas aparecem com valor zerado ou NFC-e não listam corretamente. "
            "O sistema re-processa o XML de cada nota já salva no acervo local para a empresa selecionada."
        )
        if st.button("Reparar dados do acervo", key="btn_reparar"):
            with st.spinner("Re-processando XMLs armazenados..."):
                qtd = reparar_dados_nfe(cnpj_sel)
            st.success(f"{qtd} registro(s) atualizados. Clique em **Consultar** para recarregar.")
            st.session_state.pop("siga_resultado", None)

    if consultar:
        data_ini   = f"{ano_sel}-{mes_sel:02d}-01"
        ultimo_dia = calendar.monthrange(ano_sel, mes_sel)[1]
        data_fim   = f"{ano_sel}-{mes_sel:02d}-{ultimo_dia}"
        periodo    = f"{_MESES[mes_sel - 1]}/{ano_sel}"

        with st.spinner("Consultando acervo local..."):
            rec  = listar_resultados_por_periodo(cnpj_sel, data_ini, data_fim, modelo="55", papel="Recebida")
            emit = listar_resultados_por_periodo(cnpj_sel, data_ini, data_fim, modelo="55", papel="Emitida")
            nfce = listar_resultados_por_periodo(cnpj_sel, data_ini, data_fim, modelo="65")

        st.session_state["siga_resultado"] = {
            "rec": rec, "emit": emit, "nfce": nfce,
            "cnpj": cnpj_sel, "razao": razao_sel, "periodo": periodo,
            "ano": ano_sel, "mes": mes_sel,
        }

    r = st.session_state.get("siga_resultado")
    if not r:
        st.info("Selecione a empresa e o período, depois clique em **Consultar**.")
        return

    # Avisa se o resultado é de uma empresa/período diferente da seleção atual
    if r["cnpj"] != cnpj_sel or r["ano"] != ano_sel or r["mes"] != mes_sel:
        st.warning("Resultado exibido difere da seleção atual. Clique em **Consultar** para atualizar.")

    rec  = r["rec"]
    emit = r["emit"]
    nfce = r["nfce"]
    razao_r   = r["razao"]
    cnpj_r    = r["cnpj"]
    periodo_r = r["periodo"]

    # Resumo
    tot_r = sum(float(d.get("valor_total", 0)) for d in rec)
    tot_e = sum(float(d.get("valor_total", 0)) for d in emit)
    tot_n = sum(float(d.get("valor_total", 0)) for d in nfce)

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("NF-e Recebidas",  len(rec),  _fmt_brl(tot_r))
    mc2.metric("NF-e Emitidas",   len(emit), _fmt_brl(tot_e))
    mc3.metric("NFC-e Emitidas",  len(nfce), _fmt_brl(tot_n))
    mc4.metric("Total Geral",
               len(rec) + len(emit) + len(nfce),
               _fmt_brl(tot_r + tot_e + tot_n))

    # Exportação
    ex1, ex2, _ = st.columns([1, 1, 4])
    with ex1:
        xlsx = _gerar_excel(rec, emit, nfce, cnpj_r, razao_r, periodo_r)
        st.download_button(
            "⬇ Excel",
            xlsx,
            file_name=f"fiscal_{cnpj_r}_{r['ano']}{r['mes']:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with ex2:
        pdf_bytes = _gerar_pdf(rec, emit, nfce, cnpj_r, razao_r, periodo_r)
        st.download_button(
            "⬇ PDF (DTE)",
            pdf_bytes,
            file_name=f"DTE_{cnpj_r}_{r['ano']}{r['mes']:02d}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    # Abas
    tab_rec, tab_emit, tab_nfce = st.tabs([
        f"NF-e Recebidas ({len(rec)})",
        f"NF-e Emitidas ({len(emit)})",
        f"NFC-e Emitidas ({len(nfce)})",
    ])
    with tab_rec:
        _render_tabela(rec,  "recebida")
    with tab_emit:
        _render_tabela(emit, "emitida")
    with tab_nfce:
        _render_tabela(nfce, "nfce")
