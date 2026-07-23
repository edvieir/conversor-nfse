"""views/siga_downloads.py — Relatórios do SIGA (SEFAZ-CE): NF-e, NFC-e e índices de malha.

Modelo híbrido: o cron (core/siga_scheduler.py) gera os arquivos de madrugada
em data/siga_downloads/<cnpj>/. Esta tela lista o que já está em cache (rápido,
funciona mesmo com o SIGA fora do ar) e oferece "Atualizar agora" por aba para
buscar ao vivo quando necessário.
"""
import calendar
import datetime
import io
import zipfile
from pathlib import Path
from zoneinfo import ZoneInfo

_TZ_BR = ZoneInfo("America/Fortaleza")

import streamlit as st

from auth.security import current_user
from assets.icons import icon
from db.database import (
    listar_certificados, carregar_certificado,
    enfileirar_xml, status_fila_xml,
    listar_xmls_por_periodo, listar_resultados_por_periodo,
)
from views import nav
from core.siga_scheduler import ABAS as _ABAS_DOCUMENTOS, SAIDA_DIR

XML_DIR = Path(__file__).parent.parent / "data" / "siga_xml"


def _estilos_padrao():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    return {
        "VERDE":      PatternFill("solid", fgColor="1E3A2E"),
        "VERDE_FONT": Font(bold=True, color="FFFFFF", size=10),
        "TITULO":     Font(bold=True, size=14, color="1E3A2E"),
        "SUBTITULO":  Font(bold=True, size=10, color="1E3A2E"),
        "CAPTION":    Font(size=9, italic=True, color="666666"),
        "ALT_FILL":   PatternFill("solid", fgColor="EEF4EB"),
        "BORDA":      Border(*(Side(style="thin", color="AAAAAA"),) * 4),
        "CENTRO":     Alignment(horizontal="center", vertical="center"),
    }


def _cabecalho_padrao(ws, titulo: str, razao: str, cnpj: str, ultima_col: int):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    s = _estilos_padrao()
    col_letra = get_column_letter(ultima_col)

    ws.merge_cells(f"A1:{col_letra}1")
    ws["A1"] = titulo
    ws["A1"].font = s["TITULO"]
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A3:{col_letra}3")
    ws["A3"] = f"{razao} — CNPJ: {_fmt_cnpj(cnpj)}"
    ws["A3"].font = s["SUBTITULO"]

    ws["A4"] = f"Baixado em: {datetime.datetime.now(_TZ_BR).strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A4"].font = s["CAPTION"]


def _formatar_relatorio_siga(
    conteudo: bytes, cnpj: str, razao: str, periodo: str, label: str,
) -> bytes:
    """Reformata um relatório bruto do SIGA (NF-e/NFC-e) no padrão verde."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    src = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    ws_src = src.worksheets[0]
    rows = list(ws_src.iter_rows(values_only=True))
    if not rows:
        return conteudo

    cab_orig = [str(c or "") for c in rows[0]]
    dados = rows[1:]

    col_map = []
    larguras = []
    for i, nome in enumerate(cab_orig):
        nl = nome.lower()
        if "cnpj" in nl:
            col_map.append(("CNPJ", i))
            larguras.append(20)
        elif "raz" in nl or "social" in nl:
            col_map.append(("Razão Social", i))
            larguras.append(36)
        elif "uf" == nl.strip():
            col_map.append(("UF", i))
            larguras.append(6)
        elif "número" in nl or "numero" in nl:
            col_map.append(("Número", i))
            larguras.append(12)
        elif "data" in nl:
            col_map.append(("Data Emissão", i))
            larguras.append(14)
        elif "indicador" in nl or "situação" in nl or "situacao" in nl:
            col_map.append(("Situação", i))
            larguras.append(14)
        elif "valor" in nl:
            col_map.append(("Valor R$", i))
            larguras.append(16)
        elif "chave" in nl:
            col_map.append(("Chave NF-e", i))
            larguras.append(50)
        else:
            col_map.append((nome, i))
            larguras.append(14)

    s = _estilos_padrao()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label

    ultima_col = len(col_map)
    _cabecalho_padrao(ws, f"{label.upper()} - {periodo}", razao, cnpj, ultima_col)

    r_header = 6
    for ci, (nome, _) in enumerate(col_map, 1):
        cel = ws.cell(row=r_header, column=ci, value=nome)
        cel.fill = s["VERDE"]
        cel.font = s["VERDE_FONT"]
        cel.alignment = s["CENTRO"]
        cel.border = s["BORDA"]
    ws.row_dimensions[r_header].height = 22

    idx_valor_col = None
    for ci, (nome, _) in enumerate(col_map, 1):
        if "valor" in nome.lower():
            idx_valor_col = ci

    total_valor = 0.0
    for ri, row in enumerate(dados, r_header + 1):
        fill = s["ALT_FILL"] if ri % 2 == 0 else None
        for ci, (_, src_idx) in enumerate(col_map, 1):
            v = row[src_idx] if src_idx < len(row) else ""
            if v is None:
                v = ""
            if ci == idx_valor_col:
                try:
                    v = float(v)
                    total_valor += v
                except (ValueError, TypeError):
                    v = 0.0
            cel = ws.cell(row=ri, column=ci, value=v)
            cel.border = s["BORDA"]
            if fill:
                cel.fill = fill
            if ci == idx_valor_col:
                cel.number_format = '#,##0.00'

    r_total = r_header + 1 + len(dados)
    if idx_valor_col:
        lbl_col = idx_valor_col - 1 if idx_valor_col > 1 else 1
        cel_lbl = ws.cell(row=r_total, column=lbl_col, value="TOTAL")
        cel_lbl.font = Font(bold=True)
        cel_lbl.fill = s["VERDE"]
        cel_lbl.font = s["VERDE_FONT"]
        cel_lbl.border = s["BORDA"]
        cel_lbl.alignment = s["CENTRO"]
        cel_val = ws.cell(row=r_total, column=idx_valor_col, value=total_valor)
        cel_val.font = s["VERDE_FONT"]
        cel_val.fill = s["VERDE"]
        cel_val.border = s["BORDA"]
        cel_val.number_format = '#,##0.00'
        cel_val.alignment = s["CENTRO"]

    for ci, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _planilha_nfce_combinada(
    cnpj: str, razao: str, periodo: str,
    caminho_emit: Path, caminho_canc: Path | None,
) -> bytes:
    """Gera planilha NFC-e no padrão verde com autorizadas e canceladas lado a lado."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    MESES = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
    ]

    def _extrair_por_mes(caminho: Path) -> dict[int, tuple[int, float]]:
        por_mes: dict[int, tuple[int, float]] = {m: (0, 0.0) for m in range(1, 13)}
        if not caminho or not caminho.exists():
            return por_mes
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return por_mes
        cab = [str(c or "").lower() for c in rows[0]]
        idx_data = next((i for i, c in enumerate(cab) if "data" in c), None)
        idx_valor = next((i for i, c in enumerate(cab) if "valor" in c), None)
        for row in rows[1:]:
            if idx_data is None or idx_data >= len(row) or not row[idx_data]:
                continue
            data_str = str(row[idx_data])
            try:
                if "-" in data_str:
                    mes = int(data_str.split("-")[1])
                elif "/" in data_str:
                    partes = data_str.split("/")
                    mes = int(partes[1]) if len(partes[0]) <= 2 else int(partes[1])
                else:
                    continue
            except (ValueError, IndexError):
                continue
            if 1 <= mes <= 12:
                qtd, val = por_mes[mes]
                v = 0.0
                if idx_valor is not None and idx_valor < len(row) and row[idx_valor]:
                    try:
                        v = float(row[idx_valor])
                    except (ValueError, TypeError):
                        pass
                por_mes[mes] = (qtd + 1, val + v)
        return por_mes

    emit_mes = _extrair_por_mes(caminho_emit)
    canc_mes = _extrair_por_mes(caminho_canc) if caminho_canc else {m: (0, 0.0) for m in range(1, 13)}

    s = _estilos_padrao()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    _cabecalho_padrao(ws, f"NFC-E CF-E AUTORIZADOS - {periodo}", razao, cnpj, 6)

    r = 5
    headers_l1 = [("MÊS", 1, 1), ("NFC-E AUTORIZADAS", 2, 3), ("NFC-E CANCELADAS", 4, 6)]
    for txt, c_ini, c_fim in headers_l1:
        if c_ini != c_fim:
            ws.merge_cells(start_row=r, start_column=c_ini, end_row=r, end_column=c_fim)
        cel = ws.cell(row=r, column=c_ini, value=txt)
        cel.fill = s["VERDE"]
        cel.font = s["VERDE_FONT"]
        cel.alignment = s["CENTRO"]
        for c in range(c_ini, c_fim + 1):
            ws.cell(row=r, column=c).fill = s["VERDE"]
            ws.cell(row=r, column=c).border = s["BORDA"]
    ws.row_dimensions[r].height = 22

    r = 6
    sub_headers = ["MÊS", "QTD", "VALOR R$", "QTD", "QTD%", "VALOR R$"]
    for ci, txt in enumerate(sub_headers, 1):
        cel = ws.cell(row=r, column=ci, value=txt)
        cel.fill = s["VERDE"]
        cel.font = s["VERDE_FONT"]
        cel.alignment = s["CENTRO"]
        cel.border = s["BORDA"]
    ws.row_dimensions[r].height = 20

    tot_emit_q, tot_emit_v = 0, 0.0
    tot_canc_q, tot_canc_v = 0, 0.0

    for i, mes_num in enumerate(range(1, 13), start=1):
        row = r + i
        eq, ev = emit_mes[mes_num]
        cq, cv = canc_mes[mes_num]
        tot_emit_q += eq
        tot_emit_v += ev
        tot_canc_q += cq
        tot_canc_v += cv
        pct = f"{cq / eq * 100:.0f}%" if eq > 0 and cq > 0 else "-"
        fill = s["ALT_FILL"] if i % 2 == 0 else None
        vals = [MESES[mes_num - 1], eq, ev, cq, pct, cv]
        for ci, v in enumerate(vals, 1):
            cel = ws.cell(row=row, column=ci, value=v)
            cel.border = s["BORDA"]
            cel.alignment = s["CENTRO"]
            if fill:
                cel.fill = fill
            if ci in (3, 6):
                cel.number_format = '#,##0.00'

    row_total = r + 13
    pct_total = f"{tot_canc_q / tot_emit_q * 100:.0f}%" if tot_emit_q > 0 and tot_canc_q > 0 else "-"
    vals_total = ["Total", tot_emit_q, tot_emit_v, tot_canc_q, pct_total, tot_canc_v]
    for ci, v in enumerate(vals_total, 1):
        cel = ws.cell(row=row_total, column=ci, value=v)
        cel.fill = s["VERDE"]
        cel.font = s["VERDE_FONT"]
        cel.border = s["BORDA"]
        cel.alignment = s["CENTRO"]
        if ci in (3, 6):
            cel.number_format = '#,##0.00'

    col_widths = [14, 12, 16, 12, 10, 16]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    if caminho_emit.exists():
        src = openpyxl.load_workbook(caminho_emit, data_only=True)
        ws_det = wb.create_sheet("Detalhamento")
        ws_src = src.worksheets[0]
        for ri, row in enumerate(ws_src.iter_rows(values_only=True), 1):
            for ci, v in enumerate(row, 1):
                cel = ws_det.cell(row=ri, column=ci, value=v)
                if ri == 1:
                    cel.fill = s["VERDE"]
                    cel.font = s["VERDE_FONT"]
                    cel.alignment = s["CENTRO"]
                cel.border = s["BORDA"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

_LABELS = {
    "NF_E_emitidas":               "NF-e Emitidas",
    "NF_E_recebidas":              "NF-e Recebidas",
    "NF_E_canceladas":             "NF-e Canceladas",
    "NFC_E_emitidas":              "NFC-e Emitidas",
    "INDICADORES_MALHA_pendencias":"Índices de Malha (pendências / não escrituradas)",
}


def _fmt_cnpj(c: str) -> str:
    c = (c or "").strip().zfill(14)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


def _mes_corrente() -> str:
    return datetime.date.today().strftime("%Y-%m")


_ABAS_OCULTAS = {"NFC_E_canceladas"}

def _listar_abas() -> list[tuple[str, str, str, dict]]:
    """Retorna (nome_arquivo, tipo_ou_None, nome_aba_ou_None, filtros_ou_None)."""
    abas = []
    for tipo, itens in _ABAS_DOCUMENTOS.items():
        for nome_aba, filtros in itens:
            chave = f"{tipo}_{nome_aba}"
            if chave not in _ABAS_OCULTAS:
                abas.append((chave, tipo, nome_aba, filtros))
    abas.append(("INDICADORES_MALHA_pendencias", None, None, None))
    return abas


def _caminho_cache(cnpj: str, nome_arquivo: str, periodo: str) -> Path:
    return SAIDA_DIR / cnpj / f"{nome_arquivo}_{periodo}.xlsx"


def _gerar_ao_vivo(usuario: str, cnpj: str, nome_arquivo: str, tipo, nome_aba, filtros, periodo: str) -> bytes | None:
    """Busca ao vivo no SIGA. Retorna None se não houver dado (ex.: sem pendências)."""
    from core import siga_sefaz

    cert = carregar_certificado(usuario, cnpj)
    if not cert:
        raise RuntimeError("Certificado não encontrado para esta empresa.")
    pfx_bytes, senha = cert
    sessao = siga_sefaz._sessao(pfx_bytes, senha)
    token = siga_sefaz.login(sessao)["access_token"]

    if tipo is None:  # índices de malha
        indicadores = siga_sefaz.listar_indicadores_malha(sessao, token, cnpj)
        if not indicadores:
            return None
        sid = siga_sefaz.solicitar_download_indicadores(sessao, token, cnpj)
    else:
        sid = siga_sefaz.solicitar_download(
            sessao, token, cnpj, tipo, dat_referencia=[periodo], **filtros,
        )

    return siga_sefaz.aguardar_e_baixar(sessao, token, sid)


def _segundos_restantes(proxima_consulta: str | None) -> int:
    if not proxima_consulta:
        return 0
    try:
        dt = datetime.datetime.fromisoformat(proxima_consulta)
        return max(0, int((dt - datetime.datetime.now()).total_seconds()))
    except Exception:
        return 0


def _secao_baixar_tudo(user: dict, cnpj: str, razao_social: str):
    """Baixa o XML completo (NF-e + NFC-e, emitidas e recebidas) via distribuição
    NSU da SEFAZ Nacional -- mesmo motor da página NFE/NFCE, sem o limite de
    20/hora (esse aqui é 1 consulta/hora por CNPJ, controlado por NSU)."""
    from db.database import get_nsu_cnpj

    with st.container(border=True):
        st.markdown("**Baixar tudo (XML completo via SEFAZ Nacional)**")
        st.caption(
            "NF-e e NFC-e, emitidas e recebidas, direto em XML (não resumo). "
            "Limite da SEFAZ: 1 consulta por hora por CNPJ."
        )

        estado_nsu = get_nsu_cnpj(cnpj)
        seg_rest = _segundos_restantes(estado_nsu.get("proxima_consulta"))

        if seg_rest > 0:
            minutos = seg_rest // 60
            st.warning(f"Bloqueado pela SEFAZ — tente novamente em ~{minutos} min.")
            return

        col_ini, col_fim, col_btn = st.columns([1, 1, 1])
        with col_ini:
            data_ini = st.date_input("De", value=datetime.date.today().replace(day=1), key="siga_dl_tudo_ini")
        with col_fim:
            data_fim = st.date_input("Até", value=datetime.date.today(), key="siga_dl_tudo_fim")
        with col_btn:
            st.markdown("<br>", unsafe_allow_html=True)
            disparar = st.button("Baixar tudo", type="primary", use_container_width=True, key="siga_dl_tudo_btn")

        if disparar:
            cert = carregar_certificado(user["username"], cnpj)
            if not cert:
                st.error("Certificado não encontrado para esta empresa.")
                return
            pfx_bytes, pfx_senha = cert

            progress_bar = st.progress(0, text="Conectando na SEFAZ Nacional...")
            log_placeholder = st.empty()

            def on_progress(frac: float):
                progress_bar.progress(min(frac, 1.0), text=f"Processando... {int(min(frac,1.0)*100)}%")

            def on_log(linhas: list):
                log_placeholder.code("\n".join(linhas[-30:]), language="")

            try:
                from core.nfe_sefaz import executar_consulta_sefaz

                zip_bytes, log_final = executar_consulta_sefaz(
                    pfx_bytes=pfx_bytes,
                    pfx_senha=pfx_senha,
                    empresas=[{"cnpj": cnpj, "nome": razao_social}],
                    ambiente="1",
                    uf="CE",
                    data_ini=data_ini,
                    data_fim=data_fim,
                    tipo_doc="ambos",
                    papel_filtro="ambos",
                    incluir_xml=True,
                    incluir_pdf=False,
                    incluir_excel=False,
                    log_cb=on_log,
                    progress_cb=on_progress,
                    salvar_db=True,
                )
                progress_bar.progress(1.0, text="Concluído!")
                log_placeholder.empty()

                with st.expander("Log da consulta", expanded=not bool(zip_bytes)):
                    st.code("\n".join(log_final), language="")

                if zip_bytes:
                    st.success(f"{len(zip_bytes) // 1024} KB de XML baixados.")
                    st.download_button(
                        "Baixar ZIP com os XMLs",
                        data=zip_bytes,
                        file_name=f"xml_completo_{cnpj}_{data_ini}_{data_fim}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )
                else:
                    st.info("Nenhum documento novo encontrado no período (ou já sincronizado antes).")
            except Exception as e:
                progress_bar.empty()
                log_placeholder.empty()
                st.error(f"Erro: {e}")


def _periodo_para_datas(periodo: str) -> tuple[str, str] | None:
    try:
        ano, mes = periodo.split("-")
        ano, mes = int(ano), int(mes)
        ultimo_dia = calendar.monthrange(ano, mes)[1]
        return f"{ano}-{mes:02d}-01", f"{ano}-{mes:02d}-{ultimo_dia}"
    except Exception:
        return None


def _secao_acervo_local(cnpj: str, razao_social: str, periodo: str):
    """XML e planilha do que já está sincronizado no acervo local -- sem
    chamar SEFAZ nem SIGA de novo, então sem limite de consulta."""
    datas = _periodo_para_datas(periodo)
    if not datas:
        return
    data_ini_str, data_fim_str = datas

    with st.container(border=True):
        st.markdown("**Acervo local já sincronizado**")
        st.caption(
            "XML e planilha do que já está salvo no banco para o período — "
            "não chama SEFAZ nem SIGA de novo, sem limite de consulta."
        )

        col_xml, col_xlsx = st.columns(2)

        with col_xml:
            xmls = listar_xmls_por_periodo(cnpj, data_ini_str, data_fim_str)
            st.metric("Documentos no acervo", len(xmls))
            if xmls:
                buf = io.BytesIO()
                with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for r in xmls:
                        pasta = f"{r.get('modelo','NF-e')}_{r.get('papel','?')}"
                        chave = r.get("chave", "x")
                        xml = r.get("xml_conteudo", "")
                        if xml:
                            zf.writestr(f"XMLs/{pasta}/{chave}.xml", xml)
                st.download_button(
                    "Baixar XMLs do acervo (.zip)",
                    data=buf.getvalue(),
                    file_name=f"acervo_xml_{cnpj}_{periodo}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key="siga_acervo_xml_dl",
                )

        with col_xlsx:
            rec  = listar_resultados_por_periodo(cnpj, data_ini_str, data_fim_str, modelo="55", papel="Recebida")
            emit = listar_resultados_por_periodo(cnpj, data_ini_str, data_fim_str, modelo="55", papel="Emitida")
            nfce = listar_resultados_por_periodo(cnpj, data_ini_str, data_fim_str, modelo="65")
            total = len(rec) + len(emit) + len(nfce)
            st.metric("Notas para planilha DTE", total)
            if total:
                from views.siga_consulta import _gerar_excel
                xlsx_bytes = _gerar_excel(rec, emit, nfce, cnpj, razao_social, periodo)
                st.download_button(
                    "Baixar planilha Excel (DTE)",
                    data=xlsx_bytes,
                    file_name=f"DTE_{cnpj}_{periodo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="siga_acervo_xlsx_dl",
                )


def _secao_relatorio_deconf(cnpj: str, razao_social: str, periodo: str):
    """Relatório 'PESQUISA SIGET - DECONF' -- usa o cache dos índices de
    malha (ou busca ao vivo se não existir ainda)."""
    st.markdown("---")
    st.markdown("### Relatório de malhas SIGA")
    st.caption(
        "Entradas não escrituradas, não seladas, inventários e declarações "
        "omissas, no padrão usado pela equipe. Multa dos dois primeiros "
        "grupos calculada automaticamente (10% / 20%); inventário e "
        "declarações ficam em branco pra preenchimento manual."
    )

    tipo_estabelecimento = st.selectbox("Estabelecimento", ["MATRIZ", "FILIAL"], key="deconf_tipo_estab")

    caminho_malha = _caminho_cache(cnpj, "INDICADORES_MALHA_pendencias", periodo)

    if st.button("Gerar relatório de malhas", type="primary", use_container_width=True, key="deconf_gerar"):

        with st.spinner("Buscando índices de malha..."):
            try:
                if caminho_malha.exists():
                    conteudo = caminho_malha.read_bytes()
                else:
                    from core import siga_sefaz

                    cert = carregar_certificado(current_user()["username"], cnpj)
                    if not cert:
                        st.error("Certificado não encontrado para esta empresa.")
                        return
                    pfx_bytes, senha = cert
                    sessao = siga_sefaz._sessao(pfx_bytes, senha)
                    token = siga_sefaz.login(sessao)["access_token"]
                    indicadores = siga_sefaz.listar_indicadores_malha(sessao, token, cnpj)
                    if not indicadores:
                        st.info("Sem pendências de malha para esta empresa — nada a gerar.")
                        return
                    sid = siga_sefaz.solicitar_download_indicadores(sessao, token, cnpj)
                    conteudo = siga_sefaz.aguardar_e_baixar(sessao, token, sid)
                    caminho_malha.parent.mkdir(parents=True, exist_ok=True)
                    caminho_malha.write_bytes(conteudo)

                from core.siga_sefaz import parse_indicadores_malha
                from core.relatorio_deconf import gerar_relatorio_deconf, parse_indicadores_17_18

                resumo, detalhes = parse_indicadores_malha(conteudo)
                inventarios, declaracoes = parse_indicadores_17_18(conteudo)

                xlsx_bytes = gerar_relatorio_deconf(
                    cnpj=cnpj, razao_social=razao_social,
                    resumo=resumo, detalhes=detalhes,
                    inventarios=inventarios, declaracoes=declaracoes,
                    tipo_estabelecimento=tipo_estabelecimento,
                )
                st.session_state["deconf_bytes"] = xlsx_bytes
                st.session_state["deconf_nome"] = f"DECONF_{cnpj}_{periodo}.xlsx"
                st.success("Relatório gerado.")
            except Exception as e:
                st.error(f"Erro: {e}")

    if st.session_state.get("deconf_bytes"):
        st.download_button(
            "Baixar relatório de malhas SIGA (.xlsx)",
            data=st.session_state["deconf_bytes"],
            file_name=st.session_state["deconf_nome"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="deconf_dl",
        )


def render():
    user = current_user()
    nav.render("siga_downloads")

    st.markdown("## Relatórios SIGA — SEFAZ-CE")
    st.caption(
        "NF-e, NFC-e e índices de malha fiscal direto do SIGA. "
        "Os arquivos são atualizados automaticamente todo dia; use **Atualizar agora** "
        "numa aba específica se precisar de algo mais recente."
    )

    certs = listar_certificados(user["username"])
    if not certs:
        st.warning(
            "Nenhum certificado digital cadastrado. "
            "Acesse a página **Certificados** para adicionar."
        )
        return

    opcoes = {
        f"{c['razao_social']} ({_fmt_cnpj(c['cnpj'])})": c["cnpj"]
        for c in certs
    }

    c1, c2 = st.columns([3, 1])
    with c1:
        empresa_label = st.selectbox("Empresa", list(opcoes.keys()), key="siga_dl_empresa")
    with c2:
        periodo = st.text_input(
            "Período (AAAA-MM)", value=_mes_corrente(), key="siga_dl_periodo",
            help="Só afeta NF-e/NFC-e. Índices de malha são sempre a situação atual.",
        )

    cnpj_sel = opcoes[empresa_label]
    razao_sel = next(c["razao_social"] for c in certs if c["cnpj"] == cnpj_sel)

    st.markdown("---")

    _secao_baixar_tudo(user, cnpj_sel, razao_sel)

    st.markdown("---")

    _secao_acervo_local(cnpj_sel, razao_sel, periodo)

    st.markdown("---")

    from core.siga_sefaz import extrair_chaves_xlsx

    for nome_arquivo, tipo, nome_aba, filtros in _listar_abas():
        label = _LABELS.get(nome_arquivo, nome_arquivo)
        caminho = _caminho_cache(cnpj_sel, nome_arquivo, periodo)

        with st.container(border=True):
            col_info, col_baixar, col_fila, col_atualizar = st.columns([3, 1, 1, 1])

            with col_info:
                if caminho.exists():
                    mtime = datetime.datetime.fromtimestamp(caminho.stat().st_mtime)
                    tamanho = caminho.stat().st_size
                    tamanho_fmt = f"{tamanho} bytes" if tamanho < 1024 else f"{tamanho / 1024:.1f} KB"
                    ic = icon("check-circle", 16, "#10B981")
                    st.markdown(
                        f"**{label}**<br>"
                        f'<span style="color:#64748B;font-size:.8rem;">{ic} '
                        f'Atualizado em {mtime.strftime("%d/%m/%Y %H:%M")} '
                        f'({tamanho_fmt})</span>',
                        unsafe_allow_html=True,
                    )
                else:
                    ic = icon("alert-triangle", 16, "#C97400")
                    st.markdown(
                        f"**{label}**<br>"
                        f'<span style="color:#C97400;font-size:.8rem;">{ic} '
                        f'Ainda não gerado para este período.</span>',
                        unsafe_allow_html=True,
                    )

            with col_baixar:
                if caminho.exists():
                    if nome_arquivo == "NFC_E_emitidas":
                        cam_canc = _caminho_cache(cnpj_sel, "NFC_E_canceladas", periodo)
                        dl_data = _planilha_nfce_combinada(
                            cnpj_sel, razao_sel, periodo, caminho,
                            cam_canc if cam_canc.exists() else None,
                        )
                        dl_name = f"NFC_E_{cnpj_sel}_{periodo}.xlsx"
                    elif nome_arquivo.startswith(("NF_E_", "NFC_E_")):
                        dl_data = _formatar_relatorio_siga(
                            caminho.read_bytes(), cnpj_sel, razao_sel, periodo, label,
                        )
                        dl_name = f"{nome_arquivo}_{cnpj_sel}_{periodo}.xlsx"
                    else:
                        dl_data = caminho.read_bytes()
                        dl_name = caminho.name
                    st.download_button(
                        "Baixar",
                        data=dl_data,
                        file_name=dl_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"dl_{nome_arquivo}",
                    )

            with col_fila:
                if caminho.exists():
                    if st.button("Enfileirar XML", key=f"fila_{nome_arquivo}", use_container_width=True):
                        todas = extrair_chaves_xlsx(caminho.read_bytes())
                        # A consulta avulsa da SEFAZ Nacional (consChNFe) só aceita
                        # NF-e (modelo 55) -- NFC-e (modelo 65) sempre rejeita com
                        # cStat=618, então nem vale gastar a cota de 20/hora com isso.
                        chaves = [c for c in todas if c[20:22] == "55"]
                        ignoradas = len(todas) - len(chaves)
                        qtd = enfileirar_xml(user["username"], cnpj_sel, chaves)
                        msg = f"{qtd} chave(s) de NF-e enfileirada(s) ({len(todas)} encontradas em {label})."
                        if ignoradas:
                            msg += f" {ignoradas} chave(s) de NFC-e ignorada(s) — a SEFAZ Nacional não permite consulta avulsa desse modelo."
                        st.success(msg)
                        st.rerun()

            with col_atualizar:
                if st.button("Atualizar agora", key=f"upd_{nome_arquivo}", use_container_width=True):
                    with st.spinner(f"Buscando {label} no SIGA..."):
                        try:
                            conteudo = _gerar_ao_vivo(
                                user["username"], cnpj_sel, nome_arquivo, tipo, nome_aba, filtros, periodo,
                            )
                        except Exception as e:
                            st.error(f"Erro: {e}")
                        else:
                            if conteudo is None:
                                st.info("Sem dados para esta aba (ex.: sem pendências de malha).")
                                caminho.unlink(missing_ok=True)
                            else:
                                caminho.parent.mkdir(parents=True, exist_ok=True)
                                caminho.write_bytes(conteudo)
                                try:
                                    from core.siga_sefaz import persistir_relatorio
                                    persistir_relatorio(cnpj_sel, nome_arquivo, tipo, conteudo, periodo)
                                except Exception as e_db:
                                    st.warning(f"Salvo em cache, mas falhou ao persistir no banco (Power BI): {e_db}")
                                st.success(f"{label} atualizado ({len(conteudo)} bytes).")

                            if nome_arquivo == "NFC_E_emitidas":
                                try:
                                    canc = _gerar_ao_vivo(
                                        user["username"], cnpj_sel, "NFC_E_canceladas",
                                        "NFC_E", "canceladas",
                                        {"papel_operacao": "EMITENTE", "resultado_processamento": "CANCELADA"},
                                        periodo,
                                    )
                                    cam_canc = _caminho_cache(cnpj_sel, "NFC_E_canceladas", periodo)
                                    if canc:
                                        cam_canc.parent.mkdir(parents=True, exist_ok=True)
                                        cam_canc.write_bytes(canc)
                                    else:
                                        cam_canc.unlink(missing_ok=True)
                                except Exception:
                                    pass

                            st.rerun()

    # ── Fila de XML avulso (limite de 20/hora) ───────────────────────────────
    st.markdown("---")
    st.markdown("### Fila de XML avulso")
    st.caption(
        "Os relatórios do SIGA acima só trazem a chave de acesso de cada nota, não "
        "o XML. Use \"Enfileirar XML\" em qualquer aba pra buscar o XML completo "
        "dessas chaves na SEFAZ Nacional — limite de **20 consultas por hora por "
        "certificado**, por isso roda em fila, não na hora."
    )

    status = status_fila_xml(cnpj_sel)
    pendente  = status.get("PENDENTE", 0)
    concluido = status.get("CONCLUIDO", 0)
    erro      = status.get("ERRO", 0)

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Pendentes", pendente)
    col_b.metric("Concluídos", concluido)
    col_c.metric("Com erro", erro)

    if pendente:
        horas = -(-pendente // 20)  # arredonda pra cima
        st.info(f"Fila roda a cada hora (até 20 por vez) — previsão de ~{horas}h para concluir os {pendente} pendentes.")

    pasta_xml = XML_DIR / cnpj_sel
    arquivos_xml = sorted(pasta_xml.glob("*.xml")) if pasta_xml.exists() else []
    if arquivos_xml:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in arquivos_xml:
                zf.write(f, arcname=f.name)
        st.download_button(
            f"Baixar {len(arquivos_xml)} XML(s) prontos (.zip)",
            data=buf.getvalue(),
            file_name=f"xml_fila_{cnpj_sel}.zip",
            mime="application/zip",
            use_container_width=True,
        )

    _secao_relatorio_deconf(cnpj_sel, razao_sel, periodo)
