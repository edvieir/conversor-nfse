"""
core/conversor_xlsx.py — Conversão XML → XLSX (SPED GOV)
Lógica copiada integralmente de app_web.py — NENHUMA linha de conversão alterada.
Adicionado: st.progress() no loop principal para feedback visual.
"""

import sys
import io
import os
import re
import tempfile
import contextlib
import zipfile
import xml.etree.ElementTree as _ET
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import nfse_xml_to_txt as C


# ── Conversor ZIP: SharedStrings → InlineStr + Formatos Numéricos ────────────
# Colunas com formato #,##0.00 (1-based): monetárias + alíquota
# Col 32 (Alíquota) usa o mesmo formato do "deu certo" — numFmtId=4 (#,##0.00)
_COLS_MONEY = frozenset([20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 35, 36])
_COL_ALIQ   = 32   # Alíquota usa 0.00 (sem separador de milhar)


def _col_num(letters: str) -> int:
    """'T' → 20,  'AF' → 32"""
    n = 0
    for c in letters.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _xlsx_para_inlinestr(xlsx_bytes: bytes) -> bytes:
    """
    Pós-processa o XLSX gerado pelo openpyxl para:

    A) Converter todas as células string para inlineStr (<is><t>texto</t></is>)
       conforme esperado pelo portal ISS Fortaleza (Apache POI).

    B) Garantir que células monetárias e de alíquota usem os formatos corretos
       com applyNumberFormat="1" — sem isso o portal exibe '2' em vez de '2,00'.

    Etapas inlineStr:
      1. Lê sharedStrings.xml → tabela índice→texto
      2. <c t="s"><v>N</v></c>     → <c t="inlineStr"><is><t>TEXT</t></is></c>
      3. <c t="inlineStr"><v>T</v></c> → <c t="inlineStr"><is><t>T</t></is></c>
      4. <c t="inlineStr"/>        → <c t="inlineStr"><is><t/></is></c>
      5. Remove sharedStrings.xml e suas referências

    Etapas de formato numérico:
      6. Lê styles.xml → encontra índices para numFmtId=4 (#,##0.00) e 2 (0.00)
      7. Adiciona applyNumberFormat="1" onde falta
      8. Atualiza atributo s= das células monetárias e de alíquota nas linhas de dados
    """
    _NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

    src   = zipfile.ZipFile(io.BytesIO(xlsx_bytes), 'r')
    nomes = src.namelist()

    # ── Passo 1: shared strings ───────────────────────────────────────────────
    tabela: list = []
    if 'xl/sharedStrings.xml' in nomes:
        ss_root = _ET.fromstring(src.read('xl/sharedStrings.xml').decode('utf-8'))
        for si in ss_root.iter(f'{{{_NS}}}si'):
            parts = [t.text or '' for t in si.iter(f'{{{_NS}}}t')]
            tabela.append(''.join(parts))

    # ── Passos 6-7: prepara correção de styles.xml ────────────────────────────
    idx_money  = -1   # índice cellXf para #,##0.00
    idx_aliq   = -1   # índice cellXf para 0.00
    styles_fix = None

    if 'xl/styles.xml' in nomes:
        sx = src.read('xl/styles.xml').decode('utf-8')

        def _ai(pat, s, d=0):
            m = re.search(pat, s)
            return int(m.group(1)) if m else d

        # Busca apenas dentro de <cellXfs> (o s= das células aponta para cellXfs, não cellStyleXfs)
        cellxfs_block = re.search(r'<cellXfs[^>]*>(.*?)</cellXfs>', sx, re.DOTALL)
        xfs = re.findall(r'<xf [^>]+/>', cellxfs_block.group(1)) if cellxfs_block else []
        for i, xf in enumerate(xfs):
            nfid = _ai(r'numFmtId="(\d+)"', xf)
            fnt  = _ai(r'fontId="(\d+)"',   xf)
            fll  = _ai(r'fillId="(\d+)"',   xf)
            if fnt == 0 and fll == 0:
                if nfid == 4 and idx_money < 0: idx_money = i
                if nfid == 2 and idx_aliq  < 0: idx_aliq  = i

        # Cria estilos ausentes
        cur_count = len(xfs)
        if idx_money < 0:
            entry = '<xf numFmtId="4" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            sx = sx.replace('</cellXfs>', entry + '</cellXfs>')
            idx_money = cur_count
            cur_count += 1
        if idx_aliq < 0:
            entry = '<xf numFmtId="2" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            sx = sx.replace('</cellXfs>', entry + '</cellXfs>')
            idx_aliq = cur_count

        # Adiciona applyNumberFormat="1" onde falta
        def _fix_apply(m):
            xf = m.group(0)
            nm = re.search(r'numFmtId="(\d+)"', xf)
            if nm and int(nm.group(1)) > 0 and 'applyNumberFormat' not in xf:
                xf = xf[:-2] + ' applyNumberFormat="1"/>'
            return xf

        sx = re.sub(r'<xf [^>]+/>', _fix_apply, sx)
        styles_fix = sx.encode('utf-8')

    # ── Passo 8: corrige s= nas células de dados ──────────────────────────────
    def _fix_cell_s(xml_str: str) -> str:
        if idx_money < 0 and idx_aliq < 0:
            return xml_str

        def _upd(m):
            coord = m.group(1)
            rest  = m.group(2)
            lm = re.match(r'([A-Z]+)(\d+)', coord)
            if not lm:
                return m.group(0)
            col = _col_num(lm.group(1))
            row = int(lm.group(2))
            if row < 2:
                return m.group(0)   # cabeçalho — não mexe
            if col in _COLS_MONEY and idx_money >= 0:
                target = idx_money
            elif col == _COL_ALIQ and idx_aliq >= 0:
                target = idx_aliq
            else:
                return m.group(0)
            rest2 = re.sub(r'\s*s="\d+"', '', rest)
            return f'<c r="{coord}" s="{target}"{rest2}>'

        return re.sub(r'<c r="([A-Z]+\d+)"([^>]*)>', _upd, xml_str)

    # ── Helpers inlineStr ─────────────────────────────────────────────────────
    def _esc(s: str) -> str:
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def _processar_sheet(xml_str: str) -> str:
        # Passo 2: shared strings → inlineStr
        def _rep_shared(m):
            pre, post, idx_str = m.group(1), m.group(2), m.group(3)
            idx  = int(idx_str)
            text = tabela[idx] if 0 <= idx < len(tabela) else ''
            tag  = f'<c {pre}t="inlineStr"{post}>'
            body = f'<is><t>{_esc(text)}</t></is>' if text else '<is><t/></is>'
            return f'{tag}{body}</c>'

        xml_str = re.sub(
            r'<c ((?:[^\S\n]*\S[^>]*?\s+)?)t="s"(\s*(?:[^>]*?)?)><v>(\d+)</v></c>',
            _rep_shared, xml_str
        )

        # Passo 3: inlineStr com <v>TEXTO</v>
        def _rep_v(m):
            attrs, text = m.group(1), m.group(2)
            body = f'<is><t>{_esc(text)}</t></is>' if text else '<is><t/></is>'
            return f'<c {attrs}>{body}</c>'

        xml_str = re.sub(
            r'<c ([^>]*t="inlineStr"[^>]*)><v>([^<]*)</v></c>',
            _rep_v, xml_str
        )

        # Passo 4: self-closing inlineStr
        xml_str = re.sub(
            r'<c ([^>]*t="inlineStr"[^/]*?)\s*/>',
            r'<c \1><is><t/></is></c>',
            xml_str
        )

        return xml_str

    # ── Grava ZIP de saída ────────────────────────────────────────────────────
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_out, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for item in src.infolist():
            nome    = item.filename
            content = src.read(nome)

            if nome == 'xl/sharedStrings.xml':
                continue

            elif nome == 'xl/styles.xml' and styles_fix:
                content = styles_fix

            elif nome == 'xl/_rels/workbook.xml.rels':
                s = content.decode('utf-8')
                s = re.sub(r'<Relationship[^>]+[Ss]hared[Ss]trings[^>]*/>\s*', '', s)
                content = s.encode('utf-8')

            elif nome == '[Content_Types].xml':
                s = content.decode('utf-8')
                s = re.sub(r'<Override[^>]+[Ss]hared[Ss]trings[^>]*/>\s*', '', s)
                content = s.encode('utf-8')

            elif nome.startswith('xl/worksheets/') and nome.endswith('.xml'):
                xml = _processar_sheet(content.decode('utf-8'))
                xml = _fix_cell_s(xml)          # corrige s= para formatos numéricos
                content = xml.encode('utf-8')

            zout.writestr(item, content)

    src.close()
    return buf_out.getvalue()


def processar_xlsx_sped(uploaded_files, im: str, competencia_filtro: str = ""):
    """Gera XLSX no layout exato do SPED GOV — aba 'Serviços Tomados', 43 colunas."""
    import glob as _glob
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    import streamlit as st

    # Cabeçalhos e larguras exatas do SPED GOV
    COLUNAS = [
        ("Tipo Doc.",                       24.14),
        ("Número",                          10.28),
        ("Código de Verificação",           23.57),
        ("Competência",                     14.57),
        ("Data",                            11.42),
        ("Vencimento",                      13.42),
        ("Número RPS",                      18.42),
        ("Série RPS",                       11.14),
        ("Tipo RPS",                        10.28),
        ("Natureza da Operação",            27.71),
        ("Regime Especial Tributação",      52.14),
        ("Operação Simples Nacional",       29.28),
        ("Incentivador Cultural",           22.85),
        ("Item da Lista",                   14.57),
        ("CNAE",                           123.14),
        ("ART",                              5.28),
        ("Código Obra",                     13.85),
        ("Número Empenho",                  19.42),
        ("Discriminação dos Serviços",     141.14),
        ("Valor dos Serviços",              20.28),
        ("Deduções Permitidas em Lei",      30.42),
        ("Desconto Condicionado",           25.28),
        ("Desconto Incondicionado",         27.00),
        ("Retenções Federais",              21.28),
        ("Outras Retenções",               19.42),
        ("PIS",                              4.42),
        ("COFINS",                           8.85),
        ("IRRF",                             5.71),
        ("CSLL",                             6.14),
        ("INSS",                             5.85),
        ("Base de Cálculo",                 17.28),
        ("Alíquota",                         9.85),
        ("Local da Prestação",              22.14),
        ("ISS Retido",                      11.71),
        ("Valor do ISS",                    13.85),
        ("Valor Líquido",                   14.85),
        ("Status Doc.",                     13.00),
        ("Inscrição Prestador",             21.14),
        ("CPF/CNPJ Prestador",              21.42),
        ("Razão Social/Nome do Prestador",  58.57),
        ("Escrituração",                    13.85),
        ("Origem",                           9.71),
        ("Status Aceite",                   14.85),
    ]

    IBGE_FORTALEZA = "2304400"

    def _float(v):
        try:
            return float(v) if v else 0.0
        except Exception:
            return 0.0

    def _str(v):
        """Retorna string limpa, ou '' para vazio (inlineStr via _forcar_inline)."""
        return str(v).strip() if v else ""

    def _data_fmt(iso, fmt):
        if not iso:
            return ""
        data_part = iso[:10]
        try:
            partes = data_part.split("-")
            if fmt == "mes":
                return f"{partes[1]}/{partes[0]}"
            else:
                return f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception:
            return iso

    def _local_prestacao(d):
        uf   = _str(d.get("uf", ""))
        xLP  = _str(d.get("xLP", ""))
        cLP  = _str(d.get("cLP", ""))
        xMun = _str(d.get("xMun", ""))
        nome_mun = (
            xLP
            or getattr(C, "IBGE_TO_NOME", {}).get(cLP, "")
            or xMun
        )
        if nome_mun and uf:
            return f"{nome_mun.upper()} - {uf.upper()}"
        return nome_mun.upper() if nome_mun else None

    def _cnae_desc(cnae9):
        desc = getattr(C, "CNAE9_TO_DESC", {}).get(cnae9, "")
        return f"{cnae9} - {desc}" if desc else cnae9

    def _extrair_fed(xml_path):
        """
        Extrai todas as retenções federais diretamente do XML.
        parse_nfse() não mapeia corretamente PIS/COFINS/IRRF/CSLL no modelo nacional.
        """
        import xml.etree.ElementTree as ET

        def _v(root, *tags):
            for tag in tags:
                el = next((e for e in root.iter() if e.tag.endswith(tag)), None)
                if el is not None and el.text:
                    try:
                        return float(el.text.strip())
                    except (ValueError, AttributeError):
                        pass
            return 0.0

        try:
            root = ET.parse(xml_path).getroot()
            return (
                _v(root, "vPis",    "vPIS"),
                _v(root, "vCofins", "vCOFINS"),
                _v(root, "vRetIRRF"),
                _v(root, "vRetCSLL"),
                _v(root, "vINSS"),
            )
        except Exception:
            return 0.0, 0.0, 0.0, 0.0, 0.0

    def _competencia_xml(xml_path):
        import xml.etree.ElementTree as ET
        try:
            root = ET.parse(xml_path).getroot()
            el = next((e for e in root.iter() if e.tag.endswith("dCompet")), None)
            if el is not None and el.text:
                p = el.text.strip()[:7].split("-")
                return f"{p[1]}/{p[0]}"
        except Exception:
            pass
        return ""

    with tempfile.TemporaryDirectory() as tmp:
        for uf_file in uploaded_files:
            uf_file.seek(0)
            with open(os.path.join(tmp, uf_file.name), "wb") as fh:
                fh.write(uf_file.read())

        arquivos = sorted(_glob.glob(os.path.join(tmp, "*.xml")))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços Tomados"

        header_fill = PatternFill(fill_type="solid", fgColor="C0C0C0")
        header_font = Font(bold=True, size=11)

        for col_idx, (titulo, largura) in enumerate(COLUNAS, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font  = header_font
            cell.fill  = header_fill
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = largura

        buf   = io.StringIO()
        total = 0
        erros = []
        n_arqs = max(len(arquivos), 1)

        # Barra de progresso no Streamlit
        _progress = st.progress(0, text="Iniciando processamento...")

        with contextlib.redirect_stdout(buf):
            for idx, xml_path in enumerate(arquivos):
                nome_arq = os.path.basename(xml_path)
                _progress.progress(
                    (idx + 1) / n_arqs,
                    text=f"Processando {idx + 1}/{len(arquivos)}: {nome_arq}",
                )
                try:
                    d = C.parse_nfse(xml_path)
                    if im:
                        d["im"] = im

                    # Extrai nNFSe diretamente do XML (número correto da nota)
                    import xml.etree.ElementTree as _ET2
                    try:
                        _root2 = _ET2.parse(xml_path).getroot()
                        _nNFSe = next((e.text for e in _root2.iter() if e.tag.endswith("nNFSe")), None)
                        if _nNFSe:
                            d["nNFSe"] = _nNFSe.strip()
                    except Exception:
                        pass

                    cnae9, item, aliq_cnae = C.resolver_cnae9(d)
                    dhEmi = _str(d.get("dhEmi", ""))

                    vS   = _float(d.get("vS"))
                    vISS = _float(d.get("vISS"))
                    vPIS, vCOFINS, vIRRF, vCSLL, vINSS = _extrair_fed(xml_path)
                    aliq   = _float(d.get("aliq")) or float(aliq_cnae or 0)
                    tpRet  = _str(d.get("tpRet", "1"))
                    iss_retido = (tpRet == "2")

                    ret_federais  = vPIS + vCOFINS + vCSLL

                    # Operação Simples Nacional — lê do XML
                    op_simp = _str(d.get("opSimpNac")) or ""
                    simples_nac = "Sim" if op_simp in ("1", "2") else "Não"

                    ws.append([
                        "NFS-e de Outro Município",          # [01] Tipo Doc.
                        _str(d.get("nNFSe") or d.get("nDFSe")), # [02] Número (nNFSe tem prioridade)
                        None,                                # [03] Código Verificação
                        _data_fmt(dhEmi, "mes"),            # [04] Competência
                        _data_fmt(dhEmi, "dia"),            # [05] Data
                        None,                                # [06] Vencimento
                        None,                                # [07] Número RPS
                        None,                                # [08] Série RPS
                        None,                                # [09] Tipo RPS
                        "Tributação Fora do Município",     # [10] Natureza da Operação
                        None,                                # [11] Regime Especial Tributação
                        simples_nac,                        # [12] Operação Simples Nacional
                        None,                               # [13] Incentivador Cultural
                        item,                               # [14] Item da Lista
                        _cnae_desc(cnae9),                  # [15] CNAE
                        None,                                # [16] ART
                        None,                                # [17] Código Obra
                        None,                                # [18] Número Empenho
                        _str(d.get("desc")),                # [19] Discriminação
                        vS,                                  # [20] Valor dos Serviços
                        None,                                # [21] Deduções
                        None,                                # [22] Desconto Condicionado
                        None,                                # [23] Desconto Incondicionado
                        ret_federais if ret_federais else 0.0, # [24] Retenções Federais
                        None,                                # [25] Outras Retenções
                        vPIS    if vPIS    else None,        # [26] PIS
                        vCOFINS if vCOFINS else None,        # [27] COFINS
                        vIRRF   if vIRRF   else None,        # [28] IRRF
                        vCSLL   if vCSLL   else None,        # [29] CSLL
                        vINSS   if vINSS   else None,        # [30] INSS
                        vS,                                  # [31] Base de Cálculo
                        aliq,                               # [32] Alíquota
                        _local_prestacao(d),                # [33] Local da Prestação
                        "Sim" if iss_retido else "Não",     # [34] ISS Retido
                        vISS if iss_retido else 0,          # [35] Valor do ISS (0 se não retido)
                        vS,                                  # [36] Valor Líquido
                        "NORMAL",                           # [37] Status Doc.
                        None,                                # [38] Inscrição Prestador
                        _str(d.get("cnpj")),                # [39] CPF/CNPJ Prestador
                        _str(d.get("nome")),                # [40] Razão Social
                        "Atual",                            # [41] Escrituração
                        "Prestador",                        # [42] Origem
                        "Não informada",                    # [43] Status Aceite
                    ])

                    # Formatos decimais — #,##0.00 em todos (igual ao "deu certo")
                    r = ws.max_row
                    for _col in [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 35, 36]:
                        ws.cell(row=r, column=_col).number_format = '#,##0.00'
                    ws.cell(row=r, column=32).number_format = '0.00'  # Alíquota

                    total += 1
                    print(f"  OK   {nome_arq}")
                    print(f"       NFSe {d.get('nDFSe','')} | {d.get('nome','')[:35]}")

                except Exception as exc:
                    erros.append((nome_arq, str(exc)))
                    print(f"  ERRO {nome_arq}: {exc}")

        _progress.empty()

        print(f"\n  Processadas: {total} nota(s)")
        if erros:
            print(f"  Com erro:    {len(erros)}")
            for n, e in erros:
                print(f"    - {n}: {e}")

        log   = buf.getvalue()
        saida = os.path.join(tmp, "resultado.xlsx")
        wb.save(saida)

        data = b""
        if os.path.exists(saida):
            with open(saida, "rb") as fh:
                data = fh.read()
            # Converte SharedStrings → InlineStr (compatibilidade Apache POI)
            data = _xlsx_para_inlinestr(data)

    return data, log
