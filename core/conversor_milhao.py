"""
core/conversor_milhao.py — Conversão CSV "Notas do Milhão" (portal ISS Fortaleza)
→ TXT ISS Fortaleza  e  → XLSX SPED GOV

Regras:
  - Lê o CSV exportado pelo portal ISS Fortaleza (formato antigo "Notas do Milhão")
  - Gera TXT com o mesmo layout de 46 campos usado por nfse_xml_to_txt.py
  - Gera XLSX com o mesmo layout SPED GOV usado por conversor_xlsx.py
  - Linhas com Situação = "C" (cancelada) são ignoradas
  - Linha de totais (Tipo != "2") é ignorada
  - NENHUMA linha de nfse_xml_to_txt.py é modificada — nunca modificar esse arquivo
"""

import io
import re
import csv as _csv
import unicodedata
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))


# ── Mapa UF por prefixo IBGE (2 primeiros dígitos) ───────────────────────────
_IBGE2_UF = {
    "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA",
    "16": "AP", "17": "TO", "21": "MA", "22": "PI", "23": "CE",
    "24": "RN", "25": "PB", "26": "PE", "27": "AL", "28": "SE",
    "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP",
    "41": "PR", "42": "SC", "43": "RS", "50": "MS", "51": "MT",
    "52": "GO", "53": "DF",
}

# ── Fallback IBGE para cidades mais comuns ────────────────────────────────────
_IBGE_FALLBACK = {
    ("SAO PAULO", "SP"):       "3550308",
    ("FORTALEZA", "CE"):       "2304400",
    ("RIO DE JANEIRO", "RJ"):  "3304557",
    ("BELO HORIZONTE", "MG"):  "3106200",
    ("BRASILIA", "DF"):        "5300108",
    ("CURITIBA", "PR"):        "4106902",
    ("PORTO ALEGRE", "RS"):    "4314902",
    ("SALVADOR", "BA"):        "2927408",
    ("RECIFE", "PE"):          "2611606",
    ("MANAUS", "AM"):          "1302603",
    ("HORIZONTE", "CE"):       "2305233",
    ("CAMPINAS", "SP"):        "3509502",
    ("GUARULHOS", "SP"):       "3518800",
    ("OSASCO", "SP"):          "3534401",
    ("SANTO ANDRE", "SP"):     "3547809",
    ("RIBEIRAO PRETO", "SP"):  "3543402",
    ("GOIANIA", "GO"):         "5208707",
    ("BELEM", "PA"):           "1501402",
    ("NATAL", "RN"):           "2408102",
    ("MACEIO", "AL"):          "2704302",
    ("JOAO PESSOA", "PB"):     "2507507",
    ("TERESINA", "PI"):        "2211001",
    ("CAMPO GRANDE", "MS"):    "5002704",
    ("CUIABA", "MT"):          "5103403",
    ("PORTO VELHO", "RO"):     "1100205",
    ("MACAPA", "AP"):          "1600303",
    ("BOA VISTA", "RR"):       "1400100",
    ("PALMAS", "TO"):          "1721000",
    ("RIO BRANCO", "AC"):      "1200401",
}

# ── Fallback CNAE por item LC 116 ─────────────────────────────────────────────
_CNAE_FALLBACK = {
    "1.01": "8599604", "1.02": "8599604", "1.03": "8599604", "1.04": "8599604",
    "1.05": "8599604", "1.06": "8591100", "1.07": "8599604",
    "2.01": "7020400",
    "3.02": "7111100", "3.03": "7111100", "3.04": "7111100", "3.05": "7111100",
    "4.01": "7020400", "4.02": "7020400", "4.03": "7020400",
    "5.01": "4520001", "5.02": "4520001", "5.03": "4520001", "5.04": "4520001",
    "6.01": "6201501", "6.02": "6201501", "6.03": "6201501", "6.04": "6201501",
    "6.05": "6201501", "6.29": "8299799",
    "7.01": "7410203", "7.02": "7410203", "7.03": "7490104", "7.04": "7490104",
    "8.01": "7490104", "8.02": "7490104",
    "9.01": "7320300",
    "10.01": "9200301", "10.02": "9200301", "10.03": "9200301", "10.04": "9200301",
    "10.05": "9200301", "10.06": "9200301", "10.07": "9200301", "10.08": "9200301",
    "10.09": "9200301", "10.10": "9200301",
    "11.01": "7911200", "11.02": "7990200", "11.03": "7990200", "11.04": "7990200",
    "12.01": "9001901", "12.02": "9001901", "12.03": "9001901", "12.04": "9001901",
    "12.05": "9001901", "12.06": "9001901", "12.07": "9001901", "12.08": "9001901",
    "12.09": "9001901", "12.10": "9001901", "12.11": "9001901", "12.12": "9001901",
    "12.13": "9001901", "12.14": "9001901", "12.15": "9001901", "12.16": "9001901",
    "12.17": "9001901",
    "13.01": "6612601", "13.02": "6612601", "13.03": "6612601", "13.04": "6612601",
    "13.05": "6612601",
    "14.01": "7319099",
    "15.01": "7410203", "15.02": "7410203", "15.03": "7410203", "15.04": "7410203",
    "15.05": "7410203", "15.06": "7410203", "15.07": "7410203", "15.08": "7410203",
    "15.09": "7410203", "15.10": "7410203", "15.11": "7410203", "15.12": "7410203",
    "15.13": "7410203", "15.14": "7410203", "15.15": "7410203", "15.16": "7410203",
    "15.17": "7410203", "15.18": "7410203",
    "16.01": "4399101",
    "17.01": "8299799", "17.02": "8299799", "17.03": "8299799", "17.04": "8299799",
    "17.05": "8299799", "17.06": "6311900", "17.07": "8299799", "17.08": "8299799",
    "17.09": "8299799", "17.10": "8299799", "17.11": "8299799", "17.12": "8299799",
    "17.13": "8299799", "17.14": "8299799", "17.15": "8299799", "17.16": "8299799",
    "17.17": "8299799", "17.18": "8299799", "17.19": "8299799",
    "18.01": "5811500",
    "19.01": "7490104",
    "20.01": "6612601", "20.02": "6612601", "20.03": "6612601",
    "21.01": "8711501",
    "22.01": "7500100",
    "23.01": "9609299",
    "24.01": "3812202",
    "25.01": "3600601", "25.02": "3702900", "25.03": "3811400", "25.04": "9000399",
    "26.01": "8121400",
    "27.01": "9609299",
    "28.01": "8220200",
    "29.01": "8021101",
    "30.01": "8111700",
    "31.01": "8112000", "31.02": "8112000",
    "32.01": "8599604", "32.02": "8599604", "32.03": "8599604",
    "32.04": "8599604", "32.05": "6619399",
    "33.01": "7500100", "33.02": "7500100",
    "34.01": "8591100",
    "35.01": "8011101", "35.02": "8011101", "35.03": "8011101",
    "36.01": "9609299",
    "37.01": "4711301",
    "38.01": "9609299",
    "39.01": "9609299",
    "40.01": "8591100",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _uf_de_ibge(ibge: str) -> str:
    digits = re.sub(r"[^0-9]", "", ibge or "")
    return _IBGE2_UF.get(digits[:2], "") if len(digits) >= 2 else ""


def _parse_brl(val: str) -> float:
    """'1.234,56' → 1234.56   |   '2,00' → 2.0"""
    v = (val or "").strip()
    if not v:
        return 0.0
    try:
        return float(v.replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def _centavos(val: str) -> str:
    """Valor BRL → inteiro centavos como string (vazio se zero)"""
    c = int(round(_parse_brl(val) * 100))
    return str(c) if c else ""


def _centesimos(val: str) -> str:
    """Alíquota % → centésimos como string.  '2,00' → '200'"""
    c = int(round(_parse_brl(val) * 100))
    return str(c) if c else ""


def _data_dma(val: str) -> str:
    """'28/05/2026 14:29:41' → '28/05/2026'"""
    return (val or "").strip()[:10]


def _mes(val: str) -> str:
    d = _data_dma(val)
    return d[3:5] if len(d) >= 5 else ""


def _ano(val: str) -> str:
    d = _data_dma(val)
    return d[6:10] if len(d) >= 10 else ""


def _comp_ma(val: str) -> str:
    """'28/05/2026 ...' → 'MM/AAAA'"""
    m, a = _mes(val), _ano(val)
    return f"{m}/{a}" if m and a else ""


def _strip_doc(doc: str) -> str:
    return re.sub(r"[^0-9]", "", doc or "")


def _normalizar_desc(desc: str) -> str:
    """Remove acentos e caracteres problemáticos para o TXT"""
    d = unicodedata.normalize("NFD", desc or "")
    d = "".join(c for c in d if ord(c) < 128)
    d = re.sub(r"[^A-Za-z0-9 /:.,@#%()!|-]", " ", d)
    d = re.sub(r" {2,}", " ", d).strip()
    return d


def _lgr(tipo: str, logr: str) -> str:
    t = (tipo or "").strip()
    lg = (logr or "").strip()
    return f"{t} {lg}".strip() if t else lg


def _item_de_cod(cod: str) -> str:
    """
    Converte código de serviço ISS para item LC 116.
    '3205' → '32.05'   '629' → '6.29'   '1401' → '14.01'
    """
    n = re.sub(r"[^0-9]", "", cod or "")
    if len(n) == 4:
        grp = str(int(n[:2]))
        sub = n[2:]
        return f"{grp}.{sub}"
    if len(n) == 3:
        return f"{n[0]}.{n[1:]}"
    return n


def _sem_acento(s: str) -> str:
    """Remove acentos e converte para maiúsculo para comparação."""
    return unicodedata.normalize("NFD", s or "").encode("ascii", "ignore").decode("ascii").upper().strip()


def _cnae9(item: str) -> str:
    """Lookup CNAE9 pelo item LC 116 — tenta nfse_xml_to_txt primeiro, depois fallback."""
    try:
        import nfse_xml_to_txt as C
        result = getattr(C, "ITEM_TO_CNAE9_DEFAULT", {}).get(item, "")
        if result:
            return result
    except Exception:
        pass
    return _CNAE_FALLBACK.get(item, "")


def _ibge_de_nome(cidade: str, uf: str) -> str:
    """Encontra o IBGE pelo nome da cidade + UF (sem acento, case-insensitive)."""
    cidade_n = _sem_acento(cidade)
    uf_u = (uf or "").upper().strip()
    # Fallback hardcoded primeiro (mais rápido)
    ibge = _IBGE_FALLBACK.get((cidade_n, uf_u), "")
    if ibge:
        return ibge
    # Tenta via tabela do nfse_xml_to_txt com normalização de acentos
    try:
        import nfse_xml_to_txt as C
        for ibge_key, nome in getattr(C, "IBGE_TO_NOME", {}).items():
            if _sem_acento(nome) == cidade_n and _uf_de_ibge(ibge_key) == uf_u:
                return ibge_key
    except Exception:
        pass
    return ""


def _resolver_ibge_prest(ibge_csv: str, cidade: str, uf: str) -> str:
    """
    Resolve o IBGE do município de prestação:
      1. Usa col [65] do CSV se for número válido != 0
      2. Tenta lookup IBGE_TO_NOME pelo nome da cidade + UF
      3. Retorna string vazia se não encontrar
    """
    ibge = re.sub(r"[^0-9]", "", ibge_csv or "")
    if ibge and ibge != "0":
        return ibge.zfill(7)
    # Fallback: lookup pelo nome
    encontrado = _ibge_de_nome(cidade, uf)
    return encontrado


def _local_prest(ibge: str, cidade: str, uf: str) -> str:
    """Retorna string 'CIDADE - UF' para o XLSX coluna Local da Prestação"""
    if ibge and ibge not in ("0", "0000000", ""):
        try:
            import nfse_xml_to_txt as C
            nome = getattr(C, "IBGE_TO_NOME", {}).get(ibge, "")
            if nome:
                uf2 = _uf_de_ibge(ibge)
                return f"{nome.upper()} - {uf2}" if uf2 else nome.upper()
        except Exception:
            pass
    if cidade and uf:
        return f"{cidade.upper()} - {uf.upper()}"
    return cidade.upper() if cidade else ""


def _ler_csv(conteudo: bytes) -> list:
    """
    Lê o CSV ISS Fortaleza.
    Retorna lista de listas (apenas linhas com Tipo de Registro == '2').
    Tenta latin-1 / cp1252 / utf-8 na decodificação.
    """
    texto = None
    for enc in ("latin-1", "cp1252", "utf-8", "utf-8-sig"):
        try:
            texto = conteudo.decode(enc)
            break
        except Exception:
            continue
    if texto is None:
        texto = conteudo.decode("latin-1", errors="replace")

    linhas = []
    reader = _csv.reader(io.StringIO(texto), delimiter=";")
    for row in reader:
        if not row:
            continue
        tipo = row[0].strip()
        if tipo == "2":          # apenas linhas de NFS-e
            linhas.append(row)
    return linhas


def _normalizar_linha(r: list) -> list:
    """Garante 73 colunas e remove espaços laterais"""
    r2 = [c.strip() for c in r]
    if len(r2) < 73:
        r2.extend([""] * (73 - len(r2)))
    return r2


# ── TXT ISS Fortaleza ─────────────────────────────────────────────────────────

def _gerar_linha_txt(r: list, im_padrao: str = "") -> str:
    """
    Converte uma linha do CSV para uma linha TXT de 46 campos.
    Layout idêntico ao gerado por nfse_xml_to_txt.gerar_linha().

    Índices CSV (0-based):
      [1]  N° NFS-e          [2]  Data Hora Emissão   [8]  IM Prestador
      [9]  Ind. CPF/CNPJ     [10] CPF/CNPJ Prestador  [11] Razão Social
      [12] Tipo Logradouro   [13] Logradouro           [14] Número
      [15] Complemento       [16] Bairro               [17] Cidade
      [18] UF                [19] CEP                  [20] E-mail
      [22] Situação (T/C)    [26] Valor Serviços       [28] Código Serviço
      [29] Alíquota %        [30] ISS Devido            [32] ISS Retido S/N
      [55] PIS               [56] COFINS               [57] INSS
      [59] CSLL              [65] IBGE Prestação        [72] Discriminação
    """
    c = _normalizar_linha(r)

    if c[22] == "C":           # cancelada → não gera linha
        return ""

    ind_pj = c[9]              # 1=CPF, 2=CNPJ
    cnpj_limpo = _strip_doc(c[10])

    # Campo [2] do TXT: "2" = PJ / "1" = PF (mesmo comportamento do _fix_linha)
    tipo_nota = "1" if ind_pj == "1" else "2"
    cnpj_campo = cnpj_limpo if ind_pj == "2" and len(cnpj_limpo) == 14 else ""

    uf_emit    = c[18]
    cep_emit   = re.sub(r"[^0-9]", "", c[19])
    cidade_emit = c[17]

    # IBGE do emitente (campo [8] do TXT): lookup pelo nome da cidade
    ibge_emit = _ibge_de_nome(cidade_emit, uf_emit)

    item  = _item_de_cod(c[28])
    cnae9 = _cnae9(item)

    data_emissao = _data_dma(c[2])
    m = _mes(c[2])
    a = _ano(c[2])

    iss_retido = c[32].upper() == "S"
    # Portal ISS Fortaleza: campo 21 = "1" retido, "2" a recolher
    tipo_recolh = "1" if iss_retido else "2"

    aliq_centesimos = _centesimos(c[29])    # campo 25 (índice 24) — alíquota

    desc = _normalizar_desc(c[72]).replace(";", ",")

    ibge_prest = _resolver_ibge_prest(c[65], cidade_emit, uf_emit)
    uf_prest   = _uf_de_ibge(ibge_prest) if ibge_prest else uf_emit
    natureza   = "1" if ibge_prest == "2304400" else "2"

    viss_centavos  = _centavos(c[30]) if iss_retido else ""
    vserv_centavos = _centavos(c[26])

    pis    = _centavos(c[55])
    cofins = _centavos(c[56])
    inss   = _centavos(c[57])
    csll   = _centavos(c[59])
    # Campo [40]: INSS se disponível, senão CSLL (mesmo padrão do _fix_linha)
    campo40 = inss or csll

    ind_ret = "1" if iss_retido else "0"
    im = im_padrao  # IM do tomador (usuário do sistema)

    campos = [
        "2.0",                  # [00] Versão layout
        "2",                    # [01] Tipo registro (2=nota tomada)
        tipo_nota,              # [02] Tipo nota/pessoa (2=PJ, 1=PF)
        cnpj_campo,             # [03] CNPJ emitente
        c[11],                  # [04] Razão social emitente
        "0",                    # [05] IM emitente (não usado no CSV)
        "1058",                 # [06] Código país emitente (1058=Brasil)
        uf_emit,                # [07] UF emitente
        ibge_emit,              # [08] IBGE município emitente
        cep_emit,               # [09] CEP
        _lgr(c[12], c[13]),     # [10] Logradouro
        c[14],                  # [11] Número
        c[15],                  # [12] Complemento
        c[16],                  # [13] Bairro
        "",                     # [14] Telefone (não disponível no CSV)
        c[20],                  # [15] E-mail emitente
        "7",                    # [16] Tipo documento (sempre 7 = NFSe)
        c[1],                   # [17] Número da nota fiscal
        "",                     # [18] Reservado
        data_emissao,           # [19] Data emissão (DD/MM/AAAA)
        tipo_recolh,            # [20] Tipo recolhimento (1=retido, 2=a recolher)
        m,                      # [21] Mês competência
        a,                      # [22] Ano competência
        cnae9,                  # [23] CNAE9 Fortaleza
        aliq_centesimos,        # [24] Alíquota ISS em centésimos
        desc,                   # [25] Discriminação do serviço
        "1058",                 # [26] Código país prestação
        uf_prest,               # [27] UF prestação
        ibge_prest,             # [28] IBGE município prestação
        natureza,               # [29] Natureza (1=no município, 2=fora, 3=isento)
        viss_centavos,          # [30] Valor ISS retido em centavos
        "",                     # [31] Reservado
        vserv_centavos,         # [32] Valor serviço em centavos
        "",                     # [33]
        "",                     # [34]
        "",                     # [35]
        "",                     # [36]
        "",                     # [37] Deduções
        pis,                    # [38] PIS
        cofins,                 # [39] COFINS
        campo40,                # [40] INSS / CSLL
        "",                     # [41]
        "",                     # [42]
        ind_ret,                # [43] Indicador retenção ISS (0=não, 1=sim)
        "",                     # [44]
        im,                     # [45] IM do tomador (usuário)
    ]

    return ";".join(campos)


def processar_csv_txt(conteudo: bytes, im: str = "") -> tuple:
    """
    CSV Notas do Milhão → TXT ISS Fortaleza.
    Retorna (bytes_txt, log_str).
    """
    log   = io.StringIO()
    rows  = _ler_csv(conteudo)
    linhas_txt = []
    ok = 0
    ignoradas = 0

    for i, r in enumerate(rows, 1):
        c = _normalizar_linha(r)
        situacao = c[22]
        nota = c[1] or str(i)
        prest = c[11][:35] if c[11] else ""

        if situacao == "C":
            ignoradas += 1
            log.write(f"  SKIP NFS-e {nota}: cancelada\n")
            continue

        try:
            linha = _gerar_linha_txt(r, im)
        except Exception as exc:
            ignoradas += 1
            log.write(f"  ERRO NFS-e {nota}: {exc}\n")
            continue

        if linha:
            linhas_txt.append(linha)
            ok += 1
            log.write(f"  OK   NFS-e {nota} | {prest}\n")
        else:
            ignoradas += 1
            log.write(f"  SKIP linha {i}: sem dados\n")

    log.write(f"\n  Processadas: {ok} nota(s)")
    if ignoradas:
        log.write(f"\n  Ignoradas:   {ignoradas}")

    txt_bytes = ("\n".join(linhas_txt) + ("\n" if linhas_txt else "")).encode("utf-8")
    return txt_bytes, log.getvalue()


# ── XLSX SPED GOV ─────────────────────────────────────────────────────────────

# Colunas idênticas às de conversor_xlsx.py — não modificar
_COLUNAS_XLSX = [
    ("Tipo Doc.",                       24.14),
    ("Número",                          10.28),
    ("Código de Verificação",           23.57),
    ("Competência",                     14.57),
    ("Data",                            11.42),
    ("Vencimento",                      13.42),
    ("Número RPS",                      18.42),
    ("Série RPS",                       11.14),
    ("Tipo RPS",                        10.28),
    ("Natureza da Operação",            27.71),
    ("Regime Especial Tributação",      52.14),
    ("Operação Simples Nacional",       29.28),
    ("Incentivador Cultural",           22.85),
    ("Item da Lista",                   14.57),
    ("CNAE",                           123.14),
    ("ART",                              5.28),
    ("Código Obra",                     13.85),
    ("Número Empenho",                  19.42),
    ("Discriminação dos Serviços",     141.14),
    ("Valor dos Serviços",              20.28),
    ("Deduções Permitidas em Lei",      30.42),
    ("Desconto Condicionado",           25.28),
    ("Desconto Incondicionado",         27.00),
    ("Retenções Federais",              21.28),
    ("Outras Retenções",               19.42),
    ("PIS",                              4.42),
    ("COFINS",                           8.85),
    ("IRRF",                             5.71),
    ("CSLL",                             6.14),
    ("INSS",                             5.85),
    ("Base de Cálculo",                 17.28),
    ("Alíquota",                         9.85),
    ("Local da Prestação",              22.14),
    ("ISS Retido",                      11.71),
    ("Valor do ISS",                    13.85),
    ("Valor Líquido",                   14.85),
    ("Status Doc.",                     13.00),
    ("Inscrição Prestador",             21.14),
    ("CPF/CNPJ Prestador",             21.42),
    ("Razão Social/Nome do Prestador",  58.57),
    ("Escrituração",                    13.85),
    ("Origem",                           9.71),
    ("Status Aceite",                   14.85),
]


def processar_csv_xlsx(conteudo: bytes) -> tuple:
    """
    CSV Notas do Milhão → XLSX SPED GOV.
    Layout idêntico ao gerado por conversor_xlsx.processar_xlsx_sped() para XMLs.
    Retorna (bytes_xlsx, log_str).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    import streamlit as st

    log  = io.StringIO()
    rows = _ler_csv(conteudo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serviços Tomados"

    header_fill = PatternFill(fill_type="solid", fgColor="C0C0C0")
    header_font = Font(bold=True, size=11)

    for col_idx, (titulo, largura) in enumerate(_COLUNAS_XLSX, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font  = header_font
        cell.fill  = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ok = 0
    ignoradas = 0
    n_total = max(len(rows), 1)
    _prog = st.progress(0, text="Iniciando processamento...")

    for idx, r in enumerate(rows):
        c = _normalizar_linha(r)
        nota  = c[1]
        prest = c[11][:35]
        _prog.progress(
            (idx + 1) / n_total,
            text=f"Processando {idx + 1}/{len(rows)}: NFS-e {nota}",
        )

        if c[22] == "C":
            ignoradas += 1
            log.write(f"  SKIP NFS-e {nota}: cancelada\n")
            continue

        try:
            iss_retido  = c[32].upper() == "S"
            vS          = _parse_brl(c[26])
            vISS        = _parse_brl(c[30]) if iss_retido else 0.0
            vDed        = _parse_brl(c[27])
            vPIS        = _parse_brl(c[55])
            vCOFINS     = _parse_brl(c[56])
            vIRRF       = _parse_brl(c[58])
            vCSLL       = _parse_brl(c[59])
            vINSS       = _parse_brl(c[57])
            aliq        = _parse_brl(c[29])   # % como float (ex: 2.0)
            ret_fed     = vPIS + vCOFINS + vCSLL

            item        = _item_de_cod(c[28])
            cnae9       = _cnae9(item)

            ibge_prest  = _resolver_ibge_prest(c[65], c[17], c[18])
            local_pr    = _local_prest(ibge_prest, c[17], c[18])

            simples_nac = "Sim" if c[21] not in ("0", "") else "Não"
            comp        = _comp_ma(c[2])
            data        = _data_dma(c[2])
            status_doc  = "NORMAL" if c[22] == "T" else c[22]

            ws.append([
                "NFS-e de Outro Município",    # Tipo Doc.
                c[1],                           # Número
                c[3],                           # Código de Verificação
                comp,                           # Competência MM/AAAA
                data,                           # Data DD/MM/AAAA
                "",                             # Vencimento
                c[6],                           # Número RPS
                c[5],                           # Série RPS
                c[4],                           # Tipo RPS
                "Tributação Fora do Município", # Natureza da Operação
                "",                             # Regime Especial Tributação
                simples_nac,                    # Operação Simples Nacional
                None,                           # Incentivador Cultural
                item,                           # Item da Lista
                cnae9,                          # CNAE
                "",                             # ART
                "",                             # Código Obra
                "",                             # Número Empenho
                c[72],                          # Discriminação dos Serviços
                vS,                             # Valor dos Serviços
                vDed if vDed else None,         # Deduções Permitidas em Lei
                None,                           # Desconto Condicionado
                None,                           # Desconto Incondicionado
                ret_fed if ret_fed else 0.0,    # Retenções Federais
                None,                           # Outras Retenções
                vPIS    if vPIS    else None,   # PIS
                vCOFINS if vCOFINS else None,   # COFINS
                vIRRF   if vIRRF   else None,   # IRRF
                vCSLL   if vCSLL   else None,   # CSLL
                vINSS   if vINSS   else None,   # INSS
                vS,                             # Base de Cálculo
                aliq,                           # Alíquota (%)
                local_pr if local_pr else None, # Local da Prestação
                "Sim" if iss_retido else "Não", # ISS Retido
                vISS if iss_retido else 0,      # Valor do ISS (0 se não retido)
                vS,                             # Valor Líquido
                status_doc,                     # Status Doc.
                c[8],                           # Inscrição Prestador
                c[10],                          # CPF/CNPJ Prestador
                c[11],                          # Razão Social/Nome do Prestador
                "Atual",                        # Escrituração
                "Prestador",                    # Origem
                c[66] or "Não informada",       # Status Aceite
            ])

            # Formatar células numéricas — mesmos índices de conversor_xlsx.py
            row_num = ws.max_row
            for col_mon in [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 35, 36]:
                ws.cell(row=row_num, column=col_mon).number_format = '#,##0.00'
            ws.cell(row=row_num, column=32).number_format = '#,##0.00'

            ok += 1
            log.write(f"  OK   NFS-e {nota} | {prest}\n")

        except Exception as exc:
            ignoradas += 1
            log.write(f"  ERRO NFS-e {nota}: {exc}\n")

    _prog.empty()
    log.write(f"\n  Processadas: {ok} nota(s)")
    if ignoradas:
        log.write(f"\n  Ignoradas:   {ignoradas}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), log.getvalue()
