"""
core/relatorio_deconf.py — Gera o relatório "PESQUISA SIGET - DECONF" em Excel,
a partir dos dados de índices de malha do SIGA (SEFAZ-CE).

Mapeamento de indicadores confirmado com o usuário:
  23 -> Entradas não escrituradas (NFe)
  24 -> Entradas não seladas (interestadual omissa no SITRAM)
  17 -> Quantidade de inventários anuais não entregue na EFD
  18 -> Declarações mensais omissas na EFD

Multa: 10% (não escrituradas), 20% (não seladas) sobre o valor total.
Grupos 3 e 4 (inventário/declarações) não têm base monetária conhecida
ainda -- ficam em branco para preenchimento manual.
"""
import io
from datetime import date

IND_NAO_ESCRITURADAS = "23"
IND_NAO_SELADAS = "24"
IND_INVENTARIO = "17"
IND_DECLARACAO_OMISSA = "18"

MULTA_NAO_ESCRITURADAS = 0.10
MULTA_NAO_SELADAS = 0.20


def _fmt_cnpj(cnpj: str) -> str:
    c = "".join(ch for ch in (cnpj or "") if ch.isdigit()).zfill(14)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"


def parse_indicadores_17_18(conteudo_xlsx: bytes) -> tuple[list[dict], list[dict]]:
    """
    Lê as abas "IND 17" e "IND 18" do relatório de índices de malha --
    têm colunas próprias (sem "chave"), por isso não entram no
    detalhes genérico de parse_indicadores_malha().
    Retorna (inventarios, declaracoes):
      inventarios: [{"ano": "2021", "data_declarado": ..., "codigo_motivo": ...,
                      "data_entrega_efd": ...}, ...]
      declaracoes: [{"data": "01/11/2025", "tipo_obrigacao": "EFD"}, ...]
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_xlsx), data_only=True)

    inventarios = []
    if "IND 17" in wb.sheetnames:
        rows = list(wb["IND 17"].iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            inventarios.append({
                "ano": str(row[0]).strip(),
                "data_declarado": str(row[1]) if len(row) > 1 and row[1] else "",
                "codigo_motivo": str(row[2]) if len(row) > 2 and row[2] else "",
                "data_entrega_efd": str(row[3]) if len(row) > 3 and row[3] else "",
            })

    declaracoes = []
    if "IND 18" in wb.sheetnames:
        rows = list(wb["IND 18"].iter_rows(values_only=True))
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            declaracoes.append({
                "data": str(row[0]).strip(),
                "tipo_obrigacao": str(row[1]) if len(row) > 1 and row[1] else "",
            })

    return inventarios, declaracoes


def _ano_de(data_str: str) -> str:
    """Extrai o ano de uma data 'DD/MM/AAAA' ou 'AAAA-MM-DD'."""
    data_str = (data_str or "").strip()
    if "/" in data_str:
        partes = data_str.split("/")
        if len(partes) == 3:
            return partes[2][:4]
    if "-" in data_str:
        return data_str[:4]
    return ""


def gerar_relatorio_deconf(
    cnpj: str,
    razao_social: str,
    resumo: list[dict],
    detalhes: dict[str, list[dict]],
    inventarios: list[dict],
    declaracoes: list[dict],
    tipo_estabelecimento: str = "MATRIZ",
    ano_inicial: int | None = None,
) -> bytes:
    """Monta o Excel no padrão "PESQUISA SIGET - DECONF"."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hoje = date.today()
    if ano_inicial is None:
        ano_inicial = hoje.year - 5  # janela padrão de 5 anos (decadência)
    anos = list(range(ano_inicial, hoje.year + 1))

    # ── Agrega por ano ───────────────────────────────────────────────────
    def _agrega_por_ano(itens: list[dict]) -> dict[str, tuple[int, float]]:
        agregados: dict[str, list] = {}
        for item in itens:
            ano = _ano_de(item.get("data_emissao", ""))
            if ano not in agregados:
                agregados[ano] = [0, 0.0]
            agregados[ano][0] += 1
            agregados[ano][1] += float(item.get("valor", 0) or 0)
        return {ano: (qtd, valor) for ano, (qtd, valor) in agregados.items()}

    nao_escrituradas_ano = _agrega_por_ano(detalhes.get(IND_NAO_ESCRITURADAS, []))
    nao_seladas_ano = _agrega_por_ano(detalhes.get(IND_NAO_SELADAS, []))

    inventario_por_ano: dict[str, int] = {}
    for inv in inventarios:
        ano = inv.get("ano", "")
        inventario_por_ano[ano] = inventario_por_ano.get(ano, 0) + 1

    declaracao_por_ano: dict[str, list] = {}
    for d in declaracoes:
        ano = _ano_de(d.get("data", ""))
        declaracao_por_ano.setdefault(ano, []).append(d.get("data", ""))

    # ── Workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Malhas SIGA"

    VERDE = PatternFill("solid", fgColor="1E3A2E")
    VERDE_FONT = Font(bold=True, color="FFFFFF", size=10)
    AMARELO = PatternFill("solid", fgColor="FFFF00")
    CINZA = PatternFill("solid", fgColor="D9D9D9")
    BORDA = Border(*(Side(style="thin", color="808080"),) * 4)
    CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ultima_col = 9  # A..I (3 colunas do grupo 1 + 2+2+2 dos demais grupos)

    def _titulo(linha, texto, negrito=True, tamanho=11):
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=ultima_col)
        cel = ws.cell(row=linha, column=1, value=texto)
        cel.font = Font(bold=negrito, size=tamanho)
        cel.alignment = Alignment(horizontal="center")

    _titulo(1, "RELATÓRIO DE MALHAS SIGA", tamanho=13)

    ws.cell(row=3, column=1, value=f"RAZÃO SOCIAL: {razao_social}").font = Font(bold=True)
    ws.cell(row=4, column=1, value=f"CNPJ: {_fmt_cnpj(cnpj)}").font = Font(bold=True)

    linha_matriz = 6
    ws.merge_cells(start_row=linha_matriz, start_column=1, end_row=linha_matriz, end_column=ultima_col)
    cel_matriz = ws.cell(row=linha_matriz, column=1, value=tipo_estabelecimento)
    cel_matriz.fill = VERDE
    cel_matriz.font = VERDE_FONT
    cel_matriz.alignment = CENTRO

    linha_grupo = linha_matriz + 1
    grupos = [
        ("ENTRADAS NÃO ESCRITURADAS (NFe)", 3),
        ("ENTRADAS NÃO SELADAS", 2),
        ("QUANTIDADE DE INVENTÁRIOS ANUAIS NÃO ENTREGUE NA EFD", 2),
        ("DECLARAÇÕES MENSAIS OMISSAS NA EFD", 2),
    ]
    col = 1
    for nome, largura in grupos:
        ws.merge_cells(start_row=linha_grupo, start_column=col, end_row=linha_grupo, end_column=col + largura - 1)
        cel = ws.cell(row=linha_grupo, column=col, value=nome)
        cel.fill = VERDE
        cel.font = VERDE_FONT
        cel.alignment = CENTRO
        col += largura

    linha_sub = linha_grupo + 1
    cabecalhos = [
        "ANO", "QUANTIDADE", "VALOR",
        "QUANTIDADE", "VALOR",
        "QUANTIDADE", "ANO",
        "QUANTIDADE", "PERIODO",
    ]
    ultima_col = len(cabecalhos)
    for i, h in enumerate(cabecalhos, 1):
        cel = ws.cell(row=linha_sub, column=i, value=h)
        cel.fill = CINZA
        cel.font = Font(bold=True, size=9)
        cel.alignment = CENTRO
        cel.border = BORDA

    linha_dados_ini = linha_sub + 1
    total_qtd_esc = total_val_esc = 0
    total_qtd_sel = total_val_sel = 0
    total_qtd_inv = 0
    total_qtd_decl = 0

    for i, ano in enumerate(anos):
        linha = linha_dados_ini + i
        ano_str = str(ano)

        qtd_esc, val_esc = nao_escrituradas_ano.get(ano_str, (0, 0.0))
        qtd_sel, val_sel = nao_seladas_ano.get(ano_str, (0, 0.0))
        qtd_inv = inventario_por_ano.get(ano_str, 0)
        periodos_decl = declaracao_por_ano.get(ano_str, [])
        qtd_decl = len(periodos_decl)

        valores_linha = [
            ano_str, qtd_esc or "", val_esc or "",
            qtd_sel or "", val_sel or "",
            qtd_inv or "", (ano_str if qtd_inv else ""),
            qtd_decl or "", (", ".join(periodos_decl) if periodos_decl else ""),
        ]
        for c, v in enumerate(valores_linha, 1):
            cel = ws.cell(row=linha, column=c, value=v)
            cel.border = BORDA
            cel.alignment = Alignment(horizontal="center")
            if c in (3, 5) and isinstance(v, (int, float)):
                cel.number_format = '"R$" #,##0.00'

        total_qtd_esc += qtd_esc
        total_val_esc += val_esc
        total_qtd_sel += qtd_sel
        total_val_sel += val_sel
        total_qtd_inv += qtd_inv
        total_qtd_decl += qtd_decl

    linha_total = linha_dados_ini + len(anos)
    ws.cell(row=linha_total, column=1, value="TOTAL").font = Font(bold=True)
    totais = [
        "", total_qtd_esc, total_val_esc,
        total_qtd_sel, total_val_sel,
        total_qtd_inv, "",
        total_qtd_decl, "",
    ]
    for c, v in enumerate(totais, 1):
        if c == 1:
            continue
        cel = ws.cell(row=linha_total, column=c, value=v if v != "" else None)
        cel.fill = CINZA
        cel.font = Font(bold=True)
        cel.border = BORDA
        cel.alignment = Alignment(horizontal="center")
        if c in (3, 5) and isinstance(v, (int, float)):
            cel.number_format = '"R$" #,##0.00'

    linha_multa_lbl = linha_total + 1
    linha_multa_val = linha_total + 2
    ws.cell(row=linha_multa_lbl, column=1, value="MULTA- NÃO ESCRITA")
    ws.cell(row=linha_multa_lbl, column=1).font = Font(bold=True)

    multa_esc = total_val_esc * MULTA_NAO_ESCRITURADAS
    multa_sel = total_val_sel * MULTA_NAO_SELADAS

    celulas_multa = {
        2: ("10%", None), 3: (None, multa_esc),
        4: ("20%", None), 5: (None, multa_sel),
        6: ("10%", None), 7: (None, None),
        8: ("10%", None), 9: (None, None),
    }
    for c, (label_pct, valor) in celulas_multa.items():
        if label_pct is not None:
            cel = ws.cell(row=linha_multa_lbl, column=c, value=label_pct)
            cel.fill = AMARELO
            cel.font = Font(bold=True)
            cel.alignment = Alignment(horizontal="center")
        if valor is not None:
            cel2 = ws.cell(row=linha_multa_val, column=c, value=valor)
            cel2.fill = AMARELO
            cel2.font = Font(bold=True)
            cel2.number_format = '"R$" #,##0.00'
            cel2.alignment = Alignment(horizontal="center")

    larguras = [10, 12, 14, 12, 14, 12, 10, 12, 14]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
