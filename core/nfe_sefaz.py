"""
core/nfe_sefaz.py
Consulta NFeDistribuicaoDFe da SEFAZ Nacional.
Aceita certificado como bytes (vindo do banco), sem gravar em disco.
"""

import gzip
import base64
import time
import zipfile
import io
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from typing import Callable

import requests
from requests_pkcs12 import Pkcs12Adapter

NS = "http://www.portalfiscal.inf.br/nfe"


class _RateLimitError(Exception):
    """cStat=656 — consumo indevido, aguardar 1 hora."""
    pass

UF_CODIGOS = {
    "AC": 12, "AL": 27, "AP": 16, "AM": 13, "BA": 29,
    "CE": 23, "DF": 53, "ES": 32, "GO": 52, "MA": 21,
    "MT": 51, "MS": 50, "MG": 31, "PA": 15, "PB": 25,
    "PR": 41, "PE": 26, "PI": 22, "RJ": 33, "RN": 24,
    "RS": 43, "RO": 11, "RR": 14, "SC": 42, "SP": 35,
    "SE": 28, "TO": 17,
}

URL_SEFAZ = "https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx"
SOAP_ACTION = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse"

_ENVELOPE_NSU = """<?xml version="1.0" encoding="UTF-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">
      <nfeDadosMsg>
        <distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">
          <tpAmb>{amb}</tpAmb>
          <cUFAutor>{cuf}</cUFAutor>
          <CNPJ>{cnpj}</CNPJ>
          <distNSU><ultNSU>{nsu}</ultNSU></distNSU>
        </distDFeInt>
      </nfeDadosMsg>
    </nfeDistDFeInteresse>
  </soap12:Body>
</soap12:Envelope>"""

_ENVELOPE_CHAVE = """<?xml version="1.0" encoding="UTF-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">
      <nfeDadosMsg>
        <distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">
          <tpAmb>{amb}</tpAmb>
          <cUFAutor>{cuf}</cUFAutor>
          <CNPJ>{cnpj}</CNPJ>
          <consChNFe><chNFe>{chave}</chNFe></consChNFe>
        </distDFeInt>
      </nfeDadosMsg>
    </nfeDistDFeInteresse>
  </soap12:Body>
</soap12:Envelope>"""

_HEADERS = {
    "Content-Type": "application/soap+xml; charset=utf-8",
    "SOAPAction": SOAP_ACTION,
}


def _sessao(pfx_bytes: bytes, pfx_senha: str) -> requests.Session:
    s = requests.Session()
    s.mount("https://", Pkcs12Adapter(
        pkcs12_data=pfx_bytes,
        pkcs12_password=pfx_senha.encode() if isinstance(pfx_senha, str) else pfx_senha,
    ))
    return s


def _texto(el) -> str:
    return el.text.strip() if el is not None and el.text else ""


def _descompactar_docs(xml_resposta: str) -> list[dict]:
    docs = []
    try:
        root = ET.fromstring(xml_resposta)
        lote = root.find(".//{http://www.portalfiscal.inf.br/nfe}loteDistDFeInt")
        if lote is None:
            return docs
        for doc in lote.findall("{http://www.portalfiscal.inf.br/nfe}docZip"):
            nsu = doc.get("NSU", "")
            schema = doc.get("schema", "")
            try:
                xml = gzip.decompress(base64.b64decode(doc.text)).decode("utf-8")
            except Exception:
                continue
            schema_lower = schema.lower()
            # resNFe_v1.01.xsd  → resumo NF-e
            # resNFCe_v1.00.xsd → resumo NFC-e  ("resNFe" NÃO é substring de "resNFCe")
            resumo = "resnfe" in schema_lower or "resnfce" in schema_lower
            modelo = "NFC-e" if ("nfce" in schema_lower) else "NF-e"
            docs.append({"nsu": nsu, "schema": schema, "xml": xml,
                         "resumo": resumo, "modelo": modelo})
    except Exception:
        pass
    return docs


def _extrair_chave_resumo(xml: str) -> str:
    try:
        root = ET.fromstring(xml)
        ns_nfe = "http://www.portalfiscal.inf.br/nfe"
        # resNFCe pode usar chNFCe; resNFe usa chNFe — testa ambos
        for tag in ("chNFe", "chNFCe"):
            el = root.find(f".//{{{ns_nfe}}}{tag}")
            if el is None:
                el = root.find(f".//{tag}")
            if el is not None and el.text:
                return el.text.strip()
        return ""
    except Exception:
        return ""


def _extrair_dados_resumo(xml_resumo: str, cnpj_consultado: str) -> dict:
    """Extrai dados básicos de um resNFe/resNFCe quando o XML completo não está disponível."""
    d = {
        "chave": "", "numero": "", "serie": "", "data_emissao": "",
        "cnpj_emitente": "", "nome_emitente": "",
        "cnpj_dest_doc": cnpj_consultado, "nome_dest_doc": "",
        "valor_total": 0.0, "cfop": "", "nat_operacao": "",
        "papel": "Recebida", "modelo": "NF-e",
    }
    try:
        root = ET.fromstring(xml_resumo)
        ns = {"n": NS}

        def _find(tag):
            el = root.find(f".//n:{tag}", ns)
            return el if el is not None else root.find(f".//{tag}")

        chave_el  = _find("chNFe") or _find("chNFCe")
        cnpj_e_el = _find("CNPJ")
        nome_e_el = _find("xNome")
        dh_emi_el = _find("dhEmi") or _find("dEmi")
        vnf_el    = _find("vNF")
        tp_nf_el  = _find("tpNF")
        mod_el    = _find("mod")

        if chave_el is not None and chave_el.text:
            d["chave"] = chave_el.text.strip()
        if cnpj_e_el is not None and cnpj_e_el.text:
            d["cnpj_emitente"] = cnpj_e_el.text.strip()
            if cnpj_e_el.text.strip() == cnpj_consultado:
                d["papel"] = "Emitida"
                d["cnpj_dest_doc"] = ""
        if nome_e_el is not None:
            d["nome_emitente"] = _texto(nome_e_el)
        if dh_emi_el is not None and dh_emi_el.text:
            d["data_emissao"] = dh_emi_el.text[:10]
        if vnf_el is not None and vnf_el.text:
            try:
                d["valor_total"] = float(vnf_el.text)
            except Exception:
                pass
        if tp_nf_el is not None:
            d["papel"] = "Emitida" if tp_nf_el.text == "1" else "Recebida"
        if mod_el is not None and mod_el.text == "65":
            d["modelo"] = "NFC-e"
    except Exception:
        pass
    return d


def _extrair_dados(xml_nf: str, cnpj_consultado: str) -> dict:
    d = {
        "chave": "", "numero": "", "serie": "", "data_emissao": "",
        "cnpj_emitente": "", "nome_emitente": "",
        "cnpj_dest_doc": "", "nome_dest_doc": "",
        "valor_total": 0.0, "cfop": "", "nat_operacao": "",
        "papel": "Recebida", "modelo": "NF-e",
    }
    try:
        root = ET.fromstring(xml_nf)
        ns = {"n": NS}
        inf = root.find(".//n:infNFe", ns)
        ide = root.find(".//n:ide", ns)
        emit = root.find(".//n:emit", ns)
        dest = root.find(".//n:dest", ns)
        vNF = root.find(".//n:vNF", ns)

        if inf is not None:
            d["chave"] = inf.get("Id", "").replace("NFe", "").replace("NFCe", "")
        if ide is not None:
            d["numero"] = _texto(ide.find("n:nNF", ns))
            d["serie"] = _texto(ide.find("n:serie", ns))
            d["nat_operacao"] = _texto(ide.find("n:natOp", ns))
            dt = _texto(ide.find("n:dhEmi", ns)) or _texto(ide.find("n:dEmi", ns))
            d["data_emissao"] = dt[:10] if dt else ""
            if _texto(ide.find("n:mod", ns)) == "65":
                d["modelo"] = "NFC-e"
        if emit is not None:
            cnpj_e = _texto(emit.find("n:CNPJ", ns))
            d["cnpj_emitente"] = cnpj_e
            d["nome_emitente"] = _texto(emit.find("n:xNome", ns))
            if cnpj_e == cnpj_consultado:
                d["papel"] = "Emitida"
        if dest is not None:
            d["cnpj_dest_doc"] = _texto(dest.find("n:CNPJ", ns))
            d["nome_dest_doc"] = _texto(dest.find("n:xNome", ns))
        if vNF is not None:
            d["valor_total"] = float(vNF.text or 0)
        else:
            # fallback sem namespace (alguns XMLs SVRS omitem o prefixo)
            vNF2 = root.find(".//vNF")
            if vNF2 is not None:
                d["valor_total"] = float(vNF2.text or 0)
        cfop = root.find(".//n:CFOP", ns)
        if cfop is not None:
            d["cfop"] = _texto(cfop)
    except Exception:
        pass
    return d


def _consultar_nsu(sessao, cnpj: str, nsu: str, amb: str, cuf: int) -> tuple[str, list[dict]]:
    env = _ENVELOPE_NSU.format(amb=amb, cuf=cuf, cnpj=cnpj, nsu=nsu)
    try:
        r = sessao.post(URL_SEFAZ, data=env.encode("utf-8"), headers=_HEADERS, timeout=60)
        r.raise_for_status()
        root = ET.fromstring(r.text)
        max_nsu_el = root.find(".//{http://www.portalfiscal.inf.br/nfe}maxNSU")
        novo_nsu = max_nsu_el.text if max_nsu_el is not None else nsu
        cstat_el = root.find(".//{http://www.portalfiscal.inf.br/nfe}cStat")
        cstat = cstat_el.text if cstat_el is not None else "999"
        if cstat not in ("137", "138"):
            motivo_el = root.find(".//{http://www.portalfiscal.inf.br/nfe}xMotivo")
            motivo = motivo_el.text if motivo_el is not None else "?"
            # Retorna novo_nsu para que o caller possa salvar o ultNSU da resposta
            return novo_nsu, [{"_erro": f"cStat={cstat}: {motivo}"}]
        return novo_nsu, _descompactar_docs(r.text)
    except Exception as e:
        return nsu, [{"_erro": str(e)}]


def _baixar_por_chave(sessao, cnpj: str, chave: str, amb: str, cuf: int) -> str | None:
    """Baixa XML completo por chave. Levanta _RateLimitError se cStat=656."""
    env = _ENVELOPE_CHAVE.format(amb=amb, cuf=cuf, cnpj=cnpj, chave=chave)
    try:
        r = sessao.post(URL_SEFAZ, data=env.encode("utf-8"), headers=_HEADERS, timeout=60)
        r.raise_for_status()
        root = ET.fromstring(r.text)
        cstat_el = root.find(f".//{{{NS}}}cStat")
        cstat = cstat_el.text.strip() if cstat_el is not None else "?"
        if cstat == "656":
            raise _RateLimitError()
        docs = _descompactar_docs(r.text)
        return docs[0]["xml"] if docs else None
    except _RateLimitError:
        raise
    except Exception:
        return None


def executar_consulta_sefaz(
    pfx_bytes: bytes,
    pfx_senha: str,
    empresas: list[dict],
    ambiente: str = "1",
    uf: str = "CE",
    data_ini=None,
    data_fim=None,
    tipo_doc: str = "ambos",
    papel_filtro: str = "ambos",
    incluir_xml: bool = True,
    incluir_pdf: bool = False,
    incluir_excel: bool = True,
    log_cb: Callable[[str], None] | None = None,
    progress_cb: Callable[[float], None] | None = None,
    salvar_db: bool = False,
    username: str = "",
) -> tuple[bytes | None, list[str]]:
    """
    Consulta a SEFAZ para cada empresa e retorna um ZIP com XMLs, PDFs e/ou Excel.

    tipo_doc: 'ambos' | 'nfe' | 'nfce'
    papel_filtro: 'ambos' | 'emitidas' | 'recebidas'
    """
    from datetime import date as _date
    log: list[str] = []
    todos_docs: list[dict] = []
    docs_para_db: list[dict] = []
    chaves_retry: dict[str, list[str]] = {}  # cnpj -> [chaves sem XML completo]

    # Deduplica CNPJs
    cnpjs_vistos: set[str] = set()
    empresas_unicas = []
    for emp in empresas:
        cnpj_n = "".join(c for c in emp["cnpj"] if c.isdigit())
        if cnpj_n not in cnpjs_vistos:
            cnpjs_vistos.add(cnpj_n)
            empresas_unicas.append({**emp, "cnpj": cnpj_n})

    def _log(msg: str):
        log.append(msg)
        if log_cb:
            log_cb(msg)

    def _dentro_periodo(data_emissao: str) -> bool:
        if not data_ini and not data_fim:
            return True
        if not data_emissao:
            return True
        try:
            dt = _date.fromisoformat(data_emissao[:10])
            if data_ini and dt < data_ini:
                return False
            if data_fim and dt > data_fim:
                return False
            return True
        except Exception:
            return True

    def _passa_filtros(doc: dict) -> bool:
        if tipo_doc == "nfe" and doc.get("modelo") != "NF-e":
            return False
        if tipo_doc == "nfce" and doc.get("modelo") != "NFC-e":
            return False
        if papel_filtro == "emitidas" and doc.get("papel") != "Emitida":
            return False
        if papel_filtro == "recebidas" and doc.get("papel") != "Recebida":
            return False
        if not _dentro_periodo(doc.get("data_emissao", "")):
            return False
        return True

    cuf = UF_CODIGOS.get(uf, 23)
    sessao = _sessao(pfx_bytes, pfx_senha)

    periodo_txt = ""
    if data_ini or data_fim:
        ini_str = data_ini.strftime("%d/%m/%Y") if data_ini else "início"
        fim_str = data_fim.strftime("%d/%m/%Y") if data_fim else "hoje"
        periodo_txt = f" | Período: {ini_str} a {fim_str}"

    for idx, empresa in enumerate(empresas_unicas):
        cnpj = empresa["cnpj"]
        nome = empresa.get("nome", cnpj)

        _log(f"[{idx+1}/{len(empresas_unicas)}] Consultando {nome} ({cnpj}){periodo_txt}...")
        if progress_cb:
            progress_cb(idx / len(empresas_unicas) * 0.8)

        from db.database import get_nsu_cnpj, set_nsu_cnpj
        estado = get_nsu_cnpj(cnpj)
        nsu = estado["ultimo_nsu"]
        ultima_consulta = estado.get("atualizado_em") or "nunca"
        _log(f"  Último NSU salvo: {nsu} (consulta anterior: {ultima_consulta})")

        paginas = 0
        docs_empresa = 0
        proxima_dt: datetime | None = None
        chave_rate_limited = False
        resumos_sem_xml = 0

        while paginas < 200:
            novo_nsu, docs = _consultar_nsu(sessao, cnpj, nsu, ambiente, cuf)
            paginas += 1

            if not docs:
                if novo_nsu != nsu:
                    set_nsu_cnpj(cnpj, novo_nsu)
                    _log(f"  Sem novos documentos. NSU avançado para {novo_nsu}.")
                else:
                    _log(f"  Sem novos documentos.")
                proxima_dt = datetime.now() + timedelta(minutes=55)
                break

            if "_erro" in docs[0]:
                erro_msg = docs[0]["_erro"]
                if novo_nsu != nsu:
                    set_nsu_cnpj(cnpj, novo_nsu)
                    _log(f"  NSU salvo da resposta de erro: {novo_nsu}")
                if "656" in erro_msg:
                    proxima_dt = datetime.now() + timedelta(minutes=61)
                    _log(f"  RATE LIMIT (656) no distNSU — cooldown de 61 min.")
                else:
                    proxima_dt = datetime.now() + timedelta(minutes=55)
                    _log(f"  AVISO SEFAZ: {erro_msg}")
                break

            for doc in docs:
                xml_conteudo = None
                modelo = doc.get("modelo", "NF-e")
                _log(f"  NSU={doc['nsu']} schema={doc['schema']} resumo={doc['resumo']} modelo={modelo}")

                if doc.get("resumo"):
                    chave = _extrair_chave_resumo(doc["xml"])
                    if chave and not chave_rate_limited:
                        try:
                            xml_conteudo = _baixar_por_chave(sessao, cnpj, chave, ambiente, cuf)
                        except _RateLimitError:
                            _log(f"  RATE LIMIT (656) na consulta por chave — salvando resumos restantes sem XML completo")
                            chave_rate_limited = True
                            xml_conteudo = None
                        time.sleep(0.5)
                    if not xml_conteudo:
                        resumos_sem_xml += 1
                        d_res = _extrair_dados_resumo(doc["xml"], cnpj)
                        chave_resumo = d_res.get("chave", "")
                        if salvar_db and chave_resumo:
                            docs_para_db.append({
                                "cnpj_empresa":  cnpj,
                                "chave":         chave_resumo,
                                "modelo":        d_res.get("modelo", "NF-e"),
                                "papel":         d_res.get("papel", "Recebida"),
                                "numero":        "",
                                "serie":         "",
                                "data_emissao":  d_res.get("data_emissao", ""),
                                "cnpj_emitente": d_res.get("cnpj_emitente", ""),
                                "nome_emitente": d_res.get("nome_emitente", ""),
                                "cnpj_dest_doc": d_res.get("cnpj_dest_doc", ""),
                                "nome_dest_doc": d_res.get("nome_dest_doc", ""),
                                "valor_total":   float(d_res.get("valor_total", 0) or 0),
                                "nat_operacao":  "",
                                "xml":           doc["xml"],
                            })
                        if chave_resumo:
                            chaves_retry.setdefault(cnpj, []).append(chave_resumo)
                        continue
                else:
                    xml_conteudo = doc["xml"]

                if not xml_conteudo:
                    continue

                dados = _extrair_dados(xml_conteudo, cnpj)
                # Prefere o modelo lido do XML (<mod>65</mod>) ao detectado pelo schema
                modelo_final = dados.get("modelo") or modelo
                dados.update({
                    "xml": xml_conteudo,
                    "cnpj_empresa": cnpj,
                    "nome_empresa": nome,
                    "nsu": doc["nsu"],
                    "modelo": modelo_final,
                })

                # Salva TODOS os docs no acervo (sem filtro de período/tipo)
                if salvar_db:
                    docs_para_db.append({
                        "cnpj_empresa":  cnpj,
                        "chave":         dados.get("chave", ""),
                        "modelo":        modelo_final,
                        "papel":         dados.get("papel", "Recebida"),
                        "numero":        dados.get("numero", ""),
                        "serie":         dados.get("serie", ""),
                        "data_emissao":  dados.get("data_emissao", ""),
                        "cnpj_emitente": dados.get("cnpj_emitente", ""),
                        "nome_emitente": dados.get("nome_emitente", ""),
                        "cnpj_dest_doc": dados.get("cnpj_dest_doc", ""),
                        "nome_dest_doc": dados.get("nome_dest_doc", ""),
                        "valor_total":   float(dados.get("valor_total", 0) or 0),
                        "nat_operacao":  dados.get("nat_operacao", ""),
                        "xml":           xml_conteudo,
                    })

                if not _passa_filtros(dados):
                    continue

                todos_docs.append(dados)
                docs_empresa += 1
                _log(f"  [{dados['modelo']}][{dados['papel']}] "
                     f"{dados.get('data_emissao','?')} "
                     f"{dados['chave'][:20]}... R${dados['valor_total']:.2f}")

            nsu = novo_nsu
            set_nsu_cnpj(cnpj, novo_nsu)
            if len(docs) < 50:
                proxima_dt = datetime.now() + timedelta(minutes=55)
                break
            time.sleep(2)

        if proxima_dt:
            from db.database import set_proxima_consulta
            set_proxima_consulta(cnpj, proxima_dt.strftime("%Y-%m-%dT%H:%M:%S"))
            _log(f"  Próxima consulta liberada em: {proxima_dt.strftime('%d/%m/%Y às %H:%M:%S')}")

        # ── Manifestação automática + retry consChNFe ────────────────────
        chaves_pendentes = chaves_retry.get(cnpj, [])
        nfe_pendentes = [c for c in chaves_pendentes if c[20:22] != "65"]
        if nfe_pendentes and not chave_rate_limited:
            _log(f"  Manifestando Ciência da Operação para {len(nfe_pendentes)} NF-e sem XML...")
            manifestadas = 0
            for ch in nfe_pendentes:
                try:
                    res_m = manifestar_ciencia(pfx_bytes, pfx_senha, cnpj, ch, ambiente)
                    if res_m.get("ok"):
                        manifestadas += 1
                    elif "573" in res_m.get("erro", "") or "135" in res_m.get("cStat", ""):
                        manifestadas += 1
                except Exception:
                    pass
                time.sleep(0.5)
            _log(f"  {manifestadas}/{len(nfe_pendentes)} manifestação(ões) aceita(s)")

            if manifestadas:
                _log(f"  Aguardando 5s para propagação antes do retry consChNFe...")
                time.sleep(5)
                recuperadas = 0
                for ch in nfe_pendentes:
                    try:
                        xml_ret = _baixar_por_chave(sessao, cnpj, ch, ambiente, cuf)
                    except _RateLimitError:
                        _log(f"  RATE LIMIT (656) no retry — interrompendo")
                        break
                    except Exception:
                        xml_ret = None
                    if xml_ret:
                        recuperadas += 1
                        dados = _extrair_dados(xml_ret, cnpj)
                        modelo_final = dados.get("modelo") or "NF-e"
                        dados.update({
                            "xml": xml_ret, "cnpj_empresa": cnpj,
                            "nome_empresa": nome, "nsu": "", "modelo": modelo_final,
                        })
                        if salvar_db:
                            docs_para_db.append({
                                "cnpj_empresa":  cnpj,
                                "chave":         dados.get("chave", ""),
                                "modelo":        modelo_final,
                                "papel":         dados.get("papel", "Recebida"),
                                "numero":        dados.get("numero", ""),
                                "serie":         dados.get("serie", ""),
                                "data_emissao":  dados.get("data_emissao", ""),
                                "cnpj_emitente": dados.get("cnpj_emitente", ""),
                                "nome_emitente": dados.get("nome_emitente", ""),
                                "cnpj_dest_doc": dados.get("cnpj_dest_doc", ""),
                                "nome_dest_doc": dados.get("nome_dest_doc", ""),
                                "valor_total":   float(dados.get("valor_total", 0) or 0),
                                "nat_operacao":  dados.get("nat_operacao", ""),
                                "xml":           xml_ret,
                            })
                        if _passa_filtros(dados):
                            todos_docs.append(dados)
                            docs_empresa += 1
                        chaves_pendentes.remove(ch)
                        resumos_sem_xml -= 1
                    time.sleep(0.5)
                if recuperadas:
                    _log(f"  {recuperadas} XML(s) recuperado(s) após manifestação!")

        emitidas = sum(1 for d in todos_docs if d.get("cnpj_empresa") == cnpj and d.get("papel") == "Emitida")
        recebidas = sum(1 for d in todos_docs if d.get("cnpj_empresa") == cnpj and d.get("papel") == "Recebida")
        _log(f"  Subtotal {nome}: {docs_empresa} docs (emitidas={emitidas} recebidas={recebidas})")
        if resumos_sem_xml > 0:
            _log(f"  {resumos_sem_xml} resumo(s) sem XML completo (salvos com dados parciais no acervo)")
        _log(f"  Páginas processadas: {paginas}")

    # Persiste no acervo local (independente dos filtros de período/tipo do ZIP)
    if salvar_db and docs_para_db:
        from db.database import salvar_resultados_nfe
        salvar_resultados_nfe(docs_para_db, baixado_por="auto")
        _log(f"Acervo: {len(docs_para_db)} documento(s) persistido(s) no banco.")

    if salvar_db and chaves_retry and username:
        try:
            from db.database import enfileirar_xml
            for cnpj_r, chaves_r in chaves_retry.items():
                if chaves_r:
                    qtd = enfileirar_xml(username, cnpj_r, chaves_r)
                    if qtd:
                        _log(f"  {qtd} chave(s) enfileirada(s) para retry (worker horário)")
        except Exception:
            pass

    # O mesmo documento pode aparecer mais de uma vez no feed NSU da SEFAZ
    # (ex.: republicação, ou resumo + documento completo em NSUs distintos).
    # Deduplica por (cnpj_empresa, chave) antes de gerar o ZIP e o Excel,
    # mantendo a primeira ocorrência.
    vistos: set[tuple[str, str]] = set()
    docs_unicos: list[dict] = []
    duplicados = 0
    for d in todos_docs:
        chave = d.get("chave", "")
        # Sem chave (ex.: XML de evento sem infNFe) -- usa o NSU como
        # identidade, pra não tratar documentos distintos como duplicata.
        chave_key = (d.get("cnpj_empresa", ""), chave or f"_nsu_{d.get('nsu', '')}")
        if chave_key in vistos:
            duplicados += 1
            continue
        vistos.add(chave_key)
        docs_unicos.append(d)
    if duplicados:
        _log(f"  {duplicados} documento(s) duplicado(s) no feed NSU removido(s).")
    todos_docs = docs_unicos

    if not todos_docs:
        _log("Nenhum documento novo encontrado para os filtros aplicados.")
        return None, log

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:

        for doc in todos_docs:
            chave = doc.get("chave") or f"NSU_{doc.get('nsu','x')}"
            cnpj_e = doc.get("cnpj_empresa", "")
            nome_e = "".join(c for c in doc.get("nome_empresa", "")
                             if c.isalnum() or c in " _-")[:30].strip()
            pasta = f"{cnpj_e}_{nome_e}"

            if incluir_xml:
                zf.writestr(f"XMLs/{pasta}/{chave}.xml", doc["xml"])

            if incluir_pdf:
                try:
                    from brazilfiscalreport.danfe import Danfe
                    buf_pdf = io.BytesIO()
                    danfe = Danfe(xml=doc["xml"].encode("utf-8"))
                    danfe.output(buf_pdf)
                    zf.writestr(f"PDFs/{pasta}/{chave}.pdf", buf_pdf.getvalue())
                except Exception as e_pdf:
                    _log(f"  DANFE não gerado para {chave[:20]}...: {e_pdf}")

        if incluir_excel:
            _log("Gerando relatório Excel...")
            if progress_cb:
                progress_cb(0.9)
            try:
                xlsx_bytes = _gerar_excel(todos_docs)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                zf.writestr(f"relatorio_nfe_nfce_{ts}.xlsx", xlsx_bytes)
            except Exception as e:
                _log(f"Erro ao gerar Excel: {e}")

    if progress_cb:
        progress_cb(1.0)

    _log(f"Concluído. Total: {len(todos_docs)} documento(s).")
    return zip_buf.getvalue(), log


# ── Consulta avulsa por chave de acesso (Strategy 2) ─────────────────────────

def consultar_chave_avulsa(
    pfx_bytes: bytes,
    pfx_senha: str,
    cnpj: str,
    chave: str,
    ambiente: str = "1",
    uf: str = "CE",
) -> tuple[dict | None, str]:
    """
    Baixa uma NF-e/NFC-e específica pela chave de acesso (consChNFe).
    Usa contador separado do distNSU: até 20 consultas/hora.
    Retorna (dados_com_xml, mensagem_erro).
    """
    chave = "".join(c for c in chave if c.isdigit())
    if len(chave) != 44:
        return None, f"Chave inválida: deve ter 44 dígitos (informados {len(chave)})."

    cuf = UF_CODIGOS.get(uf, 23)
    sessao = _sessao(pfx_bytes, pfx_senha)
    env = _ENVELOPE_CHAVE.format(amb=ambiente, cuf=cuf, cnpj=cnpj, chave=chave)

    try:
        r = sessao.post(URL_SEFAZ, data=env.encode("utf-8"), headers=_HEADERS, timeout=60)
        r.raise_for_status()

        root = ET.fromstring(r.text)
        ns_nfe = "http://www.portalfiscal.inf.br/nfe"
        cstat_el  = root.find(f".//{{{ns_nfe}}}cStat")
        motivo_el = root.find(f".//{{{ns_nfe}}}xMotivo")
        cstat  = cstat_el.text.strip()  if cstat_el  is not None else "999"
        motivo = motivo_el.text.strip() if motivo_el is not None else "Sem descrição"

        if cstat == "138":
            pass  # sucesso
        elif cstat == "137":
            return None, (
                "NF-e não localizada para esse CNPJ via consulta direta (cStat=137).\n\n"
                "Isso ocorre quando a NF-e foi autorizada pelo SVRS e o vínculo destinatário "
                "ainda não foi indexado no Ambiente Nacional para consulta direta por chave.\n\n"
                "Use a aba Consulta em Lote quando o bloqueio liberar — "
                "o distNSU sincroniza a fila completa e traz essas notas automaticamente."
            )
        else:
            _MENSAGENS = {
                "215": "Falha no esquema XML da requisição.",
                "217": "NF-e não consta na base de dados do Ambiente Nacional.",
                "694": "NF-e autorizada pela SEFAZ da UF emitente — não disponível no Ambiente Nacional.",
                "218": "NF-e fora do intervalo disponível (máximo 90 dias).",
                "656": "Consumo indevido — aguarde 1 hora antes de tentar novamente.",
                "573": "Duplicidade de NF-e.",
            }
            detalhe = _MENSAGENS.get(cstat, "")
            msg = f"SEFAZ retornou cStat={cstat}: {motivo}"
            if detalhe:
                msg += f"\n\nDica: {detalhe}"
            return None, msg

        docs = _descompactar_docs(r.text)
        if not docs:
            return None, (
                "SEFAZ aceitou a consulta mas não retornou o XML (cStat=138 sem docZip).\n\n"
                "Tente aguardar alguns minutos e consultar novamente."
            )

        xml_conteudo = docs[0]["xml"]
        dados = _extrair_dados(xml_conteudo, cnpj)
        dados["xml"] = xml_conteudo
        return dados, ""

    except Exception as e:
        return None, f"Erro na comunicação com SEFAZ: {e}"


# ── SAE NFC-e (NFCeListagemChaves + NFCeDownloadXML) ─────────────────────────

# Estados com SAE disponível e suas URLs base
_SAE_URLS: dict[int, str] = {
    35: "https://nfce.fazenda.sp.gov.br/ws",       # SP — implementado desde fev/2026
    # SVRS states — aguardando implementação (CE=23, AM=13, GO=52, MS=50, MT=51, PR=41, RS=43)
    # Quando disponível, adicionar: 23: "https://nfce.svrs.rs.gov.br/ws", etc.
}

# UFs que usam SVRS (quando SVRS implementar SAE, basta adicionar a URL acima)
_SVRS_UFS = {13, 23, 41, 43, 50, 51, 52}

_SAE_NS_LISTAGEM  = "http://www.portalfiscal.inf.br/nfe/wsdl/NFCeListagemChaves"
_SAE_NS_DOWNLOAD  = "http://www.portalfiscal.inf.br/nfe/wsdl/NFCeDownloadXML"
_SAE_ACTION_LIST  = f"{_SAE_NS_LISTAGEM}/nfeDadosMsg"
_SAE_ACTION_DOWN  = f"{_SAE_NS_DOWNLOAD}/nfeDadosMsg"

_ENVELOPE_SAE_LIST = """<?xml version="1.0" encoding="UTF-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDadosMsg xmlns="{ns}">
      <nfceListagemChaves versao="1.00" xmlns="http://www.portalfiscal.inf.br/nfe">
        <tpAmb>{amb}</tpAmb>
        <dataHoraInicial>{dh_ini}</dataHoraInicial>
        <dataHoraFinal>{dh_fim}</dataHoraFinal>
      </nfceListagemChaves>
    </nfeDadosMsg>
  </soap12:Body>
</soap12:Envelope>"""

_ENVELOPE_SAE_DOWN = """<?xml version="1.0" encoding="UTF-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeDadosMsg xmlns="{ns}">
      <nfceDownloadXML versao="1.00" xmlns="http://www.portalfiscal.inf.br/nfe">
        <tpAmb>{amb}</tpAmb>
        <chNFCe>{chave}</chNFCe>
      </nfceDownloadXML>
    </nfeDadosMsg>
  </soap12:Body>
</soap12:Envelope>"""


def _sae_url_base(cuf: int) -> str | None:
    """Retorna URL base SAE para o cUF, ou None se não disponível."""
    return _SAE_URLS.get(cuf)


def _listar_chaves_sae(
    sessao: requests.Session,
    url_base: str,
    amb: str,
    dh_ini: str,
    dh_fim: str,
) -> tuple[list[str], str | None]:
    """
    Chama NFCeListagemChaves e retorna (lista_de_chaves, erro_ou_None).
    Lida com cStat=101 (lista incompleta) internamente via paginação por dhEmisUltNfce.
    """
    todas_chaves: list[str] = []
    dh_atual = dh_ini
    url = f"{url_base}/NFCeListagemChaves.asmx"
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": _SAE_ACTION_LIST,
    }

    for _ in range(20):  # máximo 20 páginas de 2000 = 40.000 NFC-e
        env = _ENVELOPE_SAE_LIST.format(
            ns=_SAE_NS_LISTAGEM,
            amb=amb,
            dh_ini=dh_atual,
            dh_fim=dh_fim,
        )
        try:
            r = sessao.post(url, data=env.encode("utf-8"), headers=headers, timeout=60)
            r.raise_for_status()
        except Exception as e:
            return todas_chaves, f"Erro de comunicação: {e}"

        try:
            root = ET.fromstring(r.text)
        except Exception as e:
            return todas_chaves, f"Resposta inválida: {e}"

        ns_nfe = "http://www.portalfiscal.inf.br/nfe"
        cstat_el  = root.find(f".//{{{ns_nfe}}}cStat")
        motivo_el = root.find(f".//{{{ns_nfe}}}xMotivo")
        cstat  = cstat_el.text.strip()  if cstat_el  is not None else "999"
        motivo = motivo_el.text.strip() if motivo_el is not None else ""

        if cstat not in ("100", "101", "107"):
            return todas_chaves, f"SEFAZ SAE cStat={cstat}: {motivo}"

        chaves = [el.text.strip() for el in root.findall(f".//{{{ns_nfe}}}chNFCe") if el.text]
        todas_chaves.extend(chaves)

        if cstat == "101":
            # Lista incompleta — avança usando dhEmisUltNfce
            ult_el = root.find(f".//{{{ns_nfe}}}dhEmisUltNfce")
            if ult_el is not None and ult_el.text:
                dh_atual = ult_el.text[:16]  # AAAA-MM-DDThh:mm
            else:
                break
        else:
            break  # cStat=100 (completo) ou 107 (sem registros)

    return todas_chaves, None


def _baixar_xml_sae(
    sessao: requests.Session,
    url_base: str,
    amb: str,
    chave: str,
) -> tuple[str | None, str | None]:
    """
    Chama NFCeDownloadXML para uma chave específica.
    Retorna (xml_str, erro_ou_None).
    """
    url = f"{url_base}/NFCeDownloadXML.asmx"
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": _SAE_ACTION_DOWN,
    }
    env = _ENVELOPE_SAE_DOWN.format(ns=_SAE_NS_DOWNLOAD, amb=amb, chave=chave)

    try:
        r = sessao.post(url, data=env.encode("utf-8"), headers=headers, timeout=60)
        r.raise_for_status()
    except Exception as e:
        return None, f"Erro: {e}"

    try:
        root = ET.fromstring(r.text)
    except Exception as e:
        return None, f"Resposta inválida: {e}"

    ns_nfe = "http://www.portalfiscal.inf.br/nfe"
    cstat_el = root.find(f".//{{{ns_nfe}}}cStat")
    cstat = cstat_el.text.strip() if cstat_el is not None else "999"

    if cstat != "200":
        motivo_el = root.find(f".//{{{ns_nfe}}}xMotivo")
        motivo = motivo_el.text.strip() if motivo_el is not None else ""
        return None, f"cStat={cstat}: {motivo}"

    # Extrai o XML da NFC-e do grupo proc/nfeProc/NFe
    nfe_el = root.find(f".//{{{ns_nfe}}}NFe")
    if nfe_el is not None:
        return ET.tostring(nfe_el, encoding="unicode"), None

    # Fallback: procura nfeProc inteiro
    nfe_proc_el = root.find(f".//{{{ns_nfe}}}nfeProc")
    if nfe_proc_el is not None:
        return ET.tostring(nfe_proc_el, encoding="unicode"), None

    return None, "XML da NFC-e não encontrado na resposta"


def sincronizar_nfce_sae(
    pfx_bytes: bytes,
    pfx_senha: str,
    cnpj: str,
    cuf: int,
    ambiente: str,
    data_ini: str,
    data_fim: str,
    log_cb: Callable[[str], None] | None = None,
) -> tuple[list[dict], str | None]:
    """
    Sincroniza NFC-e emitidas via SAE (NFCeListagemChaves + NFCeDownloadXML).
    Retorna (lista_de_docs_com_xml, erro_ou_None).

    data_ini/data_fim: strings no formato 'AAAA-MM-DD'.
    """
    url_base = _sae_url_base(cuf)
    if url_base is None:
        if cuf in _SVRS_UFS:
            return [], (
                f"SAE NFC-e ainda não disponível para estados SVRS (cUF={cuf}).\n"
                "A SEFAZ-SP implementou em fev/2026 — aguardando SVRS. "
                "Quando disponível, o sistema funcionará automaticamente."
            )
        return [], f"SAE NFC-e não disponível para cUF={cuf}."

    def _log(msg: str):
        if log_cb:
            log_cb(msg)

    sessao = _sessao(pfx_bytes, pfx_senha)
    dh_ini = f"{data_ini}T00:00"
    dh_fim = f"{data_fim}T23:59"

    _log(f"SAE NFC-e: listando chaves de {data_ini} a {data_fim}...")
    chaves, erro = _listar_chaves_sae(sessao, url_base, ambiente, dh_ini, dh_fim)
    if erro:
        return [], erro

    _log(f"SAE NFC-e: {len(chaves)} chave(s) encontrada(s). Baixando XMLs...")
    docs: list[dict] = []
    for i, chave in enumerate(chaves):
        _log(f"  [{i+1}/{len(chaves)}] Baixando {chave[:20]}...")
        xml_str, err_xml = _baixar_xml_sae(sessao, url_base, ambiente, chave)
        time.sleep(0.3)  # respeita rate limit por IP/min
        if err_xml:
            _log(f"  AVISO: {err_xml}")
            continue
        if xml_str:
            dados = _extrair_dados(xml_str, cnpj)
            dados["xml"] = xml_str
            dados["cnpj_empresa"] = cnpj
            docs.append(dados)

    _log(f"SAE NFC-e: {len(docs)} NFC-e obtida(s).")
    return docs, None


# ── NfeConsultaProtocolo via SVRS ─────────────────────────────────────────────
# Consulta a situação de uma NF-e/NFC-e diretamente no autorizador (SVRS),
# independente do Ambiente Nacional. Retorna protocolo, status e datas.

_SVRS_CONSULTA_NFE  = "https://nfe.svrs.rs.gov.br/ws/NfeConsulta/NfeConsulta4.asmx"
_SVRS_CONSULTA_NFCE = "https://nfce.svrs.rs.gov.br/ws/NfeConsulta/NfeConsulta4.asmx"

_NS_CONSULTA = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeConsultaProtocolo4"

_ENVELOPE_CONSULTA_PROTOCOLO = """<?xml version="1.0" encoding="UTF-8"?>
<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xmlns:xsd="http://www.w3.org/2001/XMLSchema"
    xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
  <soap12:Body>
    <nfeConsultaNF xmlns="{ns_wsdl}">
      <nfeDadosMsg>
        <consSitNFe versao="4.00" xmlns="http://www.portalfiscal.inf.br/nfe">
          <tpAmb>{amb}</tpAmb>
          <xServ>CONSULTAR</xServ>
          <chNFe>{chave}</chNFe>
        </consSitNFe>
      </nfeDadosMsg>
    </nfeConsultaNF>
  </soap12:Body>
</soap12:Envelope>"""

_CSTAT_SITUACAO = {
    "100": "Autorizada",
    "101": "Cancelada",
    "110": "Denegada (uso indevido)",
    "150": "Autorizada fora de prazo",
    "151": "Cancelada fora de prazo",
    "155": "Cancelada por substituição",
    "301": "Uso denegado — irregularidade fiscal emitente",
    "302": "Uso denegado — irregularidade fiscal destinatário",
    "303": "Uso denegado — destinatário não habilitado a operar na UF",
}


def consultar_protocolo_svrs(
    pfx_bytes: bytes,
    pfx_senha: str,
    chave: str,
    ambiente: str = "1",
) -> dict:
    """
    Consulta situação de NF-e/NFC-e diretamente no SVRS via NfeConsultaProtocolo.
    Identifica automaticamente se é NF-e (mod 55) ou NFC-e (mod 65) pela chave.

    Retorna dict com:
      ok: bool, cStat: str, xMotivo: str, situacao: str,
      nProt: str, dhRecbto: str, chave: str, modelo: str,
      erro: str (vazio se ok)
    """
    chave = "".join(c for c in chave if c.isdigit())
    if len(chave) != 44:
        return {"ok": False, "erro": f"Chave inválida: {len(chave)} dígitos (esperado 44)."}

    modelo_cod = chave[20:22]
    if modelo_cod == "65":
        url = _SVRS_CONSULTA_NFCE
        modelo = "NFC-e"
    else:
        url = _SVRS_CONSULTA_NFE
        modelo = "NF-e"

    sessao = _sessao(pfx_bytes, pfx_senha)
    env = _ENVELOPE_CONSULTA_PROTOCOLO.format(
        ns_wsdl=_NS_CONSULTA, amb=ambiente, chave=chave,
    )
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": f"{_NS_CONSULTA}/nfeConsultaNF",
    }

    try:
        r = sessao.post(url, data=env.encode("utf-8"), headers=headers, timeout=60)
        r.raise_for_status()
    except Exception as e:
        return {"ok": False, "erro": f"Erro de comunicação com SVRS: {e}"}

    try:
        root = ET.fromstring(r.text)
    except Exception as e:
        return {"ok": False, "erro": f"Resposta XML inválida: {e}"}

    ns_nfe = "http://www.portalfiscal.inf.br/nfe"

    def _find(tag):
        return root.find(f".//{{{ns_nfe}}}{tag}")

    cstat_el  = _find("cStat")
    motivo_el = _find("xMotivo")
    nprot_el  = _find("nProt")
    dhrec_el  = _find("dhRecbto")

    cstat  = cstat_el.text.strip()  if cstat_el  is not None and cstat_el.text  else "999"
    motivo = motivo_el.text.strip() if motivo_el is not None and motivo_el.text else "Sem descrição"
    nprot  = nprot_el.text.strip()  if nprot_el  is not None and nprot_el.text  else ""
    dhrec  = dhrec_el.text.strip()  if dhrec_el  is not None and dhrec_el.text  else ""

    situacao = _CSTAT_SITUACAO.get(cstat, "")
    ok = cstat in ("100", "150")

    # Extrai eventos (cancelamento, carta de correção, etc.)
    eventos = []
    for ev_el in root.findall(f".//{{{ns_nfe}}}procEventoNFe"):
        tp_ev = _find_in(ev_el, ns_nfe, "tpEvento")
        desc_ev = _find_in(ev_el, ns_nfe, "xEvento")
        dh_ev = _find_in(ev_el, ns_nfe, "dhRegEvento")
        nprot_ev = _find_in(ev_el, ns_nfe, "nProt")
        if tp_ev or desc_ev:
            eventos.append({
                "tipo": tp_ev, "descricao": desc_ev,
                "data": dh_ev, "protocolo": nprot_ev,
            })

    return {
        "ok": ok,
        "cStat": cstat,
        "xMotivo": motivo,
        "situacao": situacao,
        "nProt": nprot,
        "dhRecbto": dhrec,
        "chave": chave,
        "modelo": modelo,
        "eventos": eventos,
        "erro": "" if ok else f"cStat={cstat}: {motivo}",
    }


def _find_in(parent, ns: str, tag: str) -> str:
    el = parent.find(f".//{{{ns}}}{tag}")
    return el.text.strip() if el is not None and el.text else ""


def consultar_protocolo_lote(
    pfx_bytes: bytes,
    pfx_senha: str,
    chaves: list[str],
    ambiente: str = "1",
    log_cb: Callable[[str], None] | None = None,
) -> list[dict]:
    """Consulta múltiplas chaves via NfeConsultaProtocolo (uma por uma com delay)."""
    resultados = []
    for i, chave in enumerate(chaves):
        if log_cb:
            log_cb(f"[{i+1}/{len(chaves)}] Consultando {chave[:20]}...")
        res = consultar_protocolo_svrs(pfx_bytes, pfx_senha, chave, ambiente)
        resultados.append(res)
        if i < len(chaves) - 1:
            time.sleep(1)
    return resultados


# ── Manifestação do Destinatário (RecepcaoEvento4) ──────────────────────────
# Envia evento 210210 (Ciência da Operação) ao Ambiente Nacional para
# permitir o download do XML completo via consChNFe.

_URL_RECEPCAO_EVENTO = "https://www1.nfe.fazenda.gov.br/NFeRecepcaoEvento4/NFeRecepcaoEvento4.asmx"
_NS_RECEPCAO_EVENTO  = "http://www.portalfiscal.inf.br/nfe/wsdl/NFeRecepcaoEvento4"
_SOAP_ACTION_EVENTO  = f"{_NS_RECEPCAO_EVENTO}/nfeRecepcaoEvento"

_EVENTO_TIPOS = {
    "210200": "Confirmacao da Operacao",
    "210210": "Ciencia da Operacao",
    "210220": "Desconhecimento da Operacao",
    "210240": "Operacao nao Realizada",
}


def _assinar_evento_xml(evento_xml: bytes, pfx_bytes: bytes, pfx_senha: str) -> str:
    """Assina o elemento <evento> contendo <infEvento> com XMLDSig (enveloped)."""
    from lxml import etree
    from signxml import XMLSigner, SignatureMethod, DigestAlgorithm, CanonicalizationMethod
    from cryptography.hazmat.primitives.serialization import pkcs12

    senha = pfx_senha.encode() if isinstance(pfx_senha, str) else pfx_senha
    private_key, cert, chain = pkcs12.load_key_and_certificates(pfx_bytes, senha)

    doc = etree.fromstring(evento_xml)
    inf = doc.find("{http://www.portalfiscal.inf.br/nfe}infEvento")
    ref_uri = f"#{inf.attrib['Id']}" if inf is not None else ""

    signer = XMLSigner(
        signature_algorithm=SignatureMethod.RSA_SHA256,
        digest_algorithm=DigestAlgorithm.SHA256,
        c14n_algorithm=CanonicalizationMethod.EXCLUSIVE_XML_CANONICALIZATION_1_0,
    )
    signed = signer.sign(
        doc,
        key=private_key,
        cert=[cert] + (list(chain) if chain else []),
        reference_uri=ref_uri,
    )
    return etree.tostring(signed, encoding="unicode")


def _montar_evento(
    cnpj: str,
    chave: str,
    tp_evento: str,
    n_seq: int,
    ambiente: str,
) -> str:
    """Monta o XML <evento> (sem assinatura) para Manifestação do Destinatário."""
    from datetime import datetime, timezone, timedelta

    desc = _EVENTO_TIPOS.get(tp_evento, "Ciencia da Operacao")
    fuso = timezone(timedelta(hours=-3))
    dh = datetime.now(fuso).strftime("%Y-%m-%dT%H:%M:%S%z")
    dh = dh[:-2] + ":" + dh[-2:]

    id_evento = f"ID{tp_evento}{chave}{n_seq:02d}"

    det_extra = ""
    if tp_evento == "210240":
        det_extra = "<xJust>Operacao nao realizada conforme conferencia interna</xJust>"

    return (
        f'<evento versao="1.00" xmlns="http://www.portalfiscal.inf.br/nfe">'
        f'<infEvento Id="{id_evento}">'
        f"<cOrgao>91</cOrgao>"
        f"<tpAmb>{ambiente}</tpAmb>"
        f"<CNPJ>{cnpj}</CNPJ>"
        f"<chNFe>{chave}</chNFe>"
        f"<dhEvento>{dh}</dhEvento>"
        f"<tpEvento>{tp_evento}</tpEvento>"
        f"<nSeqEvento>{n_seq}</nSeqEvento>"
        f"<verEvento>1.00</verEvento>"
        f'<detEvento versao="1.00">'
        f"<descEvento>{desc}</descEvento>"
        f"{det_extra}"
        f"</detEvento>"
        f"</infEvento>"
        f"</evento>"
    )


def manifestar_ciencia(
    pfx_bytes: bytes,
    pfx_senha: str,
    cnpj: str,
    chave: str,
    ambiente: str = "1",
    tp_evento: str = "210210",
    n_seq: int = 1,
) -> dict:
    """
    Envia Manifestação do Destinatário (padrão: 210210 = Ciência da Operação).

    Retorna dict com:
      ok: bool, cStat: str, xMotivo: str, nProt: str, erro: str
    """
    chave = "".join(c for c in chave if c.isdigit())
    if len(chave) != 44:
        return {"ok": False, "erro": f"Chave inválida: {len(chave)} dígitos."}

    if chave[20:22] == "65":
        return {"ok": False, "erro": "NFC-e (mod 65) não aceita Manifestação do Destinatário."}

    evento_xml = _montar_evento(cnpj, chave, tp_evento, n_seq, ambiente)

    try:
        evento_assinado = _assinar_evento_xml(evento_xml.encode("utf-8"), pfx_bytes, pfx_senha)
    except Exception as e:
        return {"ok": False, "erro": f"Erro ao assinar evento: {e}"}

    envelope = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"'
        ' xmlns:xsd="http://www.w3.org/2001/XMLSchema"'
        ' xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">'
        "<soap12:Body>"
        f'<nfeRecepcaoEvento xmlns="{_NS_RECEPCAO_EVENTO}">'
        "<nfeDadosMsg>"
        f'<envEvento versao="1.00" xmlns="{NS}">'
        "<idLote>1</idLote>"
        f"{evento_assinado}"
        "</envEvento>"
        "</nfeDadosMsg>"
        "</nfeRecepcaoEvento>"
        "</soap12:Body>"
        "</soap12:Envelope>"
    )

    sessao = _sessao(pfx_bytes, pfx_senha)
    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": _SOAP_ACTION_EVENTO,
    }

    try:
        r = sessao.post(_URL_RECEPCAO_EVENTO, data=envelope.encode("utf-8"),
                        headers=headers, timeout=60)
        r.raise_for_status()
    except Exception as e:
        return {"ok": False, "erro": f"Erro de comunicação: {e}"}

    try:
        root = ET.fromstring(r.text)
    except Exception as e:
        return {"ok": False, "erro": f"Resposta inválida: {e}"}

    cstat_lote = root.find(f".//{{{NS}}}cStat")
    cstat_lote_v = cstat_lote.text.strip() if cstat_lote is not None and cstat_lote.text else "999"

    if cstat_lote_v not in ("128", "135", "136"):
        motivo_lote = root.find(f".//{{{NS}}}xMotivo")
        motivo_txt = motivo_lote.text.strip() if motivo_lote is not None and motivo_lote.text else ""
        return {"ok": False, "cStat": cstat_lote_v, "xMotivo": motivo_txt,
                "erro": f"Lote rejeitado: cStat={cstat_lote_v} — {motivo_txt}"}

    ret = root.find(f".//{{{NS}}}infEvento")
    if ret is None:
        return {"ok": False, "erro": "Resposta sem infEvento."}

    cstat_ev = ret.find(f"{{{NS}}}cStat")
    motivo_ev = ret.find(f"{{{NS}}}xMotivo")
    nprot_ev = ret.find(f"{{{NS}}}nProt")

    cstat_v = cstat_ev.text.strip() if cstat_ev is not None and cstat_ev.text else "999"
    motivo_v = motivo_ev.text.strip() if motivo_ev is not None and motivo_ev.text else ""
    nprot_v = nprot_ev.text.strip() if nprot_ev is not None and nprot_ev.text else ""

    ok = cstat_v in ("135", "136")
    return {
        "ok": ok,
        "cStat": cstat_v,
        "xMotivo": motivo_v,
        "nProt": nprot_v,
        "chave": chave,
        "tp_evento": tp_evento,
        "desc_evento": _EVENTO_TIPOS.get(tp_evento, ""),
        "erro": "" if ok else f"cStat={cstat_v}: {motivo_v}",
    }


def manifestar_lote(
    pfx_bytes: bytes,
    pfx_senha: str,
    cnpj: str,
    chaves: list[str],
    ambiente: str = "1",
    tp_evento: str = "210210",
    log_cb: Callable[[str], None] | None = None,
) -> list[dict]:
    """Manifesta Ciência da Operação em lote (uma por uma com delay)."""
    resultados = []
    nfe_chaves = [c for c in chaves if c[20:22] != "65"]
    if log_cb and len(nfe_chaves) < len(chaves):
        log_cb(f"  {len(chaves) - len(nfe_chaves)} NFC-e ignorada(s) (não aceita manifestação)")

    for i, chave in enumerate(nfe_chaves):
        if log_cb:
            log_cb(f"[{i+1}/{len(nfe_chaves)}] Manifestando {chave[:20]}...")
        res = manifestar_ciencia(pfx_bytes, pfx_senha, cnpj, chave, ambiente, tp_evento)
        resultados.append(res)
        if log_cb:
            status = "OK" if res["ok"] else res.get("erro", "?")
            log_cb(f"  → {status}")
        if i < len(nfe_chaves) - 1:
            time.sleep(1)
    return resultados


# ── Geração do Excel ──────────────────────────────────────────────────────────

def _gerar_excel(docs: list[dict]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    BOR = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    def cel(ws, r, c, v, bold=False, bg=None, fc="000000", align="left", fmt=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name="Arial", bold=bold, color=fc, size=10)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            cell.number_format = fmt
        cell.border = BOR
        return cell

    def fmt_cnpj(cnpj: str) -> str:
        c = "".join(d for d in (cnpj or "") if d.isdigit())
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else cnpj

    COLUNAS = [
        ("NSU", 15), ("Modelo", 10), ("Papel", 12), ("Chave", 48),
        ("Nº NF", 10), ("Série", 7), ("Data Emissão", 15),
        ("CNPJ Emitente", 20), ("Emitente", 40),
        ("CNPJ Destinatário", 20), ("Destinatário", 38),
        ("Nat. Operação", 28), ("CFOP", 8), ("Valor Total (R$)", 18),
    ]

    wb = openpyxl.Workbook()
    ws_res = wb.active
    ws_res.title = "RESUMO GERAL"

    # ── Resumo ──────────────────────────────────────────────────────────────
    empresas: dict[str, list] = {}
    for d in docs:
        key = d.get("cnpj_empresa", "")
        empresas.setdefault(key, []).append(d)

    res_cols = [("Empresa", 38), ("CNPJ", 20), ("Total Docs", 12),
                ("Emitidas", 12), ("Recebidas", 12), ("NFC-e", 10),
                ("Valor Total (R$)", 20)]
    ws_res.merge_cells(f"A1:{get_column_letter(len(res_cols))}1")
    t = ws_res["A1"]
    t.value = "RESUMO GERAL — NF-e e NFC-e (Emitidas e Recebidas)"
    t.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    t.fill = PatternFill("solid", start_color="0D2137")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 30
    ws_res.merge_cells(f"A2:{get_column_letter(len(res_cols))}2")
    sub = ws_res["A2"]
    sub.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    sub.font = Font(name="Arial", italic=True, color="555555", size=9)
    sub.alignment = Alignment(horizontal="center", vertical="center")

    for ci, (nm, larg) in enumerate(res_cols, 1):
        cel(ws_res, 3, ci, nm, bold=True, bg="1F4E79", fc="FFFFFF", align="center")
        ws_res.column_dimensions[get_column_letter(ci)].width = larg
    ws_res.row_dimensions[3].height = 20

    tot_q = tot_emit = tot_rec = tot_nfce = 0
    tot_val = 0.0
    for i, (cnpj_e, ddocs) in enumerate(empresas.items()):
        r = 4 + i
        bg = "D6E4F0" if i % 2 == 0 else "FFFFFF"
        vals = [d.get("valor_total", 0) for d in ddocs]
        emit = sum(1 for d in ddocs if d.get("papel") == "Emitida")
        rec = sum(1 for d in ddocs if d.get("papel") == "Recebida")
        nfce = sum(1 for d in ddocs if d.get("modelo") == "NFC-e")
        nome_e = ddocs[0].get("nome_empresa", cnpj_e) if ddocs else cnpj_e
        cel(ws_res, r, 1, nome_e, bg=bg)
        cel(ws_res, r, 2, fmt_cnpj(cnpj_e), bg=bg, align="center")
        cel(ws_res, r, 3, len(ddocs), bg=bg, align="center")
        cel(ws_res, r, 4, emit, bg=bg, align="center")
        cel(ws_res, r, 5, rec, bg=bg, align="center")
        cel(ws_res, r, 6, nfce, bg=bg, align="center")
        c7 = cel(ws_res, r, 7, sum(vals), bg=bg, align="right")
        c7.number_format = "#,##0.00"
        tot_q += len(ddocs); tot_emit += emit; tot_rec += rec
        tot_nfce += nfce; tot_val += sum(vals)
        ws_res.row_dimensions[r].height = 16

    rt = 4 + len(empresas)
    ws_res.merge_cells(f"A{rt}:B{rt}")
    lbl = ws_res[f"A{rt}"]
    lbl.value = "TOTAL GERAL"
    lbl.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    lbl.fill = PatternFill("solid", start_color="2E75B6")
    lbl.alignment = Alignment(horizontal="right", vertical="center")
    lbl.border = BOR
    for ci, v in [(3, tot_q), (4, tot_emit), (5, tot_rec), (6, tot_nfce)]:
        cel(ws_res, rt, ci, v, bold=True, bg="2E75B6", fc="FFFFFF", align="center")
    ctot = cel(ws_res, rt, 7, tot_val, bold=True, bg="2E75B6", fc="FFFFFF", align="right")
    ctot.number_format = "#,##0.00"
    ws_res.freeze_panes = "A4"

    # ── Aba por empresa ──────────────────────────────────────────────────────
    for cnpj_e, ddocs in empresas.items():
        nome_e = ddocs[0].get("nome_empresa", cnpj_e) if ddocs else cnpj_e
        titulo_aba = "".join(c for c in nome_e if c.isalnum() or c in " -_")[:28].strip()
        ws = wb.create_sheet(title=titulo_aba)

        emit = sum(1 for d in ddocs if d.get("papel") == "Emitida")
        rec = sum(1 for d in ddocs if d.get("papel") == "Recebida")
        nfce = sum(1 for d in ddocs if d.get("modelo") == "NFC-e")

        ws.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
        tt = ws["A1"]
        tt.value = f"NF-e / NFC-e — {nome_e} | CNPJ: {fmt_cnpj(cnpj_e)}"
        tt.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        tt.fill = PatternFill("solid", start_color="0D2137")
        tt.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f"A2:{get_column_letter(len(COLUNAS))}2")
        sb = ws["A2"]
        sb.value = (
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  "
            f"Total: {len(ddocs)}  |  Emitidas: {emit}  |  Recebidas: {rec}  |  NFC-e: {nfce}"
        )
        sb.font = Font(name="Arial", italic=True, color="555555", size=9)
        sb.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 16

        for ci, (nm, larg) in enumerate(COLUNAS, 1):
            cel(ws, 3, ci, nm, bold=True, bg="1F4E79", fc="FFFFFF", align="center")
            ws.column_dimensions[get_column_letter(ci)].width = larg
        ws.row_dimensions[3].height = 20

        LI = 4
        for i, doc in enumerate(ddocs):
            r = LI + i
            if doc.get("modelo") == "NFC-e":
                bg = "FFF2CC" if i % 2 == 0 else "FFFDE7"
            elif doc.get("papel") == "Emitida":
                bg = "E2EFDA" if i % 2 == 0 else "F0F7EB"
            else:
                bg = "D6E4F0" if i % 2 == 0 else "FFFFFF"
            ws.row_dimensions[r].height = 16

            cel(ws, r,  1, doc.get("nsu", ""),               bg=bg, align="center")
            cel(ws, r,  2, doc.get("modelo", "NF-e"),         bg=bg, align="center")
            cel(ws, r,  3, doc.get("papel", "Recebida"),      bg=bg, align="center")
            cel(ws, r,  4, doc.get("chave", ""),              bg=bg, align="center")
            cel(ws, r,  5, doc.get("numero", ""),             bg=bg, align="center")
            cel(ws, r,  6, doc.get("serie", ""),              bg=bg, align="center")
            cel(ws, r,  7, doc.get("data_emissao", ""),       bg=bg, align="center")
            cel(ws, r,  8, fmt_cnpj(doc.get("cnpj_emitente", "")), bg=bg, align="center")
            cel(ws, r,  9, doc.get("nome_emitente", ""),      bg=bg)
            cel(ws, r, 10, fmt_cnpj(doc.get("cnpj_dest_doc", "")), bg=bg, align="center")
            cel(ws, r, 11, doc.get("nome_dest_doc", ""),      bg=bg)
            cel(ws, r, 12, doc.get("nat_operacao", ""),       bg=bg)
            cel(ws, r, 13, doc.get("cfop", ""),               bg=bg, align="center")
            cv = cel(ws, r, 14, doc.get("valor_total", 0.0),  bg=bg, align="right")
            cv.number_format = "#,##0.00"

        if ddocs:
            rt2 = LI + len(ddocs)
            ultima = LI + len(ddocs) - 1
            col_v = get_column_letter(14)
            ws.merge_cells(f"A{rt2}:{get_column_letter(13)}{rt2}")
            lb2 = ws[f"A{rt2}"]
            lb2.value = f"TOTAL  ({len(ddocs)} docs  |  {emit} emitidas  |  {rec} recebidas)"
            lb2.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            lb2.fill = PatternFill("solid", start_color="2E75B6")
            lb2.alignment = Alignment(horizontal="right", vertical="center")
            lb2.border = BOR
            ctot2 = cel(ws, rt2, 14, f"=SUM({col_v}{LI}:{col_v}{ultima})",
                        bold=True, bg="2E75B6", fc="FFFFFF", align="right")
            ctot2.number_format = "#,##0.00"
            ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUNAS))}{ultima}"
        ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
