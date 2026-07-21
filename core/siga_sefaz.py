"""
core/siga_sefaz.py
Cliente para a API REST do SIGA (Painel Indicador) da SEFAZ-CE.
Aceita certificado como bytes (vindo do banco), sem gravar em disco.

Mapeamento de endpoints obtido do OpenAPI publico em
https://siga.sefaz.ce.gov.br/api/v3/api-docs (sem autenticacao).

Login (OIDC/Keycloak + mTLS via broker "certificado-digital") validado
contra o site real: sso.sefaz.ce.gov.br delega para um realm dedicado a
mTLS em cert-sso.sefaz.ce.gov.br, que e onde o certificado do cliente e
de fato exigido na camada TLS.
"""

import base64
import hashlib
import io
import os
import re
import time
import zipfile
from urllib.parse import urlencode, urlparse, parse_qs

import requests
from requests_pkcs12 import Pkcs12Adapter

SSO_REALM = "https://sso.sefaz.ce.gov.br/auth/realms/sefaz-ad-realm"
SSO_AUTH_ENDPOINT = f"{SSO_REALM}/protocol/openid-connect/auth"
SSO_TOKEN_ENDPOINT = f"{SSO_REALM}/protocol/openid-connect/token"
SSO_CLIENT_ID = "painelind-frontend"
SSO_REDIRECT_URI = "https://siga.sefaz.ce.gov.br/ui/"

# "Entrar com Certificado Digital" é um identity broker do Keycloak que
# delega para um realm dedicado a mTLS em outro host (cert-sso). O TLS
# client cert só é pedido nesse host -- daí a necessidade de montar o
# adapter também nele.
CERT_IDP_ALIAS = "certificado-digital"
CERT_SSO_HOST = "https://cert-sso.sefaz.ce.gov.br"

SIGA_API = "https://siga.sefaz.ce.gov.br/api"

# tipo -> (segmento do path, prefixo do path: "unidades" usa CNPJ completo,
# "empresas" usa CNPJ base de 8 digitos)
TIPOS_DOCUMENTO = {
    "NF_E":  ("nf-e", "unidades"),
    "NFC_E": ("nfc-e", "unidades"),
    "CT_E":  ("ct-e", "unidades"),
    "MDF_E": ("mdf-e", "unidades"),
    "CF_E":  ("cf-e", "unidades"),
    "NF3_E": ("nf3-e", "empresas"),
    "NFS_E": ("nfs-e", "empresas"),
}

STATUS_CONCLUIDO = "CONCLUIDO"
STATUS_ERRO = "ERRO"
STATUS_EM_ANDAMENTO = {"SOLICITADO", "PROCESSANDO", "GERANDO_DADOS"}


def _sessao(pfx_bytes: bytes, pfx_senha: str) -> requests.Session:
    s = requests.Session()
    adapter = Pkcs12Adapter(
        pkcs12_data=pfx_bytes,
        pkcs12_password=pfx_senha.encode() if isinstance(pfx_senha, str) else pfx_senha,
    )
    # mTLS certificate-bound tokens: o certificado precisa ir tanto no login
    # (nos 3 hosts do fluxo de broker) quanto em toda chamada subsequente à API.
    s.mount("https://sso.sefaz.ce.gov.br", adapter)
    s.mount(CERT_SSO_HOST, adapter)
    s.mount("https://siga.sefaz.ce.gov.br", adapter)
    return s


def _pkce_pair() -> tuple[str, str]:
    verifier = base64.urlsafe_b64encode(os.urandom(40)).rstrip(b"=").decode()
    challenge = base64.urlsafe_b64encode(
        hashlib.sha256(verifier.encode()).digest()
    ).rstrip(b"=").decode()
    return verifier, challenge


def login(sessao: requests.Session) -> dict:
    """
    Executa o fluxo Authorization Code + PKCE do Keycloak usando o
    certificado digital (mTLS), confirmado via captura de rede real:

    1. GET no endpoint de autorização do realm principal (sefaz-ad-realm)
       -> retorna a página de login (com AUTH_SESSION_ID em cookie) contendo
       um link para o broker "certificado-digital".
    2. GET nesse link de broker (/broker/certificado-digital/login) ->
       Keycloak redireciona (303) para o realm dedicado a mTLS, hospedado
       em outro host (cert-sso.sefaz.ce.gov.br, realm sefaz-x509-realm).
       É *nesse* host que o TLS pede o certificado do cliente.
    3. cert-sso valida o certificado e redireciona (302) de volta para o
       endpoint do broker no realm principal, que por sua vez redireciona
       para SSO_REDIRECT_URI com "#code=...&state=..." -- fluxo OIDC
       original concluído.

    Retorna o dict de resposta do token endpoint (access_token,
    refresh_token, expires_in, ...).
    """
    verifier, challenge = _pkce_pair()
    state = base64.urlsafe_b64encode(os.urandom(16)).rstrip(b"=").decode()
    nonce = base64.urlsafe_b64encode(os.urandom(16)).rstrip(b"=").decode()

    auth_params = {
        "client_id": SSO_CLIENT_ID,
        "redirect_uri": SSO_REDIRECT_URI,
        "state": state,
        "response_mode": "fragment",
        "response_type": "code",
        "scope": "openid",
        "nonce": nonce,
        "code_challenge": challenge,
        "code_challenge_method": "S256",
    }
    r = sessao.get(
        f"{SSO_AUTH_ENDPOINT}?{urlencode(auth_params)}",
        timeout=30,
        allow_redirects=True,
    )
    r.raise_for_status()

    m = re.search(
        rf'href="([^"]*broker/{CERT_IDP_ALIAS}/login[^"]*)"', r.text,
    )
    if not m:
        raise RuntimeError(
            "Link do broker 'certificado-digital' não encontrado na página "
            "de login. O layout da tela pode ter mudado -- reinspecionar "
            f"a página em {r.url}."
        )
    broker_url = m.group(1).replace("&amp;", "&")
    if broker_url.startswith("/"):
        broker_url = f"https://sso.sefaz.ce.gov.br{broker_url}"

    # Segue: broker/login -> (mTLS) cert-sso -> broker/endpoint -> SSO_REDIRECT_URI
    r2 = sessao.get(broker_url, timeout=30, allow_redirects=True)
    r2.raise_for_status()

    final_url = r2.url
    fragment = urlparse(final_url).fragment
    qs = parse_qs(fragment)
    code = qs.get("code", [None])[0]
    if not code:
        raise RuntimeError(
            "Não foi possível extrair o 'code' após o broker de certificado. "
            f"URL final: {final_url}."
        )

    token_params = {
        "grant_type": "authorization_code",
        "client_id": SSO_CLIENT_ID,
        "redirect_uri": SSO_REDIRECT_URI,
        "code": code,
        "code_verifier": verifier,
    }
    rt = sessao.post(SSO_TOKEN_ENDPOINT, data=token_params, timeout=30)
    rt.raise_for_status()
    return rt.json()


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def _kebab_para_camel(s: str) -> str:
    partes = s.split("-")
    return partes[0] + "".join(p.capitalize() for p in partes[1:])


def _filtro_bate(filtro: dict, params_enviados: dict) -> bool:
    """Compara o filtro de uma solicitação já existente com os params que
    acabamos de enviar. dat-referencia é comparado por prefixo, pois a API
    normaliza "2026-07" para "2026-07-01" ao salvar. Se a chave nem existir no
    filtro salvo (alguns tipos, como INDICADORES_MALHA, não ecoam certos
    params, ex. formato-arquivo), não conta como divergência -- só ignora."""
    filtro = filtro or {}
    for chave_kebab, valor_enviado in params_enviados.items():
        chave_camel = _kebab_para_camel(chave_kebab)
        if chave_camel not in filtro:
            continue
        valor_salvo = filtro.get(chave_camel)
        if chave_kebab == "dat-referencia":
            enviados = valor_enviado if isinstance(valor_enviado, list) else [valor_enviado]
            salvos = valor_salvo if isinstance(valor_salvo, list) else ([valor_salvo] if valor_salvo else [])
            if not all(any(str(s or "").startswith(str(e)) for s in salvos) for e in enviados):
                return False
            continue
        if str(valor_salvo) != str(valor_enviado):
            return False
    return True


def _solicitar(
    sessao: requests.Session, token: str, url: str, params: dict,
    cnpj: str, tipo: str, descricao: str,
) -> str:
    """POST genérico de solicitação de download assíncrona. Trata corpo vazio
    (comum: 200/202 sem JSON) e 409 (solicitação igual já existe) buscando o
    id em /v1/solicitacoes -- comparando os filtros de verdade, não só
    cnpj+tipo, para não reaproveitar uma solicitação com outro período/papel."""
    r = sessao.post(
        url, headers=_headers(token), params=params,
        json={"descricao": descricao}, timeout=30,
    )
    # 409 = já existe uma solicitação igual (mesmo tipo/cnpj/filtros) pendente
    # ou concluída -- não é erro, só reaproveita a existente.
    if r.status_code != 409:
        r.raise_for_status()
        try:
            body = r.json()
            solicitacao_id = body.get("data", {}).get("id") or body.get("id")
            if solicitacao_id:
                return solicitacao_id
        except ValueError:
            pass

    solicitacoes = listar_solicitacoes(sessao, token)
    candidatas = [
        s for s in solicitacoes
        if s.get("cnpj") == cnpj and s.get("tipo") == tipo
        and _filtro_bate(s.get("filtro"), params)
    ]
    if not candidatas:
        raise RuntimeError(
            f"Solicitação de {tipo} para {cnpj} criada (status {r.status_code}), "
            f"mas nenhuma em /v1/solicitacoes bate com os filtros enviados ({params})."
        )
    # listar_solicitacoes já vem ordenado por criacao desc -- pega a mais nova.
    return candidatas[0]["id"]


def solicitar_download(
    sessao: requests.Session,
    token: str,
    cnpj: str,
    tipo: str,
    formato_arquivo: str = "xlsx",
    **filtros,
) -> str:
    """
    Dispara a geração assíncrona de um relatório (NF-e, NFC-e, CT-e, MDF-e,
    CF-e, NF3-e ou NFS-e). Retorna o id da solicitação criada.

    `filtros` aceita os parâmetros de query documentados na API, por
    exemplo: dat_referencia=["2026-06"], tipo_operacao="ENTRADA",
    papel_operacao="DESTINATARIO", resultado_processamento="...".
    Use underscore; a função converte para o formato kebab-case da API.
    """
    if tipo not in TIPOS_DOCUMENTO:
        raise ValueError(f"tipo inválido: {tipo}. Opções: {list(TIPOS_DOCUMENTO)}")

    segmento, prefixo = TIPOS_DOCUMENTO[tipo]
    cnpj_path = cnpj if prefixo == "unidades" else cnpj[:8]
    url = f"{SIGA_API}/v1/{prefixo}/{cnpj_path}/documentos-fiscais/{segmento}/download"

    params = {"formato-arquivo": formato_arquivo}
    for chave, valor in filtros.items():
        params[chave.replace("_", "-")] = valor

    descricao = f"{tipo} {cnpj} {filtros.get('dat_referencia', '')}"
    return _solicitar(sessao, token, url, params, cnpj, tipo, descricao)


def listar_indicadores_malha(sessao: requests.Session, token: str, cnpj: str, dat_referencia: str | None = None) -> list[dict]:
    """Lista os indícios de irregularidade (malha fiscal) do CNPJ. Lista vazia = sem pendências."""
    params = {}
    if dat_referencia:
        params["dat-referencia"] = dat_referencia
    r = sessao.get(
        f"{SIGA_API}/v1/unidades/{cnpj}/malhas-fiscais/indicadores",
        headers=_headers(token), params=params, timeout=30,
    )
    if r.status_code == 404:
        return []  # CNPJ sem malha calculada -- equivale a "sem pendências"
    r.raise_for_status()
    return r.json().get("data", [])


def solicitar_download_indicadores(
    sessao: requests.Session, token: str, cnpj: str,
    formato_arquivo: str = "xlsx", **filtros,
) -> str:
    """Dispara a geração assíncrona do relatório de índices de malha fiscal (pendências)."""
    url = f"{SIGA_API}/v1/unidades/{cnpj}/malhas-fiscais/indicadores/download"
    params = {"formato-arquivo": formato_arquivo}
    for chave, valor in filtros.items():
        params[chave.replace("_", "-")] = valor
    descricao = f"INDICADORES_MALHA {cnpj} {filtros.get('dat_referencia', '')}"
    return _solicitar(sessao, token, url, params, cnpj, "INDICADORES_MALHA", descricao)


def listar_solicitacoes(sessao: requests.Session, token: str) -> list[dict]:
    r = sessao.get(
        f"{SIGA_API}/v1/solicitacoes",
        headers=_headers(token), params={"sort": "criacao,desc"}, timeout=30,
    )
    r.raise_for_status()
    return r.json().get("data", [])


def _desembrulhar_zip_aninhado(conteudo: bytes) -> bytes:
    """
    Alguns tipos (ex.: INDICADORES_MALHA) retornam um .zip contendo um único
    .xlsx dentro, em vez do .xlsx puro (que os demais tipos retornam). Um
    .xlsx normal também é um zip, então a distinção é: se o zip tem
    "[Content_Types].xml" na raiz, já é o próprio xlsx; senão, é um wrapper
    -- extrai o primeiro .xlsx interno.
    """
    try:
        zf = zipfile.ZipFile(io.BytesIO(conteudo))
    except zipfile.BadZipFile:
        return conteudo
    nomes = zf.namelist()
    if "[Content_Types].xml" in nomes:
        return conteudo
    interno = next((n for n in nomes if n.lower().endswith(".xlsx")), None)
    if interno:
        return zf.read(interno)
    return conteudo


def baixar_solicitacao(sessao: requests.Session, token: str, solicitacao_id: str) -> bytes:
    r = sessao.get(
        f"{SIGA_API}/v1/solicitacoes/{solicitacao_id}/download",
        headers=_headers(token), timeout=60,
    )
    r.raise_for_status()
    conteudo = _desembrulhar_zip_aninhado(r.content)
    # Um .xlsx (ou o .zip que o embrulha) sempre começa com a assinatura PK.
    # Se não bater, a solicitação encontrada não é o arquivo esperado --
    # melhor falhar alto do que salvar um arquivo quebrado como "sucesso".
    if conteudo[:2] != b"PK":
        raise RuntimeError(
            f"Conteúdo inesperado para a solicitação {solicitacao_id} "
            f"({len(conteudo)} bytes, não é um arquivo válido)."
        )
    return conteudo


def aguardar_e_baixar(
    sessao: requests.Session,
    token: str,
    solicitacao_id: str,
    timeout: int = 180,
    intervalo: int = 3,
) -> bytes:
    """Faz polling em /v1/solicitacoes até o status virar CONCLUIDO e baixa o arquivo."""
    decorrido = 0
    while decorrido < timeout:
        solicitacoes = listar_solicitacoes(sessao, token)
        alvo = next((s for s in solicitacoes if s.get("id") == solicitacao_id), None)
        if alvo:
            status = alvo.get("status")
            if status == STATUS_CONCLUIDO:
                return baixar_solicitacao(sessao, token, solicitacao_id)
            if status == STATUS_ERRO:
                raise RuntimeError(f"Solicitação {solicitacao_id} terminou com erro no SIGA.")
        time.sleep(intervalo)
        decorrido += intervalo
    raise TimeoutError(f"Solicitação {solicitacao_id} não concluiu em {timeout}s.")


def extrair_chaves_xlsx(conteudo_xlsx: bytes) -> list[str]:
    """
    Extrai as chaves de acesso de qualquer relatório do SIGA (índices de
    malha, NF-e/NFC-e emitidas/recebidas/canceladas etc.) -- procura em
    todas as abas exceto "Resumo" por uma coluna com "chave" no nome;
    abas sem essa coluna (ex.: saldo credor) são ignoradas.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_xlsx), data_only=True)
    chaves: set[str] = set()
    for ws in wb.worksheets:
        if ws.title == "Resumo":
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        cabecalho = [str(c or "") for c in rows[0]]
        idx_chave = next((i for i, c in enumerate(cabecalho) if "chave" in c.lower()), None)
        if idx_chave is None:
            continue
        for row in rows[1:]:
            if idx_chave < len(row) and row[idx_chave]:
                chaves.add(str(row[idx_chave]).strip())
    return sorted(chaves)


def _indice_coluna(cabecalho: list[str], *palavras_chave: str):
    cab_lower = [c.lower() for c in cabecalho]
    for palavra in palavras_chave:
        idx = next((i for i, c in enumerate(cab_lower) if palavra in c), None)
        if idx is not None:
            return idx
    return None


def parse_documentos_fiscais(conteudo_xlsx: bytes, tipo: str, aba: str, periodo: str) -> list[dict]:
    """
    Lê um relatório de documentos fiscais (NF-e/NFC-e emitidas/recebidas/
    canceladas) e retorna linhas prontas pra upsert_documentos_fiscais().
    Colunas variam entre NF-e (completo) e NFC-e (só chave/data/valor) --
    usa busca por palavra-chave no cabeçalho em vez de nome fixo.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_xlsx), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    cabecalho = [str(c or "") for c in rows[0]]
    idx_chave = _indice_coluna(cabecalho, "chave")
    if idx_chave is None:
        return []
    idx_numero = _indice_coluna(cabecalho, "número", "numero")
    idx_data   = _indice_coluna(cabecalho, "data")
    idx_valor  = _indice_coluna(cabecalho, "valor")
    idx_uf     = _indice_coluna(cabecalho, "uf")
    idx_cnpj   = _indice_coluna(cabecalho, "cnpj")
    idx_nome   = _indice_coluna(cabecalho, "razão social", "razao social")
    idx_ind    = _indice_coluna(cabecalho, "indicador")

    def _val(row, idx):
        return row[idx] if idx is not None and idx < len(row) and row[idx] is not None else ""

    docs = []
    for row in rows[1:]:
        chave = _val(row, idx_chave)
        if not chave:
            continue
        docs.append({
            "tipo": tipo, "aba": aba, "periodo": periodo,
            "chave": str(chave).strip(),
            "numero": str(_val(row, idx_numero)),
            "data_emissao": str(_val(row, idx_data)),
            "valor": float(_val(row, idx_valor) or 0),
            "uf": str(_val(row, idx_uf)),
            "contraparte_cnpj": str(_val(row, idx_cnpj)),
            "contraparte_nome": str(_val(row, idx_nome)),
            "situacao": str(_val(row, idx_ind)),
        })
    return docs


def parse_indicadores_malha(conteudo_xlsx: bytes) -> tuple[list[dict], dict[str, list[dict]]]:
    """
    Lê o relatório de índices de malha. Retorna (resumo, detalhes):
    - resumo: lista de {indicador, descricao, grupo_cfop, valor, unidade, qtd_indicios}
      a partir da aba "Resumo".
    - detalhes: {codigo_indicador: [ {chave, numero, data_emissao, valor,
      contraparte_cnpj}, ... ]} a partir das abas "IND {codigo}" que tiverem
      coluna de chave (indicadores sem chave, como saldo credor, ficam de fora).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo_xlsx), data_only=True)

    resumo = []
    if "Resumo" in wb.sheetnames:
        rows = list(wb["Resumo"].iter_rows(values_only=True))
        if rows:
            cabecalho = [str(c or "") for c in rows[0]]
            idx_ind   = _indice_coluna(cabecalho, "indicador")
            idx_desc  = _indice_coluna(cabecalho, "descrição", "descricao")
            idx_valor = _indice_coluna(cabecalho, "valor")
            idx_unid  = _indice_coluna(cabecalho, "unidade")
            idx_qtd   = _indice_coluna(cabecalho, "qtd", "quantidade")
            for row in rows[1:]:
                if idx_ind is None or idx_ind >= len(row) or row[idx_ind] is None:
                    continue
                indicador_val = row[idx_ind]
                indicador_str = str(int(indicador_val)) if isinstance(indicador_val, float) and indicador_val.is_integer() else str(indicador_val)
                resumo.append({
                    "indicador": indicador_str,
                    "descricao": str(row[idx_desc]) if idx_desc is not None and idx_desc < len(row) else "",
                    "grupo_cfop": "",
                    "valor": float(row[idx_valor] or 0) if idx_valor is not None and idx_valor < len(row) else 0,
                    "unidade": str(row[idx_unid]) if idx_unid is not None and idx_unid < len(row) else "",
                    "qtd_indicios": int(row[idx_qtd] or 0) if idx_qtd is not None and idx_qtd < len(row) else 0,
                })

    detalhes: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        if ws.title == "Resumo":
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        cabecalho = [str(c or "") for c in rows[0]]
        idx_chave = _indice_coluna(cabecalho, "chave")
        if idx_chave is None:
            continue
        idx_numero = _indice_coluna(cabecalho, "número", "numero")
        idx_data   = _indice_coluna(cabecalho, "data")
        idx_valor  = _indice_coluna(cabecalho, "valor")
        idx_cnpj   = _indice_coluna(cabecalho, "cnpj")

        codigo = ws.title.replace("IND ", "").strip() or ws.title
        itens = []
        for row in rows[1:]:
            if idx_chave >= len(row) or not row[idx_chave]:
                continue
            itens.append({
                "chave": str(row[idx_chave]).strip(),
                "numero": str(row[idx_numero]) if idx_numero is not None and idx_numero < len(row) and row[idx_numero] else "",
                "data_emissao": str(row[idx_data]) if idx_data is not None and idx_data < len(row) and row[idx_data] else "",
                "valor": float(row[idx_valor] or 0) if idx_valor is not None and idx_valor < len(row) else 0,
                "contraparte_cnpj": str(row[idx_cnpj]) if idx_cnpj is not None and idx_cnpj < len(row) and row[idx_cnpj] else "",
            })
        if itens:
            detalhes[codigo] = itens

    return resumo, detalhes


def persistir_relatorio(cnpj: str, nome_arquivo: str, tipo: str | None, conteudo: bytes, periodo: str):
    """
    Estrutura o conteúdo de um relatório já baixado (xlsx) e grava nas
    tabelas siga_documentos_fiscais / siga_indicadores_malha(_detalhe) --
    é a fonte que o Power BI consulta via Postgres. Chamar sempre depois de
    salvar o .xlsx em cache, tanto no scheduler quanto no "Atualizar agora".
    """
    from db.database import (
        upsert_documentos_fiscais, upsert_indicadores_malha, upsert_indicadores_malha_detalhe,
    )

    if nome_arquivo == "INDICADORES_MALHA_pendencias":
        resumo, detalhes = parse_indicadores_malha(conteudo)
        upsert_indicadores_malha(cnpj, resumo)
        for codigo, itens in detalhes.items():
            upsert_indicadores_malha_detalhe(cnpj, codigo, itens)
    else:
        # nome_arquivo é "{TIPO}_{aba}", ex.: "NF_E_emitidas"
        aba = nome_arquivo[len(tipo) + 1:] if tipo and nome_arquivo.startswith(tipo) else nome_arquivo
        docs = parse_documentos_fiscais(conteudo, tipo or "", aba, periodo)
        for d in docs:
            d["cnpj"] = cnpj
        upsert_documentos_fiscais(docs)
