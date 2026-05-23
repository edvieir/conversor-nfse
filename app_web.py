#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor NFSe — Interface Web  v1.3
Uso: streamlit run app_web.py
"""

import sys
import os
import io
import tempfile
import contextlib
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

for _mod in (
    "tkinter", "tkinter.simpledialog", "tkinter.filedialog",
    "tkinter.messagebox", "tkinter.ttk", "tkinter.font",
):
    sys.modules.setdefault(_mod, MagicMock())

import streamlit as st

st.set_page_config(
    page_title="Conversor NFS-e  |  ISS Fortaleza",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed",
)

import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth
import bcrypt

sys.path.insert(0, str(Path(__file__).parent))

_CONVERSOR_OK  = False
_CONVERSOR_ERR = ""
try:
    import nfse_xml_to_txt as C
    _CONVERSOR_OK = True
except Exception as _e:
    _CONVERSOR_ERR = str(_e)


# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ── RESET / BASE ── */
#MainMenu, footer, header { visibility: hidden; }
*, *::before, *::after { box-sizing: border-box; }

html, body, [data-testid="stAppViewContainer"] {
    background: #0A0E1A !important;
    font-family: 'Inter', -apple-system, sans-serif;
}
[data-testid="stHeader"] { background: transparent !important; height: 0 !important; }
[data-testid="stSidebar"] {
    background: #0D1117 !important;
    border-right: 1px solid #1C2333 !important;
}
.block-container { max-width: 740px; padding: 0 1.5rem 5rem; margin: 0 auto; }

body, p, div, span, li { color: #8B949E; }
h1,h2,h3,h4 { color: #E6EDF3 !important; font-weight: 700; }
label {
    color: #6E7681 !important; font-size: .75rem !important;
    font-weight: 600 !important; letter-spacing: .4px !important;
    text-transform: uppercase !important;
}

/* ── ANIMAÇÃO ── */
@keyframes slideUp {
    from { opacity: 0; transform: translateY(18px); }
    to   { opacity: 1; transform: translateY(0); }
}

/* ── TOPBAR (banner de topo decorativo) ── */
.topbar {
    background: linear-gradient(90deg, #0D1117 0%, #131A2E 50%, #0D1117 100%);
    border-bottom: 1px solid #1C2333;
    padding: 12px 20px 10px;
    margin: 0 -1.5rem .2rem;
    display: flex; align-items: center;
    gap: 12px;
}
.topbar-logo {
    width: 30px; height: 30px;
    background: linear-gradient(135deg, #2F6FEB 0%, #7B3FE4 100%);
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: .95rem;
    box-shadow: 0 0 16px rgba(47,111,235,.35);
    flex-shrink: 0;
}
.topbar-name {
    color: #E6EDF3; font-size: .95rem; font-weight: 700; letter-spacing: -.2px;
}
.topbar-divider {
    width: 1px; height: 14px; background: #1C2333;
}
.topbar-tag {
    color: #484F58; font-size: .68rem; font-weight: 500; letter-spacing: .2px;
}
.topbar-spacer { flex: 1; }

/* ── NAVBAR (barra de ações/usuário) ── */
.navbar {
    background: #0D1117;
    border-bottom: 1px solid #1C2333;
    padding: 8px 20px;
    margin: 0 -1.5rem 1.8rem;
    display: flex; align-items: center; gap: 10px;
}
.navbar-user {
    display: flex; align-items: center; gap: 8px;
}
.navbar-avatar {
    width: 26px; height: 26px;
    background: linear-gradient(135deg, #2F6FEB, #7B3FE4);
    border-radius: 6px;
    display: flex; align-items: center; justify-content: center;
    font-size: .72rem; font-weight: 800; color: #fff;
}
.navbar-name {
    color: #8B949E; font-size: .78rem; font-weight: 500;
}
.navbar-spacer { flex: 1; }
.navbar-btn {
    background: #161B27; border: 1px solid #2F3E55;
    border-radius: 7px; color: #8B949E;
    font-size: .74rem; font-weight: 600;
    padding: 5px 12px; cursor: pointer;
    text-decoration: none; display: inline-flex;
    align-items: center; gap: 5px;
    transition: all .2s;
}
.navbar-btn:hover { background: #1C2540; color: #E6EDF3; border-color: #2F6FEB; }

/* ── STEP CONTAINERS ── */
[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1px solid #1C2333 !important;
    border-radius: 14px !important;
    background: #0D1117 !important;
    padding: 6px 18px 18px !important;
    margin-bottom: 14px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,.4), 0 0 0 1px rgba(255,255,255,.02) !important;
    animation: slideUp .3s ease-out forwards !important;
    transition: border-color .2s, box-shadow .2s !important;
}
[data-testid="stVerticalBlockBorderWrapper"]:hover {
    border-color: #2F3E55 !important;
    box-shadow: 0 4px 20px rgba(0,0,0,.5), 0 0 0 1px rgba(255,255,255,.03) !important;
}

/* ── STEP HEADER ── */
.step-header {
    display: flex; align-items: center; gap: 12px;
    margin: 10px 0 16px;
    padding-bottom: 14px;
    border-bottom: 1px solid #1C2333;
}
.step-num {
    min-width: 26px; height: 26px;
    background: #161B27;
    border: 1px solid #2F3E55;
    color: #2F6FEB;
    border-radius: 7px;
    display: flex; align-items: center; justify-content: center;
    font-size: .78rem; font-weight: 800; flex-shrink: 0;
    letter-spacing: -.3px;
}
.step-title { color: #C9D1D9; font-weight: 600; font-size: .9rem; letter-spacing: -.1px; }
.step-desc  { color: #484F58; font-size: .75rem; margin-top: 2px; }

/* ── LOGIN FORM ── */
[data-testid="stForm"] {
    background: #0D1117 !important;
    border: 1px solid #1C2333 !important;
    border-radius: 16px !important;
    padding: 28px 28px !important;
    box-shadow: 0 24px 64px rgba(0,0,0,.6) !important;
}
/* Esconde o título "Login" / "Entrar" gerado pelo streamlit-authenticator */
[data-testid="stForm"] h1,
[data-testid="stForm"] h2,
[data-testid="stForm"] h3,
[data-testid="stForm"] [data-testid="stMarkdownContainer"] h1,
[data-testid="stForm"] [data-testid="stMarkdownContainer"] h2,
[data-testid="stForm"] [data-testid="stMarkdownContainer"] h3 {
    display: none !important;
}

div[data-testid="stFormSubmitButton"] > button,
[data-testid="stFormSubmitButton"] > button {
    background: linear-gradient(135deg, #2F6FEB, #1A56C8) !important;
    color: #fff !important; border: none !important;
    border-radius: 8px !important; width: 100% !important;
    height: 3em !important; font-weight: 700 !important;
    font-size: .92rem !important; letter-spacing: .1px !important;
    box-shadow: 0 2px 10px rgba(47,111,235,.3) !important;
    transition: all .2s ease !important; margin-top: .5rem !important;
}
div[data-testid="stFormSubmitButton"] > button:hover,
[data-testid="stFormSubmitButton"] > button:hover {
    background: linear-gradient(135deg, #4080F0, #2563EB) !important;
    box-shadow: 0 4px 18px rgba(47,111,235,.45) !important;
    transform: translateY(-1px) !important;
}

/* ── ADMIN HERO ── */
.admin-hero {
    background: #0D1117; border: 1px solid #1C2333;
    border-radius: 12px; padding: 24px 28px 20px; margin-bottom: 20px;
}
.admin-hero-title { color: #E6EDF3; font-size: 1.3rem; font-weight: 700; margin: 0 0 4px; }
.admin-hero-sub   { color: #484F58; font-size: .84rem; }

/* ── BADGES ── */
.badge {
    display: inline-block; background: #131A2E; color: #2F6FEB;
    border: 1px solid #1C3060; border-radius: 20px;
    font-size: .68rem; font-weight: 700; letter-spacing: .5px;
    padding: 3px 10px; margin-bottom: 10px; text-transform: uppercase;
}

/* ── FORMAT CARDS ── */
.format-card {
    background: #0A0E1A; border: 1px solid #1C2333;
    border-radius: 10px; padding: 16px 14px 12px; margin-bottom: 10px;
    text-align: center; transition: border-color .2s, box-shadow .2s;
}
.format-card:hover { border-color: #2F6FEB; box-shadow: 0 4px 14px rgba(47,111,235,.12); }
.format-icon { font-size: 1.7rem; margin-bottom: 7px; }
.format-name { color: #C9D1D9; font-weight: 700; font-size: .88rem; }
.format-desc { color: #484F58; font-size: .73rem; margin-top: 4px; line-height: 1.4; }

/* ── STATUS BOXES ── */
.info-box {
    background: #0B1525; border: 1px solid #1A2E50; border-left: 3px solid #2F6FEB;
    border-radius: 8px; padding: 11px 15px; margin: 8px 0;
    color: #5B8DD9; font-size: .8rem; line-height: 1.55;
}
.warn-box {
    background: #14100A; border: 1px solid #3A2800; border-left: 3px solid #C77D0A;
    border-radius: 8px; padding: 11px 15px; margin: 10px 0;
    color: #C77D0A; font-size: .8rem;
}
.error-box {
    background: #130B0B; border: 1px solid #3A1010; border-left: 3px solid #D93025;
    border-radius: 8px; padding: 11px 15px; margin: 10px 0;
    color: #D93025; font-size: .8rem;
}
.success-box {
    background: #071412; border: 1px solid #0D3027; border-left: 3px solid #1AB87A;
    border-radius: 8px; padding: 11px 15px; margin: 10px 0;
    color: #1AB87A; font-size: .8rem;
}

/* ── TABELA DE USUÁRIOS ── */
.user-table {
    width: 100%; border-collapse: collapse; background: #0D1117;
    border-radius: 10px; overflow: hidden; border: 1px solid #1C2333; margin: 10px 0 18px;
}
.user-table th {
    background: #0A0E1A; color: #484F58; font-size: .7rem;
    font-weight: 700; letter-spacing: .5px; text-transform: uppercase;
    padding: 10px 14px; border-bottom: 1px solid #1C2333; text-align: left;
}
.user-table td {
    padding: 10px 14px; border-bottom: 1px solid #1C2333; color: #C9D1D9; font-size: .82rem;
}
.user-table tr:last-child td { border-bottom: none; }
.user-table tr:hover td { background: #0A0E1A; }
.user-badge {
    display: inline-block; background: #0B1525; color: #2F6FEB;
    border: 1px solid #1A2E50; border-radius: 12px;
    font-size: .65rem; font-weight: 700; padding: 2px 8px;
}
.user-badge-admin { background: #14100A; color: #C77D0A; border-color: #3A2800; }

/* ── RESULTADO ── */
.result-success {
    background: #0D1117; border: 1px solid #1C2333;
    border-radius: 12px; padding: 18px 22px; margin: 14px 0 10px;
    display: flex; align-items: center; gap: 16px;
}
.result-success-icon  { font-size: 2rem; flex-shrink: 0; }
.result-success-title { color: #1AB87A; font-weight: 700; font-size: .95rem; }
.result-success-meta  { color: #484F58; font-size: .77rem; margin-top: 4px; }

/* ── BOTÕES GERAIS ── */
div[data-testid="stButton"] > button {
    background: #161B27 !important; color: #C9D1D9 !important;
    border: 1px solid #2F3E55 !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: .84rem !important;
    height: 2.7em !important; width: 100% !important;
    transition: all .2s ease !important;
}
div[data-testid="stButton"] > button:hover {
    background: #1C2540 !important; color: #E6EDF3 !important;
    border-color: #2F6FEB !important;
    box-shadow: 0 0 0 3px rgba(47,111,235,.15) !important;
}
div[data-testid="stButton"] > button:disabled {
    background: #0D1117 !important; color: #2F3E55 !important;
    border-color: #1C2333 !important; cursor: not-allowed !important;
    box-shadow: none !important; transform: none !important;
}

/* ── BOTÕES PRIMÁRIOS (Gerar TXT / XLSX) ── */
div[data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, #2F6FEB, #1A56C8) !important;
    color: #fff !important; border: none !important;
    font-weight: 700 !important; font-size: .9rem !important; height: 3em !important;
    box-shadow: 0 2px 8px rgba(47,111,235,.3) !important;
}
div[data-testid="stButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #4080F0, #2563EB) !important;
    box-shadow: 0 4px 16px rgba(47,111,235,.4) !important;
    transform: translateY(-1px) !important;
}

/* ── DOWNLOAD ── */
div[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #1AB87A, #0D9B64) !important;
    color: #fff !important; border: none !important; border-radius: 8px !important;
    font-weight: 700 !important; font-size: .9rem !important;
    height: 3em !important; width: 100% !important;
    box-shadow: 0 2px 8px rgba(26,184,122,.25) !important; transition: all .2s !important;
}
div[data-testid="stDownloadButton"] > button:hover {
    background: linear-gradient(135deg, #20D48D, #12AD70) !important;
    box-shadow: 0 4px 14px rgba(26,184,122,.4) !important;
    transform: translateY(-1px) !important;
}

/* ── INPUTS ── */
div[data-testid="stTextInput"] input {
    background: #0A0E1A !important; color: #E6EDF3 !important;
    border: 1px solid #2F3E55 !important; border-radius: 8px !important;
    font-size: .87rem !important; padding: 10px 14px !important;
    transition: border-color .2s, box-shadow .2s !important;
}
div[data-testid="stTextInput"] input:focus {
    border-color: #2F6FEB !important;
    box-shadow: 0 0 0 3px rgba(47,111,235,.18) !important;
    outline: none !important;
}
div[data-testid="stTextInput"] input::placeholder { color: #3A4455 !important; }

/* ── FILE UPLOADER ── */
div[data-testid="stFileUploader"] {
    background: #0A0E1A !important; border: 2px dashed #2F3E55 !important;
    border-radius: 12px !important; padding: 18px !important;
    transition: border-color .2s, box-shadow .2s !important;
}
div[data-testid="stFileUploader"]:hover {
    border-color: #2F6FEB !important;
    box-shadow: 0 0 0 3px rgba(47,111,235,.1) !important;
}
div[data-testid="stFileUploader"] * { color: #484F58 !important; }
div[data-testid="stFileUploader"] button {
    background: #2F3E55 !important; color: #C9D1D9 !important; border: none !important;
}

/* ── LOG / CODE ── */
[data-testid="stCode"], pre {
    background: #0A0E1A !important; border: 1px solid #1C2333 !important;
    border-radius: 8px !important; color: #6E7681 !important;
    font-size: .77rem !important; line-height: 1.65 !important;
}

/* ── EXPANDER ── */
[data-testid="stExpander"] {
    background: #0D1117 !important; border: 1px solid #1C2333 !important;
    border-radius: 8px !important;
}

/* ── SIDEBAR ── */
[data-testid="stSidebar"] { min-width: 220px !important; max-width: 260px !important; }
[data-testid="stSidebar"] * { color: #6E7681 !important; }
[data-testid="stSidebar"] strong { color: #E6EDF3 !important; }
.user-info-card {
    background: #161B27; border: 1px solid #1C2333;
    border-radius: 10px; padding: 14px 16px; margin-bottom: 12px;
}
.user-avatar {
    width: 36px; height: 36px;
    background: linear-gradient(135deg, #2F6FEB, #7B3FE4);
    border-radius: 8px; display: flex; align-items: center;
    justify-content: center; font-size: .9rem; font-weight: 800;
    color: #fff; margin-bottom: 10px;
}
.user-name  { color: #E6EDF3 !important; font-weight: 700; font-size: .88rem; }
.user-login { color: #484F58 !important; font-size: .73rem; margin-top: 2px; }

/* ── FOOTER ── */
.footer {
    text-align: center; color: #2F3E55; font-size: .7rem;
    margin-top: 3rem; padding-top: 1.2rem; border-top: 1px solid #161B27;
    letter-spacing: .2px;
}

/* ── HR / CAPTION / SELECT ── */
hr { border-color: #1C2333 !important; margin: 1rem 0 !important; }
[data-testid="stCaptionContainer"] { color: #6E7681 !important; font-size: .77rem !important; }
[data-testid="stSelectbox"] > div > div {
    background: #0A0E1A !important; border: 1px solid #2F3E55 !important;
    border-radius: 8px !important; color: #E6EDF3 !important;
}
</style>
""", unsafe_allow_html=True)


# ── AUTENTICAÇÃO ──────────────────────────────────────────────────────────────
_CFG_PATH = Path(__file__).parent / "config.yaml"


def _carregar_auth():
    if not _CFG_PATH.exists():
        st.error("Arquivo `config.yaml` não encontrado.\n\nExecute: `python gerar_senha.py`")
        st.stop()
    with open(_CFG_PATH) as f:
        cfg = yaml.load(f, Loader=SafeLoader)
    if not cfg.get("credentials", {}).get("usernames"):
        st.warning("Nenhum usuário cadastrado. Execute: `python gerar_senha.py`")
        st.stop()
    return stauth.Authenticate(
        cfg["credentials"],
        cfg["cookie"]["name"],
        cfg["cookie"]["key"],
        cfg["cookie"]["expiry_days"],
    )


auth = _carregar_auth()

if st.session_state.get("authentication_status") is not True:
    # Esconde sidebar na tela de login
    st.markdown("""
    <style>
        [data-testid="stSidebar"] { display: none !important; }
        [data-testid="stSidebarCollapsedControl"] { display: none !important; }
        .block-container { max-width: 480px !important; }
    </style>""", unsafe_allow_html=True)
    # Cabeçalho da tela de login
    st.markdown('<div style="height:48px;"></div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align:center; margin-bottom:2rem;">
        <div style="
            width:54px; height:54px;
            background:linear-gradient(135deg,#2F6FEB 0%,#7B3FE4 100%);
            border-radius:14px;
            display:inline-flex; align-items:center; justify-content:center;
            font-size:1.6rem; margin-bottom:1.4rem;
            box-shadow:0 8px 28px rgba(47,111,235,.35);
        ">📊</div>
        <div style="color:#E6EDF3; font-size:1.45rem; font-weight:800; letter-spacing:-.4px; line-height:1.2;">
            Conversor NFS-e
        </div>
        <div style="color:#484F58; font-size:.82rem; margin-top:.55rem; letter-spacing:.1px;">
            ISS Fortaleza &nbsp;·&nbsp; SPED GOV &nbsp;·&nbsp; Modelo Nacional 2026
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Formulário de login
    auth.login(fields={
        "Form name": "",
        "Username": "Usuário",
        "Password": "Senha",
        "Login": "Entrar",
    })

    _status = st.session_state.get("authentication_status")
    if _status is False:
        st.error("Usuário ou senha incorretos. Tente novamente.")
    if _status is not True:
        st.markdown("""
        <div style="text-align:center; color:#334155; font-size:.72rem; margin-top:1.2rem;">
            🔒 Acesso restrito · Entre com suas credenciais acima
        </div>""", unsafe_allow_html=True)
        st.stop()

_status = st.session_state.get("authentication_status")
if _status is not True:
    st.stop()


# ── SIDEBAR ───────────────────────────────────────────────────────────────────
_usuario_atual = st.session_state.get("username", "")
_nome_atual    = st.session_state.get("name", "Usuário")
_is_admin      = (_usuario_atual == "admin")

if "pagina" not in st.session_state:
    st.session_state.pagina = "conversor"

if not _CONVERSOR_OK and st.session_state.pagina == "conversor":
    st.error(f"Não foi possível carregar `nfse_xml_to_txt.py`:\n\n`{_CONVERSOR_ERR}`")
    st.stop()


# ── HELPER ────────────────────────────────────────────────────────────────────
def processar_uploads(uploaded_files, im: str, modo: str, competencia_filtro: str = ""):
    import xml.etree.ElementTree as _ET
    import unicodedata as _ud
    import re as _re

    def _comp_bytes(b):
        """Extrai competência MM/AAAA do conteúdo do XML."""
        try:
            root = _ET.fromstring(b)
            el = next((e for e in root.iter() if e.tag.endswith("dCompet")), None)
            if el is not None and el.text:
                p = el.text.strip()[:7].split("-")
                return f"{p[1]}/{p[0]}"
        except Exception:
            pass
        return ""

    # ── Pré-leitura: monta lookup de retenção a partir dos XMLs originais ──────
    # Necessário porque nfse_xml_to_txt.py tem dois bugs que precisamos corrigir:
    #   BUG 1 – Campo 21 (tpRetISSQN): o script passa o valor da NFS-e nacional
    #           diretamente, mas os significados são invertidos:
    #           NFS-e tpRetISSQN=2 = "retido pelo tomador"
    #           ISS Fortaleza campo 21=1 = "retido" (campo 21=2 = "a recolher")
    #           → notas com tpRetISSQN=2 ficam com campo 21=2 ("a recolher") — ERRADO.
    #   BUG 2 – Campos 39/40 (PIS/COFINS): o script busca tags 'vPIS'/'vCOFINS'
    #           (maiúsculas), mas o XML usa 'vPis'/'vCofins' (misto). Case-sensitive
    #           no ElementTree → valores nunca são encontrados → ficam vazios.
    _ret_lookup = {}   # {numero_nota: {tpRet, vPis, vCofins, vCSLL}}
    if modo == "txt":
        for _uf in uploaded_files:
            try:
                _uf.seek(0)
                _root = _ET.fromstring(_uf.read())
                # Identificadores da nota (campo 18 do TXT pode ser nDFSe ou nNFSe)
                _nDFSe = next((e.text or "" for e in _root.iter()
                               if e.tag.endswith("nDFSe")), "").strip()
                _nNFSe = next((e.text or "" for e in _root.iter()
                               if e.tag.endswith("nNFSe")), "").strip()
                # Tipo retenção ISS (1=não retido, 2=retido pelo tomador)
                _tpRet = next((e.text or "1" for e in _root.iter()
                               if e.tag.endswith("tpRetISSQN")), "1")
                # Tipo retenção PIS/COFINS (1=só PIS, 2=só COFINS, 3=ambos)
                _tpRetPC = next((e.text or "0" for e in _root.iter()
                                 if e.tag.endswith("tpRetPisCofins")), "0")
                # Valores de PIS, COFINS e CSLL
                # Busca case-insensitive (XML usa vPis/vCofins, script buscava vPIS/vCOFINS)
                _vPis    = next((e.text or "" for e in _root.iter()
                                 if e.tag.lower().endswith("vpis")), "")
                _vCofins = next((e.text or "" for e in _root.iter()
                                 if e.tag.lower().endswith("vcofins")), "")
                _vCSLL   = next((e.text or "" for e in _root.iter()
                                 if e.tag.lower().endswith("vretcsll")), "")
                # Alíquota ISS: busca pAliqAplic (nível NFSe) ou pAliq (nível DPS/tribMun)
                _aliq = (next((e.text or "" for e in _root.iter()
                                if e.tag.endswith("pAliqAplic")), "")
                         or next((e.text or "" for e in _root.iter()
                                  if e.tag.endswith("pAliq")), ""))
                # Valor ISS retido (vISSQN)
                _vISS = next((e.text or "" for e in _root.iter()
                              if e.tag.endswith("vISSQN")), "")
                _rinfo = {
                    "tpRet": _tpRet,
                    "aliq":  _aliq,
                    "vISS":  _vISS,
                    # PIS/COFINS não entram nos campos 39/40:
                    # quando são "Débito Apuração Própria" o emitente os recolhe
                    # por conta própria — não são retidos pelo tomador em Fortaleza.
                    # O campo "Contribuições Sociais Retidas" (vRetCSLL) já consolida
                    # o total das retenções federais e vai sozinho no campo 41.
                    "vCSLL": _vCSLL,
                }
                for _nk in [_nDFSe, _nNFSe]:
                    if _nk:
                        _ret_lookup[_nk] = _rinfo
            except Exception:
                pass
            finally:
                try:
                    _uf.seek(0)
                except Exception:
                    pass

    with tempfile.TemporaryDirectory() as tmp:
        ignorados = 0
        for uf in uploaded_files:
            uf.seek(0)
            content = uf.read()
            if competencia_filtro:
                comp = _comp_bytes(content)
                if comp and comp != competencia_filtro:
                    ignorados += 1
                    continue
            with open(os.path.join(tmp, uf.name), "wb") as fh:
                fh.write(content)
        ext   = "txt" if modo == "txt" else "xlsx"
        saida = os.path.join(tmp, f"resultado.{ext}")
        buf   = io.StringIO()
        if ignorados:
            buf.write(f"  FILTRO  {ignorados} arquivo(s) ignorado(s) — competência ≠ {competencia_filtro}\n\n")
        with contextlib.redirect_stdout(buf):
            try:
                if modo == "txt":
                    C.processar(tmp, saida, im_padrao=im)
                else:
                    C.processar_sped(tmp, saida, im_padrao=im)
            except Exception as exc:
                print(f"\nERRO FATAL: {exc}")
        log  = buf.getvalue()
        data = b""
        if os.path.exists(saida):
            with open(saida, "rb") as fh:
                data = fh.read()

            if modo == "txt" and data:
                def _fix_linha(linha: str) -> str:
                    cs = linha.split(";")
                    if len(cs) != 46:
                        return linha

                    # Campo 26 (índice 25) – descrição: remove acentos E pontuações
                    _desc = "".join(
                        c for c in _ud.normalize("NFD", cs[25]) if ord(c) < 128
                    )
                    _desc = _re.sub(r"[^A-Za-z0-9 ]", " ", _desc)
                    cs[25] = _re.sub(r" {2,}", " ", _desc).strip()

                    # Campo 30 (índice 29) – natureza: prestação em Fortaleza → "1"
                    if cs[28] == "2304400":
                        cs[29] = "1"

                    # Lookup pelo número da nota (campo 18 = índice 17)
                    _rinfo = _ret_lookup.get(cs[17].strip(), {})

                    # Campo 21 (índice 20) – tipo recolhimento
                    # tpRetISSQN=2 → ISS retido pelo tomador
                    # ISS Fortaleza: campo 21=1 significa "retido" (não "2"!)
                    if _rinfo.get("tpRet") == "2" and cs[20] == "2":
                        cs[20] = "1"

                    # Campo 25 (índice 24) – alíquota ISS em centésimos
                    # Fica logo após o CNAE. Só preenche quando ISS é retido (campo 21=1).
                    # Ex.: pAliqAplic="2.00" → campo 25="200"
                    if cs[20] == "1" and not cs[24].strip() and _rinfo.get("aliq"):
                        try:
                            cs[24] = str(int(round(float(_rinfo["aliq"]) * 100)))
                        except Exception:
                            pass

                    # Campo 31 (índice 30) – valor ISS retido em centavos
                    # Ex.: vISSQN="6.84" → campo 31="684"
                    if cs[20] == "1" and not cs[30].strip() and _rinfo.get("vISS"):
                        try:
                            cs[30] = str(int(round(float(_rinfo["vISS"]) * 100)))
                        except Exception:
                            pass

                    # Campo 41 (índice 40) – Contribuições Sociais retidas (CSLL)
                    # O XML traz vRetCSLL = total das contribuições sociais retidas
                    # pelo tomador (consolidado). Campos 39/40 (PIS/COFINS) ficam
                    # vazios pois são "Débito Apuração Própria" do emitente.
                    if _rinfo.get("vCSLL") and not cs[40].strip():
                        try:
                            cs[40] = str(int(round(float(_rinfo["vCSLL"]) * 100)))
                        except Exception:
                            pass

                    return ";".join(cs)

                texto  = data.decode("utf-8")
                linhas = texto.split("\n")
                data   = "\n".join(
                    _fix_linha(l) if ";" in l else l
                    for l in linhas
                ).encode("utf-8")
    return data, log


def _carregar_cfg() -> dict:
    with open(_CFG_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _salvar_cfg(cfg: dict):
    with open(_CFG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)


def _hash_senha(senha: str) -> str:
    return bcrypt.hashpw(senha.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")


def processar_xlsx_sped(uploaded_files, im: str, competencia_filtro: str = ""):
    """Gera XLSX no layout exato do SPED GOV — aba 'Serviços Tomados', 43 colunas."""
    import glob as _glob
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Cabeçalhos e larguras exatas do SPED GOV
    COLUNAS = [
        ("Tipo Doc.",                       24.14),
        ("Número",                          10.28),
        ("Código de Verificação",           23.57),
        ("Competência",                     14.57),
        ("Data",                            11.42),
        ("Vencimento",                      13.42),
        ("Número RPS",                      18.42),
        ("Série RPS",                       11.14),
        ("Tipo RPS",                        10.28),
        ("Natureza da Operação",            27.71),
        ("Regime Especial Tributação",      52.14),
        ("Operação Simples Nacional",       29.28),
        ("Incentivador Cultural",           22.85),
        ("Item da Lista",                   14.57),
        ("CNAE",                           123.14),
        ("ART",                              5.28),
        ("Código Obra",                     13.85),
        ("Número Empenho",                  19.42),
        ("Discriminação dos Serviços",     141.14),
        ("Valor dos Serviços",              20.28),
        ("Deduções Permitidas em Lei",      30.42),
        ("Desconto Condicionado",           25.28),
        ("Desconto Incondicionado",         27.00),
        ("Retenções Federais",              21.28),
        ("Outras Retenções",               19.42),
        ("PIS",                              4.42),
        ("COFINS",                           8.85),
        ("IRRF",                             5.71),
        ("CSLL",                             6.14),
        ("INSS",                             5.85),
        ("Base de Cálculo",                 17.28),
        ("Alíquota",                         9.85),
        ("Local da Prestação",              22.14),
        ("ISS Retido",                      11.71),
        ("Valor do ISS",                    13.85),
        ("Valor Líquido",                   14.85),
        ("Status Doc.",                     13.00),
        ("Inscrição Prestador",             21.14),
        ("CPF/CNPJ Prestador",              21.42),
        ("Razão Social/Nome do Prestador",  58.57),
        ("Escrituração",                    13.85),
        ("Origem",                           9.71),
        ("Status Aceite",                   14.85),
    ]

    IBGE_FORTALEZA = "2304400"

    def _float(v):
        try:
            return float(v) if v else 0.0
        except Exception:
            return 0.0

    def _str(v):
        return str(v).strip() if v else ""

    def _data_fmt(iso, fmt):
        """Converte ISO date para DD/MM/YYYY ou MM/YYYY."""
        if not iso:
            return ""
        data_part = iso[:10]  # YYYY-MM-DD
        try:
            partes = data_part.split("-")
            if fmt == "mes":
                return f"{partes[1]}/{partes[0]}"
            else:
                return f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception:
            return iso

    def _local_prestacao(d):
        # Tenta xLP (nome do município de prestação), depois lookup pelo código IBGE, depois xMun (emitente)
        uf   = _str(d.get("uf", ""))
        xLP  = _str(d.get("xLP", ""))
        cLP  = _str(d.get("cLP", ""))
        xMun = _str(d.get("xMun", ""))
        nome_mun = (
            xLP
            or getattr(C, "IBGE_TO_NOME", {}).get(cLP, "")
            or xMun
        )
        if nome_mun and uf:
            return f"{nome_mun.upper()} - {uf.upper()}"
        return nome_mun.upper() if nome_mun else ""

    def _cnae_desc(cnae9):
        desc = getattr(C, "CNAE9_TO_DESC", {}).get(cnae9, "")
        return f"{cnae9} - {desc}" if desc else cnae9

    def _extrair_fed(xml_path):
        """
        Extrai todas as retenções federais diretamente do XML.

        parse_nfse() tenta vPIS/vCOFINS como filhos diretos de tribFed,
        mas no modelo nacional eles ficam em piscofins/vPis e piscofins/vCofins.
        vRetIRRF e vRetCSLL também não são mapeados por parse_nfse().

        Retorna: (vPIS, vCOFINS, vIRRF, vCSLL, vINSS)
        """
        import xml.etree.ElementTree as ET

        def _v(root, *tags):
            """Busca o primeiro elemento cujo tag termina com qualquer um dos nomes."""
            for tag in tags:
                el = next((e for e in root.iter() if e.tag.endswith(tag)), None)
                if el is not None and el.text:
                    try:
                        return float(el.text.strip())
                    except (ValueError, AttributeError):
                        pass
            return 0.0

        try:
            root = ET.parse(xml_path).getroot()
            return (
                _v(root, "vPis",    "vPIS"),     # PIS   (piscofins/vPis ou tribFed/vPIS)
                _v(root, "vCofins", "vCOFINS"),  # COFINS
                _v(root, "vRetIRRF"),             # IRRF  (imposto de renda — coluna separada)
                _v(root, "vRetCSLL"),             # CSLL  (contribuição social)
                _v(root, "vINSS"),                # INSS  (se presente)
            )
        except Exception:
            return 0.0, 0.0, 0.0, 0.0, 0.0

    def _competencia_xml(xml_path):
        """Extrai competência MM/AAAA do campo dCompet do XML."""
        import xml.etree.ElementTree as ET
        try:
            root = ET.parse(xml_path).getroot()
            el = next((e for e in root.iter() if e.tag.endswith("dCompet")), None)
            if el is not None and el.text:
                p = el.text.strip()[:7].split("-")
                return f"{p[1]}/{p[0]}"
        except Exception:
            pass
        return ""

    with tempfile.TemporaryDirectory() as tmp:
        for uf_file in uploaded_files:
            uf_file.seek(0)
            with open(os.path.join(tmp, uf_file.name), "wb") as fh:
                fh.write(uf_file.read())

        arquivos = sorted(_glob.glob(os.path.join(tmp, "*.xml")))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Serviços Tomados"

        # Cabeçalho — bold, tamanho 11, fill indexado 55 (igual ao original)
        header_fill = PatternFill(fill_type="solid", fgColor="C0C0C0")  # cinza prata próximo ao original
        header_font = Font(bold=True, size=11)

        for col_idx, (titulo, largura) in enumerate(COLUNAS, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.font = header_fill and header_font
            cell.fill = header_fill
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = largura

        buf   = io.StringIO()
        total = 0
        erros = []

        with contextlib.redirect_stdout(buf):
            for xml_path in arquivos:
                nome_arq = os.path.basename(xml_path)
                try:
                    # Filtro de competência — ignora XMLs de outros meses
                    if competencia_filtro:
                        comp = _competencia_xml(xml_path)
                        if comp and comp != competencia_filtro:
                            print(f"  SKIP {nome_arq}: competência {comp} ≠ {competencia_filtro}")
                            continue

                    d = C.parse_nfse(xml_path)

                    if im:
                        d["im"] = im

                    cnae9, item, aliq_cnae = C.resolver_cnae9(d)
                    dhEmi = _str(d.get("dhEmi", ""))

                    # Fixo para SPED GOV — serviços tomados de outros municípios
                    tipo_doc = "NFS-e de Outro Município"
                    natureza = "Tributação Fora do Município"

                    vS   = _float(d.get("vS"))
                    vISS = _float(d.get("vISS"))
                    # parse_nfse() não extrai corretamente PIS/COFINS/IRRF/CSLL no modelo nacional
                    # (nomes e caminhos divergem do XML real) — usamos _extrair_fed() diretamente
                    vPIS, vCOFINS, vIRRF, vCSLL, vINSS = _extrair_fed(xml_path)
                    # Alíquota: usa o valor do XML; se vazio/zero, usa sugerido pela tabela CNAE
                    aliq   = _float(d.get("aliq")) or float(aliq_cnae or 0)
                    tpRet  = _str(d.get("tpRet", "1"))
                    iss_retido = (tpRet == "2")

                    # Contribuições Sociais Retidas = PIS + COFINS + CSLL
                    # IRRF é imposto de renda (não é contribuição social) — vai só na coluna AB
                    ret_federais = vPIS + vCOFINS + vCSLL

                    # Valor do ISS: só mostra se o tomador retém (tpRet=2), senão 0
                    valor_iss_col = vISS if iss_retido else 0

                    ws.append([
                        "NFS-e de Outro Município",        # A  Tipo Doc.        (FIXO)
                        _str(d.get("nDFSe")),             # B  Número           (XML)
                        None,                              # C  Cód. Verificação (FIXO vazio)
                        _data_fmt(dhEmi, "mes"),           # D  Competência      (XML MM/AAAA)
                        _data_fmt(dhEmi, "dia"),           # E  Data             (XML DD/MM/AAAA)
                        "",                                # F  Vencimento       (FIXO vazio)
                        "",                                # G  Número RPS       (FIXO vazio)
                        "",                                # H  Série RPS        (FIXO vazio)
                        "",                                # I  Tipo RPS         (FIXO vazio)
                        "Tributação Fora do Município",    # J  Natureza         (FIXO)
                        "",                                # K  Regime Especial  (FIXO vazio)
                        "Não",                             # L  Simples Nacional (FIXO)
                        None,                              # M  Incentivador     (FIXO vazio)
                        item,                              # N  Item da Lista    (XML)
                        _cnae_desc(cnae9),                 # O  CNAE             (XML cnae9+desc)
                        "",                                # P  ART              (FIXO vazio)
                        "",                                # Q  Código Obra      (FIXO vazio)
                        "",                                # R  Nº Empenho       (FIXO vazio)
                        _str(d.get("desc")),               # S  Discriminação    (XML)
                        vS,                                # T  Valor Serviços   (XML float)
                        "",                                # U  Deduções         (FIXO vazio)
                        "",                                # V  Desc. Condic.    (FIXO vazio)
                        "",                                # W  Desc. Incondic.  (FIXO vazio)
                        ret_federais,                      # X  Ret. Federais    (XML soma, 0 se nenhuma)
                        "",                                # Y  Outras Retenções (FIXO vazio)
                        vPIS    if vPIS    else "",        # Z  PIS              (XML, vazio se 0)
                        vCOFINS if vCOFINS else "",        # AA COFINS           (XML, vazio se 0)
                        vIRRF   if vIRRF   else "",        # AB IRRF             (XML vRetIRRF)
                        vCSLL   if vCSLL   else "",        # AC CSLL             (XML vRetCSLL)
                        vINSS   if vINSS   else "",        # AD INSS             (XML, vazio se 0)
                        vS,                                # AE Base de Cálculo  (XML = vS)
                        aliq,                              # AF Alíquota         (XML ou CNAE)
                        _local_prestacao(d),               # AG Local Prestação  (XML município+UF)
                        "Sim" if iss_retido else "Não",    # AH ISS Retido       (XML tpRet)
                        valor_iss_col,                     # AI Valor do ISS     (XML se retido, 0 se não)
                        vS,                                # AJ Valor Líquido    (XML = vS)
                        "NORMAL",                          # AK Status Doc.      (FIXO)
                        "",                                # AL Inscrição Prest. (FIXO vazio)
                        _str(d.get("cnpj")),               # AM CPF/CNPJ Prest.  (XML)
                        _str(d.get("nome")),               # AN Razão Social     (XML)
                        "Atual",                           # AO Escrituração     (FIXO)
                        "Prestador",                       # AP Origem           (FIXO)
                        "Não informada",                   # AQ Status Aceite    (FIXO)
                    ])

                    # Formatação numérica com 2 casas decimais nas colunas monetárias
                    r = ws.max_row
                    for col_mon in [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 35, 36]:
                        ws.cell(row=r, column=col_mon).number_format = '#,##0.00'
                    ws.cell(row=r, column=32).number_format = '0.00'  # Alíquota

                    total += 1
                    print(f"  OK   {nome_arq}")
                    print(f"       NFSe {d.get('nDFSe','')} | {d.get('nome','')[:35]}")
                except Exception as exc:
                    erros.append((nome_arq, str(exc)))
                    print(f"  ERRO {nome_arq}: {exc}")

        print(f"\n  Processadas: {total} nota(s)")
        if erros:
            print(f"  Com erro:    {len(erros)}")
            for n, e in erros:
                print(f"    - {n}: {e}")

        log   = buf.getvalue()
        saida = os.path.join(tmp, "resultado.xlsx")
        wb.save(saida)

        data = b""
        if os.path.exists(saida):
            with open(saida, "rb") as fh:
                data = fh.read()

    return data, log


# ── PÁGINA: GERENCIAR USUÁRIOS ─────────────────────────────────────────────────
def pagina_usuarios():
    if not _is_admin:
        st.error("Acesso restrito ao administrador.")
        return

    # ── Top bar ──
    inicial = _nome_atual[0].upper() if _nome_atual else "U"
    st.markdown("""
    <div class="topbar">
        <div class="topbar-logo">📊</div>
        <span class="topbar-name">Conversor NFS-e</span>
        <div class="topbar-divider"></div>
        <span class="topbar-tag">Painel Administrativo</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Navbar: usuário + botões ──
    _ua1, _ua2, _ua3 = st.columns([5, 1.8, 1.2])
    with _ua1:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:8px;padding:6px 0;">
            <div style="width:26px;height:26px;background:linear-gradient(135deg,#2F6FEB,#7B3FE4);
                border-radius:6px;display:flex;align-items:center;justify-content:center;
                font-size:.72rem;font-weight:800;color:#fff;">{inicial}</div>
            <span style="color:#8B949E;font-size:.78rem;font-weight:500;">{_nome_atual}</span>
        </div>""", unsafe_allow_html=True)
    with _ua2:
        if st.button("📄 Conversor", key="nav_conv_admin", use_container_width=True):
            st.session_state.pagina = "conversor"
            st.rerun()
    with _ua3:
        auth.logout("↩ Sair", key="logout_admin")

    st.markdown('<div style="height:.6rem;"></div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="admin-hero">
        <div class="admin-hero-title">👥 Gerenciar Usuários</div>
        <div class="admin-hero-sub">Adicione ou remova logins de acesso ao sistema</div>
    </div>
    """, unsafe_allow_html=True)

    cfg = _carregar_cfg()
    usuarios = cfg.get("credentials", {}).get("usernames", {})

    # ── Lista de usuários ──
    st.markdown("#### Usuários cadastrados")

    if not usuarios:
        st.markdown('<div class="warn-box">⚠️ Nenhum usuário cadastrado.</div>', unsafe_allow_html=True)
    else:
        linhas = ""
        for login, dados in usuarios.items():
            badge = '<span class="user-badge user-badge-admin">admin</span>' if login == "admin" else '<span class="user-badge">usuário</span>'
            nome_u = dados.get("name", "—")
            email_u = dados.get("email", "—")
            linhas += f"""
            <tr>
                <td>{badge} <strong style="color:#e6edf3">{login}</strong></td>
                <td>{nome_u}</td>
                <td>{email_u}</td>
            </tr>"""
        st.markdown(f"""
        <table class="user-table">
            <thead><tr>
                <th>Login</th>
                <th>Nome</th>
                <th>E-mail</th>
            </tr></thead>
            <tbody>{linhas}</tbody>
        </table>
        """, unsafe_allow_html=True)

    st.divider()

    # ── Adicionar usuário ──
    st.markdown("#### ➕ Adicionar novo usuário")

    with st.form("form_add_user", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            novo_login = st.text_input("Login *", placeholder="ex: joao")
            novo_nome  = st.text_input("Nome completo *", placeholder="ex: João Silva")
        with col2:
            novo_email = st.text_input("E-mail", placeholder="ex: joao@empresa.com")
            nova_senha = st.text_input("Senha *  (mín. 6 caracteres)", type="password")

        confirmar_senha = st.text_input("Confirmar senha *", type="password")

        submitted = st.form_submit_button("✅  Criar usuário", use_container_width=True, type="primary")

        if submitted:
            erros = []
            if not novo_login.strip():
                erros.append("Login é obrigatório.")
            elif " " in novo_login.strip():
                erros.append("Login não pode ter espaços.")
            if not novo_nome.strip():
                erros.append("Nome completo é obrigatório.")
            if len(nova_senha) < 6:
                erros.append("Senha deve ter pelo menos 6 caracteres.")
            if nova_senha != confirmar_senha:
                erros.append("As senhas não coincidem.")
            if novo_login.strip() in usuarios:
                erros.append(f"Já existe um usuário com o login '{novo_login.strip()}'.")

            if erros:
                for e in erros:
                    st.markdown(f'<div class="error-box">❌ {e}</div>', unsafe_allow_html=True)
            else:
                with st.spinner("Gerando hash da senha…"):
                    hashed = _hash_senha(nova_senha)

                cfg["credentials"]["usernames"][novo_login.strip()] = {
                    "name": novo_nome.strip(),
                    "email": novo_email.strip() or f"{novo_login.strip()}@exemplo.com",
                    "password": hashed,
                    "failed_login_attempts": 0,
                    "logged_in": False,
                }
                _salvar_cfg(cfg)
                st.markdown(f'<div class="success-box">✅ Usuário <strong>{novo_login.strip()}</strong> criado com sucesso!</div>', unsafe_allow_html=True)
                st.rerun()

    st.divider()

    # ── Remover usuário ──
    st.markdown("#### 🗑️ Remover usuário")

    opcoes_remover = [u for u in usuarios.keys() if u != _usuario_atual]

    if not opcoes_remover:
        st.markdown('<div class="info-box">💡 Não há outros usuários para remover.</div>', unsafe_allow_html=True)
    else:
        with st.form("form_remove_user"):
            login_remover = st.selectbox(
                "Selecione o usuário a remover",
                options=opcoes_remover,
                format_func=lambda u: f"{u}  —  {usuarios[u].get('name', '')}",
            )
            st.markdown('<div class="warn-box">⚠️ Esta ação é irreversível. O usuário perderá o acesso imediatamente.</div>', unsafe_allow_html=True)
            confirmar_remocao = st.form_submit_button("🗑️  Remover usuário", type="primary", use_container_width=True)

            if confirmar_remocao:
                del cfg["credentials"]["usernames"][login_remover]
                _salvar_cfg(cfg)
                st.markdown(f'<div class="success-box">✅ Usuário <strong>{login_remover}</strong> removido com sucesso.</div>', unsafe_allow_html=True)
                st.rerun()

    st.divider()

    # ── Salvar logins permanentemente ──
    st.markdown("#### 💾 Salvar logins permanentemente")
    st.markdown("""
    <div class="warn-box">
        ⚠️ <strong>Importante:</strong> Logins criados aqui ficam salvos enquanto o servidor estiver rodando.
        Para que não se percam após atualizações do sistema, baixe o arquivo abaixo e envie para o administrador técnico commitar no repositório.
    </div>
    """, unsafe_allow_html=True)

    with open(_CFG_PATH, "rb") as f_cfg:
        cfg_bytes = f_cfg.read()

    st.download_button(
        label="⬇  Baixar config.yaml atualizado",
        data=cfg_bytes,
        file_name="config.yaml",
        mime="text/plain",
        use_container_width=True,
    )

    st.markdown("""
    <div class="footer">
        Conversor NFS-e v1.4&nbsp;·&nbsp; Painel Administrativo
    </div>
    """, unsafe_allow_html=True)


# ── PÁGINA: CONVERSOR ─────────────────────────────────────────────────────────
def pagina_conversor():
    # ── Top bar ──
    st.markdown("""
    <div class="topbar">
        <div class="topbar-logo">📊</div>
        <span class="topbar-name">Conversor NFS-e</span>
        <div class="topbar-divider"></div>
        <span class="topbar-tag">ISS Fortaleza · SPED GOV · Modelo Nacional 2026</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Navbar: usuário + botões ──
    inicial = _nome_atual[0].upper() if _nome_atual else "U"
    if _is_admin:
        _nc1, _nc2, _nc3 = st.columns([5, 1.8, 1.2])
    else:
        _nc1, _nc3 = st.columns([7, 1.2])

    with _nc1:
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:8px;padding:6px 0;">
            <div style="width:26px;height:26px;background:linear-gradient(135deg,#2F6FEB,#7B3FE4);
                border-radius:6px;display:flex;align-items:center;justify-content:center;
                font-size:.72rem;font-weight:800;color:#fff;">{inicial}</div>
            <span style="color:#8B949E;font-size:.78rem;font-weight:500;">{_nome_atual}</span>
        </div>""", unsafe_allow_html=True)

    if _is_admin:
        with _nc2:
            if st.button("👥 Usuários", key="nav_usuarios_main", use_container_width=True):
                st.session_state.pagina = "usuarios"
                st.rerun()

    with _nc3:
        auth.logout("↩ Sair", key="logout_main")

    st.markdown('<div style="height:.6rem;"></div>', unsafe_allow_html=True)

    # Chave dinâmica para permitir limpeza total do uploader
    if "upload_key" not in st.session_state:
        st.session_state["upload_key"] = 0

    # ── ETAPA 1 ────────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("""
        <div class="step-header">
            <div class="step-num">1</div>
            <div class="step-info">
                <div class="step-title">Arquivos XML</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "xml_upload", type=["xml"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key=f"uploader_{st.session_state['upload_key']}",
        )

        # Toast dinâmico ao detectar novos arquivos
        n_atual = len(uploaded) if uploaded else 0
        if n_atual != st.session_state.get("_n_xml_prev", 0):
            if n_atual > 0:
                st.toast(
                    f"{n_atual} arquivo{'s' if n_atual > 1 else ''} carregado{'s' if n_atual > 1 else ''}",
                    icon="📂",
                )
            st.session_state["_n_xml_prev"] = n_atual

        if uploaded:
            total = len(uploaded)
            col_status, col_limpar = st.columns([5, 1])
            with col_status:
                st.markdown(
                    f'<div style="color:#10B981; font-size:.82rem; font-weight:600; padding:4px 0;">'
                    f'✅&nbsp; {total} arquivo{"s" if total > 1 else ""} selecionado{"s" if total > 1 else ""}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            with col_limpar:
                if st.button("✕ Limpar", key="btn_clear", use_container_width=True,
                             help="Remover todos os arquivos"):
                    st.session_state["upload_key"] += 1
                    st.session_state["_n_xml_prev"] = 0
                    st.rerun()

    # ── ETAPA 2 ────────────────────────────────────────────────────────────────
    with st.container(border=True):
        st.markdown("""
        <div class="step-header">
            <div class="step-num">2</div>
            <div class="step-info">
                <div class="step-title">Parâmetros</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        col_im, col_comp = st.columns(2, gap="medium")
        with col_im:
            st.markdown('<div style="color:#475569; font-size:.7rem; font-weight:600; letter-spacing:.4px; text-transform:uppercase; margin-bottom:3px;">Inscrição Municipal</div>', unsafe_allow_html=True)
            im_input = st.text_input("im", placeholder="Ex: 12345678-0", label_visibility="collapsed")
            if not im_input.strip():
                st.markdown(
                    '<div style="color:#C97400; font-size:.72rem; margin-top:3px;">'
                    '⚠️ Obrigatória para gerar TXT (ISS Fortaleza)'
                    '</div>',
                    unsafe_allow_html=True,
                )
        with col_comp:
            st.markdown('<div style="color:#475569; font-size:.7rem; font-weight:600; letter-spacing:.4px; text-transform:uppercase; margin-bottom:3px;">Competência</div>', unsafe_allow_html=True)
            comp_input = st.text_input("comp", placeholder="Ex: 05/2026  (opcional)", label_visibility="collapsed")

    # ── ETAPA 3 ────────────────────────────────────────────────────────────────
    tem_arquivos = bool(uploaded)

    with st.container(border=True):
        st.markdown("""
        <div class="step-header">
            <div class="step-num">3</div>
            <div class="step-info">
                <div class="step-title">Gerar arquivo</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if not tem_arquivos:
            st.markdown(
                '<div style="color:#475569; font-size:.8rem; margin-bottom:.6rem;">'
                '⚠️&nbsp; Faça o upload dos XMLs na etapa 1 para liberar a conversão.'
                '</div>',
                unsafe_allow_html=True,
            )

        col1, col2 = st.columns(2, gap="medium")

        with col1:
            st.markdown("""
            <div class="format-card">
                <div class="format-icon">📄</div>
                <div class="format-name">ISS Fortaleza</div>
            </div>
            """, unsafe_allow_html=True)
            btn_txt = st.button(
                "Gerar TXT",
                disabled=not tem_arquivos or not im_input.strip(),
                use_container_width=True,
                type="primary",
                key="btn_txt",
            )

        with col2:
            st.markdown("""
            <div class="format-card">
                <div class="format-icon">📊</div>
                <div class="format-name">SPED GOV</div>
            </div>
            """, unsafe_allow_html=True)
            btn_xlsx = st.button(
                "Gerar XLSX",
                disabled=not tem_arquivos,
                use_container_width=True,
                type="primary",
                key="btn_xlsx",
            )

    # ── Processamento ──────────────────────────────────────────────────────────
    # Normalizar filtro de competência (aceita MM/AAAA ou MM.AAAA)
    comp_filtro = ""
    if comp_input.strip():
        try:
            parts = comp_input.strip().replace(".", "/").split("/")
            if len(parts) == 2:
                mes, ano = int(parts[0]), int(parts[1])
                if 1 <= mes <= 12 and ano >= 2020:
                    comp_filtro = f"{mes:02d}/{ano}"
        except (ValueError, IndexError):
            pass
        if not comp_filtro:
            st.markdown('<div class="warn-box">⚠️ Competência inválida — use o formato MM/AAAA (ex: 05/2026).</div>', unsafe_allow_html=True)

    if btn_txt or btn_xlsx:
        if not uploaded:
            st.markdown('<div class="warn-box">⚠️ Selecione pelo menos um arquivo XML na Etapa 1.</div>', unsafe_allow_html=True)
        else:
            modo       = "txt" if btn_txt else "xlsx"
            tipo_label = "ISS Fortaleza (TXT)" if modo == "txt" else "SPED GOV (XLSX)"

            with st.spinner(f"⏳  Processando {len(uploaded)} arquivo(s) — {tipo_label}…"):
                if modo == "xlsx":
                    # IM é ignorada para SPED GOV — passa string vazia
                    dados_saida, log = processar_xlsx_sped(uploaded, "", comp_filtro)
                else:
                    # IM obrigatória para TXT (botão já fica desabilitado se vazia)
                    im = im_input.strip()
                    dados_saida, log = processar_uploads(uploaded, im, modo, comp_filtro)

            if dados_saida:
                ext        = "txt" if modo == "txt" else "xlsx"
                mime       = "text/plain" if modo == "txt" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                nome       = f"nfse_{modo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
                tamanho_kb = round(len(dados_saida) / 1024, 1)
                icone      = "📄" if modo == "txt" else "📊"

                st.toast(f"Arquivo {tipo_label} gerado com sucesso!", icon="✅")

                st.markdown(f"""
                <div class="result-success">
                    <div class="result-success-icon">✅</div>
                    <div class="result-success-body">
                        <div class="result-success-title">Arquivo gerado com sucesso!</div>
                        <div class="result-success-meta">
                            {icone} {nome} &nbsp;·&nbsp; {tamanho_kb} KB &nbsp;·&nbsp;
                            {len(uploaded)} XML{'s' if len(uploaded) > 1 else ''} processado{'s' if len(uploaded) > 1 else ''}
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.download_button(
                    label=f"⬇  Baixar  {nome}",
                    data=dados_saida, file_name=nome, mime=mime,
                    use_container_width=True,
                )
            else:
                st.markdown('<div class="error-box">❌ Nenhum arquivo foi gerado. Verifique o log abaixo.</div>', unsafe_allow_html=True)

            with st.expander("📋  Ver log de processamento"):
                st.code(log.strip() if log.strip() else "(nenhuma saída registrada)", language="")

    st.markdown("""
    <div class="footer">
        Conversor NFS-e &nbsp;v1.5 &nbsp;·&nbsp; Modelo Nacional 2026 &nbsp;·&nbsp;
        ISS Fortaleza &nbsp;/&nbsp; SPED GOV
    </div>
    """, unsafe_allow_html=True)


# ── ROTEADOR ─────────────────────────────────────────────────────────────────
pagina = st.session_state.get("pagina", "conversor")

if pagina == "usuarios":
    pagina_usuarios()
else:
    pagina_conversor()
